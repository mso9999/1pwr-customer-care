"""
Extract customer types (with HH1/HH2/HH3 granularity) from site survey master files.

Survey file structure (consistent across all sites):
  - Sheets named {SITE}_{NN}_{Village} contain per-household data
  - Header row: "Site number" in column B (found by scanning rows 1-25)
  - Data starts header_row + 2
  - Column B (2): site number
  - Column C (3): type (HH, HH1, HH2, HH3, SME, CHU, SCP, SCH, HC, HHSME, N/A)
  - Column F (6): score (1=Low/HH1, 2=Medium/HH2, 3=High/HH3)
  - Column G (7): latitude
  - Column H (8): longitude

HH subtype reconstruction:
  - If type is already HH1/HH2/HH3 -> use as-is
  - If type is "HH" and score exists -> map score to HH1/HH2/HH3
  - If type is "HH" and no score -> HH (unresolved)

Output: JSON array to /tmp/all_survey_types.json
"""
import openpyxl
import json
import sys
from collections import Counter

SCORE_TO_HH = {1: "HH1", 2: "HH2", 3: "HH3"}

SURVEY_FILES = {
    "KET": (
        "/Users/mattmso/Dropbox/0_0_LS/0_8 1PWR KET/(0) 1PWR KET WBS/"
        "(1) Community Survey, Outreach, Recruitment/1.2. KET/"
        "(1) Baseline Survey Data/(3) KET Site Survey/(4) KET Master Files/"
        "KET Master File 20220815.xlsx"
    ),
    "MAK": (
        "/Users/mattmso/Dropbox/0_0_LS/0_0 1PWR MAK/(0) 1PWR MAK WBS/"
        "(1) Community Survey, Outreach, Recruitment/1.2. MAK/"
        "Baseline Survey Data/(3) MAK Site Survey/MAK Master File/"
        "MAK Master File 20250512.xlsx"
    ),
    "MAS": (
        "/Users/mattmso/Dropbox/0_0_LS/0_6A 1PWR MAS/(0) 1PWR MAS WBS/"
        "(1) Community Survey, Outreach, Recruitment/1.2. MAS/"
        "(1) Baseline Survey Data/3) MAS Site Survey/(4) MAS Master file/"
        "MAS Site Form Master 20250612.xlsx"
    ),
    "MAT": (
        "/Users/mattmso/Dropbox/0_0_LS/0_2 1PWR MAT/(0) 1PWR MAT WBS/"
        "(1) Community Survey, Outreach, Recruitment/1.2. MAT/"
        "(1) Baseline Survey Data/(3) MAT Site Survey/(4) MAT Master file/"
        "MAT Master Excel file 20250606.xlsx"
    ),
    "SEH": (
        "/Users/mattmso/Dropbox/0_0_LS/0_4 1PWR SEH/(0) 1PWR SEH WBS/"
        "(1) Community Survey, Outreach, Recruitment/1.2. SEH/"
        "(1) Baseline Survey Data/(3) SEH Site Survey/(4) SEH Master Files/"
        "SEH Master File 20220727.xlsx"
    ),
    "SHG": (
        "/Users/mattmso/Dropbox/0_0_LS/0_6B 1PWR SHG/(0) 1PWR SHG WBS/"
        "(1) Community Survey, Outreach, Recruitment/1.2 SHG/"
        "(1) Baseline Survey Data/(3) SHG Site Survey/(4) SHG Master file/"
        "SHG Site Form Master 20250520.xlsx"
    ),
    "LEB": (
        "/Users/mattmso/Dropbox/0_0_LS/0_1 1PWR LEB/(0) 1PWR LEB WBS/"
        "(1) Community Survey, Outreach, Recruitment/1.2 LEB/"
        "(1) Baseline Survey Data/(3) LEB Site Survey/(4) LEB Master File/"
        "LEB Site Form Master 20220705.xlsx"
    ),
    "SEB": (
        "/Users/mattmso/Dropbox/0_0_LS/0_3A 1PWR SEB/(0) 1PWR SEB WBS/"
        "(1) Community Survey, Outreach, Recruitment/1.2 SEB/"
        "(1) Baseline Survey Data/(3) SEB Site Survey/(4) SEB Master files/"
        "SEB Survey Master File v2.0 20220502.xlsx"
    ),
    "TOS": (
        "/Users/mattmso/Dropbox/0_0_LS/0_3B 1PWR TOS/(0) 1PWR TOS WBS/"
        "(1) Community Survey, Outreach, Recruitment/1.2 TOS/"
        "(1) Baseline Survey Data/(3) TOS Site Survey/(4) TOS Master File/"
        "TOS Survey Master File 20220502.xlsx"
    ),
    "RIB": (
        "/Users/mattmso/Dropbox/0_0_LS/0_7 1PWR RIB/(0) 1PWR RIB WBS/"
        "(1) Community Survey, Outreach, Recruitment/1.2. RIB/"
        "(1) Baseline Survey Data/(3) RIB Site Survey/(4) RIB Master file/"
        "RIB Site Form Master 20220413 .xlsx"
    ),
    "RAL": (
        "/Users/mattmso/Dropbox/0_0_LS/0_9 1PWR RAL/(0) 1PWR RAL WBS/"
        "(1) Community Survey, Outreach, Recruitment/1.2. RAL/"
        "(1) Baseline Survey Data/(3) RAL Site Survey/(4) RAL Master file/"
        "RAL Site Form Master 20220106 14-56.xlsx"
    ),
    "TLH": (
        "/Users/mattmso/Dropbox/0_0_LS/0_5 1PWR TLH/(0) 1PWR TLH WBS/"
        "(1) Community Survey, Outreach, Recruitment/1.2 TLH/"
        "(1) Baseline Survey Data/(3) TLH Site Survey/(4) TLH Master file/"
        "TLH_Master_Site_Survey_(22_06_2022).xlsx"
    ),
    "SUA": (
        "/Users/mattmso/Dropbox/0_0_LS/0_11 1PWR SUA/(0) 1PWR SUA WBS/"
        "(1) Community Survey, Outreach, Recruitment/1.2 SUA/"
        "(1) Baseline Survey Data/(4) SUA Master file/"
        "LHDA_Ha_Suane_Master File.xlsx"
    ),
}


def extract_site(site_code: str, path: str) -> list:
    """Extract survey entries from a site master file."""
    import warnings
    warnings.filterwarnings("ignore", category=UserWarning)

    wb = openpyxl.load_workbook(path, data_only=True)
    entries = []

    village_sheets = [
        s for s in wb.sheetnames
        if s.startswith(f"{site_code}_")
        and "Summary" not in s
        and "Site_Number" not in s
        and "Sites_Summary" not in s
    ]

    # RAL has data in "Sheet1" instead of village-named sheets
    if not village_sheets:
        for s in wb.sheetnames:
            if s.lower() in ("sheet1", "sheet 1"):
                village_sheets = [s]
                break

    for sheet_name in village_sheets:
        ws = wb[sheet_name]
        header_row = None
        for r in range(1, 30):
            v = str(ws.cell(r, 2).value or "").strip().lower()
            if "site" in v and "number" in v:
                header_row = r
                break

        if not header_row:
            continue

        for r in range(header_row + 2, ws.max_row + 1):
            site_num = ws.cell(r, 2).value
            raw_type = str(ws.cell(r, 3).value or "").strip().upper()
            score_val = ws.cell(r, 6).value
            lat = ws.cell(r, 7).value
            lon = ws.cell(r, 8).value

            if site_num is None or not raw_type:
                continue
            try:
                site_num = int(float(str(site_num)))
            except (ValueError, TypeError):
                continue

            # Parse GPS (optional for type extraction but useful for validation)
            try:
                lat = float(lat) if lat is not None else None
                lon = float(lon) if lon is not None else None
            except (ValueError, TypeError):
                lat = lon = None

            # Parse score
            try:
                score = int(float(str(score_val))) if score_val is not None else None
            except (ValueError, TypeError):
                score = None

            # Resolve HH subtype
            if raw_type in ("HH1", "HH2", "HH3"):
                resolved_type = raw_type
            elif raw_type == "HH" and score in SCORE_TO_HH:
                resolved_type = SCORE_TO_HH[score]
            elif raw_type in ("SME", "CHU", "SCP", "SCH", "HC", "PWH", "GOV", "COM", "IND"):
                resolved_type = raw_type
            elif raw_type == "HHSME":
                resolved_type = "SME"
            elif raw_type in ("CLI", "CLINIC"):
                resolved_type = "HC"
            elif raw_type in ("CHRCH", "CHURCH"):
                resolved_type = "CHU"
            elif raw_type in ("N/A", "NA", ""):
                continue
            else:
                resolved_type = raw_type

            account = f"{site_num:04d}{site_code}"
            entries.append({
                "account": account,
                "site": site_code,
                "type": resolved_type,
                "type_raw": raw_type,
                "score": score,
                "lat": lat,
                "lon": lon,
                "village": sheet_name,
            })

    wb.close()
    return entries


def main():
    all_entries = []
    for site_code, path in sorted(SURVEY_FILES.items()):
        entries = extract_site(site_code, path)
        types = Counter(e["type"] for e in entries)
        print(f"{site_code}: {len(entries)} entries  {dict(types.most_common())}")
        all_entries.extend(entries)

    print(f"\n{'='*60}")
    print(f"Total: {len(all_entries)} entries across {len(SURVEY_FILES)} sites")
    total_types = Counter(e["type"] for e in all_entries)
    print(f"Type distribution: {dict(total_types.most_common())}")

    hh_total = sum(v for k, v in total_types.items() if k.startswith("HH"))
    hh_resolved = sum(v for k, v in total_types.items() if k in ("HH1", "HH2", "HH3"))
    hh_unresolved = total_types.get("HH", 0)
    print(f"\nHH breakdown: {hh_resolved} resolved ({hh_resolved*100/hh_total:.0f}%), "
          f"{hh_unresolved} unresolved ({hh_unresolved*100/hh_total:.0f}%)")

    with open("/tmp/all_survey_types.json", "w") as f:
        json.dump(all_entries, f)
    print(f"\nWrote {len(all_entries)} entries to /tmp/all_survey_types.json")


if __name__ == "__main__":
    main()
