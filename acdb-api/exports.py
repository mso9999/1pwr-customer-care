"""
Data export endpoints (CSV, XLSX).
"""

import csv
import io
import logging
from typing import Optional

from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse

from models import CurrentUser
from middleware import require_employee

logger = logging.getLogger("acdb-api.exports")

router = APIRouter(prefix="/api/export", tags=["export"])


def _get_connection():
    from customer_api import get_connection
    return get_connection()


def _row_to_dict(cursor, row) -> dict:
    columns = [desc[0] for desc in cursor.description]
    d = {}
    for col, val in zip(columns, row):
        if val is not None and not isinstance(val, (str, int, float, bool)):
            val = str(val)
        d[col] = val
    return d


@router.get("/{table_name}")
def export_table(
    table_name: str,
    format: str = Query("csv", regex="^(csv|xlsx)$"),
    filter_col: Optional[str] = Query(None),
    filter_val: Optional[str] = Query(None),
    search: Optional[str] = Query(None),
    user: CurrentUser = Depends(require_employee),
):
    """
    Export a table as CSV or XLSX.
    Optionally filter by column value or search text.
    """
    with _get_connection() as conn:
        cursor = conn.cursor()

        # Verify table exists via information_schema
        cursor.execute(
            "SELECT EXISTS (SELECT 1 FROM information_schema.tables "
            "WHERE table_schema = 'public' AND table_name = %s)",
            (table_name,),
        )
        found = cursor.fetchone()[0]
        if not found:
            raise HTTPException(status_code=404, detail=f"Table '{table_name}' not found")

        # Build query
        where_clauses = []
        params = []

        if filter_col and filter_val:
            where_clauses.append(f"{filter_col} = %s")
            params.append(filter_val)

        if search:
            # Find text columns via information_schema
            cursor.execute(
                "SELECT column_name FROM information_schema.columns "
                "WHERE table_schema = 'public' AND table_name = %s "
                "AND data_type IN ('character varying', 'text')",
                (table_name,),
            )
            text_cols = [row[0] for row in cursor.fetchall()]
            if text_cols:
                search_parts = [f"{c} LIKE %s" for c in text_cols[:10]]
                where_clauses.append(f"({' OR '.join(search_parts)})")
                params.extend([f"%{search}%"] * len(search_parts))

        where_sql = f" WHERE {' AND '.join(where_clauses)}" if where_clauses else ""
        sql = f"SELECT * FROM {table_name}{where_sql}"

        cursor.execute(sql, params)

        # Get column names
        columns = [desc[0] for desc in cursor.description]
        rows = cursor.fetchall()

    if format == "csv":
        return _export_csv(table_name, columns, rows)
    else:
        return _export_xlsx(table_name, columns, rows)


def _export_csv(table_name: str, columns: list, rows: list) -> StreamingResponse:
    """Generate CSV streaming response."""
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(columns)
    for row in rows:
        writer.writerow([str(v) if v is not None else "" for v in row])

    output.seek(0)
    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="text/csv",
        headers={"Content-Disposition": f"attachment; filename={table_name}.csv"},
    )


def _export_xlsx(table_name: str, columns: list, rows: list) -> StreamingResponse:
    """Generate XLSX streaming response."""
    try:
        from openpyxl import Workbook
    except ImportError:
        raise HTTPException(status_code=500, detail="openpyxl not installed; XLSX export unavailable")

    wb = Workbook()
    ws = wb.active
    ws.title = table_name[:31]  # Excel sheet name max 31 chars

    # Header row
    for col_idx, col_name in enumerate(columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = cell.font.copy(bold=True)

    # Data rows
    for row_idx, row in enumerate(rows, 2):
        for col_idx, val in enumerate(row, 1):
            if val is not None and not isinstance(val, (str, int, float, bool)):
                val = str(val)
            ws.cell(row=row_idx, column=col_idx, value=val)

    # Auto-width
    for col_idx, col_name in enumerate(columns, 1):
        max_len = len(str(col_name))
        for row_idx in range(2, min(len(rows) + 2, 102)):  # Sample first 100 rows
            cell_val = ws.cell(row=row_idx, column=col_idx).value
            if cell_val:
                max_len = max(max_len, len(str(cell_val)))
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = min(max_len + 2, 50)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={table_name}.xlsx"},
    )
