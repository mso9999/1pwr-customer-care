"""
Programs admin: funder program CRUD + bulk account tagging + Odyssey token
issuance.

Mounted at ``/api/admin/programs/...`` and restricted to **superadmin**. The
public read-only consumer of this data is :mod:`odyssey_api`.

Bulk-tag semantics
------------------

``POST /api/admin/programs/{code}/memberships/bulk`` accepts the *union* of:

* ``country_codes`` -- every account in any site of those countries
* ``site_codes``    -- every account in those sites
* ``account_numbers`` -- explicit list

Resolved to a set of account numbers, then ``add`` upserts membership rows
(idempotent) and ``remove`` deletes them. Returns counts so the UI can show
what actually changed.

Token issuance
--------------

Plaintext tokens are returned **once** at creation. Only the sha256 hash is
stored. ``token_prefix`` (first 8 chars of the plaintext) is kept so the UI
can identify a token in lists without revealing the secret.

The plaintext format is ``ody_<32 url-safe random chars>``.
"""

from __future__ import annotations

import hashlib
import io
import logging
import secrets
from datetime import datetime, timedelta, timezone
from typing import Any, Dict, List, Optional

from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from pydantic import BaseModel, Field

from customer_api import get_connection
from middleware import require_role
from models import CCRole, CurrentUser
from mutations import try_log_mutation

logger = logging.getLogger("cc-api.programs")

router = APIRouter(prefix="/api/admin/programs", tags=["programs"])

# Default token lifetime if the caller doesn't specify one. 90 days aligns
# with the documented JWT-secret rotation policy.
DEFAULT_TOKEN_LIFETIME_DAYS = 90


# ---------------------------------------------------------------------------
# Pydantic models
# ---------------------------------------------------------------------------

class ProgramCreate(BaseModel):
    code: str = Field(..., min_length=2, max_length=64)
    name: str = Field(..., min_length=1, max_length=200)
    funder: Optional[str] = Field(None, max_length=200)
    country_code: Optional[str] = Field(None, min_length=2, max_length=2)
    description: Optional[str] = None
    active: bool = True


class ProgramUpdate(BaseModel):
    name: Optional[str] = None
    funder: Optional[str] = None
    country_code: Optional[str] = Field(None, min_length=2, max_length=2)
    description: Optional[str] = None
    active: Optional[bool] = None


class ProgramOut(BaseModel):
    id: int
    code: str
    name: str
    funder: Optional[str] = None
    country_code: Optional[str] = None
    description: Optional[str] = None
    active: bool
    created_at: str
    member_count: int = 0
    active_token_count: int = 0


class BulkMembershipRequest(BaseModel):
    action: str = Field(..., description="'add' or 'remove'")
    country_codes: List[str] = Field(default_factory=list)
    site_codes: List[str] = Field(default_factory=list)
    account_numbers: List[str] = Field(default_factory=list)
    claim_milestone: Optional[str] = None
    notes: Optional[str] = None


class BulkMembershipResult(BaseModel):
    action: str
    requested_count: int
    affected_count: int
    skipped_unknown: List[str] = Field(default_factory=list)


class MembershipOut(BaseModel):
    account_number: str
    customer_id_legacy: Optional[str] = None
    customer_name: Optional[str] = None
    site_id: Optional[str] = None
    joined_at: str
    claim_milestone: Optional[str] = None
    notes: Optional[str] = None
    added_by: Optional[str] = None


class TokenIssueRequest(BaseModel):
    label: str = Field(..., min_length=1, max_length=120)
    lifetime_days: Optional[int] = Field(
        DEFAULT_TOKEN_LIFETIME_DAYS, ge=1, le=3650,
        description="Days until expiry. Pass null for a non-expiring token.",
    )


class TokenSummaryOut(BaseModel):
    id: int
    label: str
    token_prefix: str
    issued_at: str
    issued_by: Optional[str] = None
    expires_at: Optional[str] = None
    revoked_at: Optional[str] = None
    last_used_at: Optional[str] = None
    last_used_ip: Optional[str] = None


class TokenIssueResponse(BaseModel):
    token: str  # plaintext -- shown ONCE
    summary: TokenSummaryOut


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _iso(value: Any) -> Optional[str]:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.astimezone(timezone.utc).isoformat() if value.tzinfo else value.isoformat()
    return str(value)


def _resolve_program(cur, code: str) -> Dict[str, Any]:
    cur.execute(
        """
        SELECT id, code, name, funder, country_code, description, active, created_at
          FROM programs
         WHERE code = %s
         LIMIT 1
        """,
        (code,),
    )
    row = cur.fetchone()
    if not row:
        raise HTTPException(status_code=404, detail=f"Unknown program: {code}")
    return {
        "id": row[0],
        "code": row[1],
        "name": row[2],
        "funder": row[3],
        "country_code": row[4],
        "description": row[5],
        "active": row[6],
        "created_at": row[7],
    }


def _hash_token(plaintext: str) -> str:
    return hashlib.sha256(plaintext.encode("utf-8")).hexdigest()


def _new_token() -> str:
    """Return a fresh plaintext bearer token in the form ``ody_<32 chars>``."""
    return f"ody_{secrets.token_urlsafe(24)}"  # ~32 url-safe chars


def _sites_for_countries(country_codes: List[str]) -> List[str]:
    """Return the union of every site code registered to any of the given
    countries via ``country_config._REGISTRY``.
    """
    if not country_codes:
        return []
    from country_config import _REGISTRY  # type: ignore[attr-defined]

    out: List[str] = []
    for cc in country_codes:
        cfg = _REGISTRY.get(cc.strip().upper())
        if cfg:
            out.extend(cfg.site_abbrev.keys())
    return out


def _resolve_target_accounts(
    cur, country_codes: List[str], site_codes: List[str], account_numbers: List[str]
) -> Dict[str, Any]:
    """Resolve the union of (country -> site -> account, site -> account,
    explicit account_numbers) into a deduped set of valid account numbers.

    Returns ``{accounts: [...], skipped_unknown: [...]}``. Unknown account
    numbers (not in ``accounts`` table) are filtered out and reported back so
    the caller can warn the user.
    """
    sites = set(s.strip().upper() for s in site_codes if s and s.strip())
    sites.update(_sites_for_countries(country_codes))
    explicit = [a.strip() for a in account_numbers if a and a.strip()]

    found: set[str] = set()

    if sites:
        cur.execute(
            "SELECT account_number FROM accounts WHERE community = ANY(%s)",
            (list(sites),),
        )
        found.update(r[0] for r in cur.fetchall() if r[0])

    skipped: List[str] = []
    if explicit:
        cur.execute(
            "SELECT account_number FROM accounts WHERE account_number = ANY(%s)",
            (explicit,),
        )
        valid = {r[0] for r in cur.fetchall() if r[0]}
        for a in explicit:
            if a in valid:
                found.add(a)
            else:
                skipped.append(a)

    return {"accounts": sorted(found), "skipped_unknown": skipped}


# ---------------------------------------------------------------------------
# Programs CRUD
# ---------------------------------------------------------------------------

@router.get("", response_model=List[ProgramOut])
def list_programs(user: CurrentUser = Depends(require_role(CCRole.superadmin))):
    """List all funder programs, with member and token counts."""
    with get_connection() as conn:
        cur = conn.cursor()
        cur.execute(
            """
            SELECT p.id, p.code, p.name, p.funder, p.country_code, p.description,
                   p.active, p.created_at,
                   (SELECT COUNT(*) FROM program_memberships pm WHERE pm.program_id = p.id) AS members,
                   (SELECT COUNT(*) FROM odyssey_api_tokens t
                     WHERE t.program_id = p.id
                       AND t.revoked_at IS NULL
                       AND (t.expires_at IS NULL OR t.expires_at > NOW())) AS tokens
              FROM programs p
             ORDER BY p.created_at DESC, p.id DESC
            """
        )
        out: List[ProgramOut] = []
        for row in cur.fetchall():
            out.append(
                ProgramOut(
                    id=row[0],
                    code=row[1],
                    name=row[2],
                    funder=row[3],
                    country_code=row[4],
                    description=row[5],
                    active=row[6],
                    created_at=_iso(row[7]) or "",
                    member_count=int(row[8] or 0),
                    active_token_count=int(row[9] or 0),
                )
            )
        return out


@router.post("", response_model=ProgramOut, status_code=201)
def create_program(
    req: ProgramCreate,
    user: CurrentUser = Depends(require_role(CCRole.superadmin)),
):
    """Create a new funder program."""
    code = req.code.strip().upper()
    cc = (req.country_code or "").strip().upper() or None

    with get_connection() as conn:
        cur = conn.cursor()
        cur.execute("SELECT 1 FROM programs WHERE code = %s", (code,))
        if cur.fetchone():
            raise HTTPException(status_code=409, detail=f"Program {code} already exists.")
        cur.execute(
            """
            INSERT INTO programs (code, name, funder, country_code, description, active, created_by)
            VALUES (%s, %s, %s, %s, %s, %s, %s)
            RETURNING id, created_at
            """,
            (code, req.name, req.funder, cc, req.description, req.active, user.user_id),
        )
        new_id, created_at = cur.fetchone()
        conn.commit()

    try_log_mutation(
        user, "create", "programs", str(new_id),
        new_values={"code": code, "name": req.name, "funder": req.funder,
                    "country_code": cc, "active": req.active},
        metadata={"origin": "programs_admin"},
    )

    return ProgramOut(
        id=new_id, code=code, name=req.name, funder=req.funder, country_code=cc,
        description=req.description, active=req.active,
        created_at=_iso(created_at) or "",
        member_count=0, active_token_count=0,
    )


@router.patch("/{code}", response_model=ProgramOut)
def update_program(
    code: str,
    req: ProgramUpdate,
    user: CurrentUser = Depends(require_role(CCRole.superadmin)),
):
    """Update a program. Only fields explicitly set on the request are
    changed.
    """
    code = code.strip().upper()
    fields: Dict[str, Any] = {}
    if req.name is not None:
        fields["name"] = req.name
    if req.funder is not None:
        fields["funder"] = req.funder
    if req.country_code is not None:
        fields["country_code"] = req.country_code.strip().upper() or None
    if req.description is not None:
        fields["description"] = req.description
    if req.active is not None:
        fields["active"] = req.active

    if not fields:
        raise HTTPException(status_code=400, detail="No fields provided.")

    set_sql = ", ".join(f"{k} = %s" for k in fields)
    params = list(fields.values()) + [code]

    with get_connection() as conn:
        cur = conn.cursor()
        prog = _resolve_program(cur, code)
        cur.execute(
            f"UPDATE programs SET {set_sql}, updated_at = NOW() WHERE code = %s",
            params,
        )
        conn.commit()

    try_log_mutation(
        user, "update", "programs", str(prog["id"]),
        new_values=fields, metadata={"origin": "programs_admin"},
    )

    with get_connection() as conn:
        cur = conn.cursor()
        prog = _resolve_program(cur, code)
        cur.execute(
            "SELECT COUNT(*) FROM program_memberships WHERE program_id = %s",
            (prog["id"],),
        )
        members = int(cur.fetchone()[0] or 0)
        cur.execute(
            """
            SELECT COUNT(*) FROM odyssey_api_tokens
             WHERE program_id = %s AND revoked_at IS NULL
               AND (expires_at IS NULL OR expires_at > NOW())
            """,
            (prog["id"],),
        )
        tokens = int(cur.fetchone()[0] or 0)

    return ProgramOut(
        id=prog["id"], code=prog["code"], name=prog["name"], funder=prog["funder"],
        country_code=prog["country_code"], description=prog["description"],
        active=prog["active"], created_at=_iso(prog["created_at"]) or "",
        member_count=members, active_token_count=tokens,
    )


# ---------------------------------------------------------------------------
# Memberships -- bulk tagging + listing
# ---------------------------------------------------------------------------

@router.post("/{code}/memberships/bulk", response_model=BulkMembershipResult)
def bulk_membership(
    code: str,
    req: BulkMembershipRequest,
    user: CurrentUser = Depends(require_role(CCRole.superadmin)),
):
    """Add or remove the union of (country_codes, site_codes, account_numbers)
    to/from the program. Idempotent (``add`` upserts; ``remove`` deletes).
    """
    code = code.strip().upper()
    action = (req.action or "").strip().lower()
    if action not in ("add", "remove"):
        raise HTTPException(status_code=400, detail="action must be 'add' or 'remove'.")

    with get_connection() as conn:
        cur = conn.cursor()
        prog = _resolve_program(cur, code)
        resolved = _resolve_target_accounts(
            cur, req.country_codes, req.site_codes, req.account_numbers
        )
        accounts = resolved["accounts"]
        skipped = resolved["skipped_unknown"]

        if not accounts:
            return BulkMembershipResult(
                action=action,
                requested_count=len(req.account_numbers) + len(req.site_codes) + len(req.country_codes),
                affected_count=0,
                skipped_unknown=skipped,
            )

        if action == "add":
            cur.executemany(
                """
                INSERT INTO program_memberships
                    (program_id, account_number, claim_milestone, notes, added_by)
                VALUES (%s, %s, %s, %s, %s)
                ON CONFLICT (program_id, account_number) DO UPDATE
                   SET claim_milestone = COALESCE(EXCLUDED.claim_milestone, program_memberships.claim_milestone),
                       notes           = COALESCE(EXCLUDED.notes,           program_memberships.notes)
                """,
                [
                    (prog["id"], a, req.claim_milestone, req.notes, user.user_id)
                    for a in accounts
                ],
            )
            affected = cur.rowcount if cur.rowcount >= 0 else len(accounts)
        else:  # remove
            cur.execute(
                "DELETE FROM program_memberships "
                " WHERE program_id = %s AND account_number = ANY(%s)",
                (prog["id"], accounts),
            )
            affected = cur.rowcount if cur.rowcount >= 0 else 0
        conn.commit()

    try_log_mutation(
        user, action, "program_memberships", code,
        new_values={"accounts": accounts[:200], "claim_milestone": req.claim_milestone},
        metadata={"origin": "programs_bulk_tag", "site_codes": req.site_codes,
                  "country_codes": req.country_codes,
                  "explicit_count": len(req.account_numbers),
                  "skipped_unknown": skipped},
    )

    return BulkMembershipResult(
        action=action,
        requested_count=len(accounts) + len(skipped),
        affected_count=int(affected),
        skipped_unknown=skipped,
    )


@router.get("/{code}/memberships")
def list_memberships(
    code: str,
    site: Optional[str] = Query(None, description="Filter by site code."),
    search: Optional[str] = Query(None, description="Match account_number prefix or customer name."),
    page: int = Query(1, ge=1),
    page_size: int = Query(50, ge=1, le=500),
    user: CurrentUser = Depends(require_role(CCRole.superadmin)),
) -> Dict[str, Any]:
    """List the accounts tagged into a program with customer / site context."""
    code = code.strip().upper()

    with get_connection() as conn:
        cur = conn.cursor()
        prog = _resolve_program(cur, code)

        where = ["pm.program_id = %s"]
        params: List[Any] = [prog["id"]]
        if site:
            where.append("a.community = %s")
            params.append(site.strip().upper())
        if search:
            where.append(
                "(pm.account_number ILIKE %s OR "
                " COALESCE(c.first_name, '') || ' ' || COALESCE(c.last_name, '') ILIKE %s)"
            )
            like = f"%{search.strip()}%"
            params.extend([like, like])

        where_sql = " AND ".join(where)
        offset = (page - 1) * page_size

        cur.execute(
            f"""
            SELECT COUNT(*)
              FROM program_memberships pm
              JOIN accounts a ON a.account_number = pm.account_number
              LEFT JOIN customers c ON c.id = a.customer_id
             WHERE {where_sql}
            """,
            params,
        )
        total = int(cur.fetchone()[0] or 0)

        cur.execute(
            f"""
            SELECT pm.account_number,
                   c.customer_id_legacy,
                   COALESCE(c.first_name, '') || ' ' ||
                   COALESCE(NULLIF(c.middle_name, ''), '') || ' ' ||
                   COALESCE(c.last_name, '') AS customer_name,
                   a.community,
                   pm.joined_at,
                   pm.claim_milestone,
                   pm.notes,
                   pm.added_by
              FROM program_memberships pm
              JOIN accounts a ON a.account_number = pm.account_number
              LEFT JOIN customers c ON c.id = a.customer_id
             WHERE {where_sql}
             ORDER BY pm.joined_at DESC, pm.account_number ASC
             LIMIT %s OFFSET %s
            """,
            params + [page_size, offset],
        )
        rows = cur.fetchall()

    items: List[MembershipOut] = []
    for r in rows:
        items.append(
            MembershipOut(
                account_number=r[0],
                customer_id_legacy=str(r[1]) if r[1] is not None else None,
                customer_name=" ".join(str(r[2] or "").split()) or None,
                site_id=r[3],
                joined_at=_iso(r[4]) or "",
                claim_milestone=r[5],
                notes=r[6],
                added_by=r[7],
            )
        )

    return {
        "program": code,
        "total": total,
        "page": page,
        "page_size": page_size,
        "items": [m.model_dump() for m in items],
    }


# ---------------------------------------------------------------------------
# Tokens -- issue / list / revoke
# ---------------------------------------------------------------------------

@router.post("/{code}/tokens", response_model=TokenIssueResponse, status_code=201)
def issue_token(
    code: str,
    req: TokenIssueRequest,
    user: CurrentUser = Depends(require_role(CCRole.superadmin)),
):
    """Issue a fresh bearer token for a program. **The plaintext token is
    returned once** -- store it immediately. Only the sha256 hash and
    plaintext prefix are persisted.
    """
    code = code.strip().upper()
    plaintext = _new_token()
    th = _hash_token(plaintext)
    prefix = plaintext[:8]
    expires_at: Optional[datetime] = None
    if req.lifetime_days:
        expires_at = datetime.now(timezone.utc) + timedelta(days=req.lifetime_days)

    with get_connection() as conn:
        cur = conn.cursor()
        prog = _resolve_program(cur, code)
        cur.execute(
            """
            INSERT INTO odyssey_api_tokens
                (program_id, token_hash, token_prefix, label, issued_by, expires_at)
            VALUES (%s, %s, %s, %s, %s, %s)
            RETURNING id, issued_at
            """,
            (prog["id"], th, prefix, req.label, user.user_id, expires_at),
        )
        token_id, issued_at = cur.fetchone()
        conn.commit()

    try_log_mutation(
        user, "create", "odyssey_api_tokens", str(token_id),
        new_values={"program_code": code, "label": req.label,
                    "expires_at": _iso(expires_at), "token_prefix": prefix},
        metadata={"origin": "programs_token_issue"},
    )
    logger.info(
        "Odyssey token issued: program=%s id=%d prefix=%s by=%s",
        code, token_id, prefix, user.user_id,
    )

    return TokenIssueResponse(
        token=plaintext,
        summary=TokenSummaryOut(
            id=token_id,
            label=req.label,
            token_prefix=prefix,
            issued_at=_iso(issued_at) or "",
            issued_by=user.user_id,
            expires_at=_iso(expires_at),
        ),
    )


@router.get("/{code}/tokens", response_model=List[TokenSummaryOut])
def list_tokens(
    code: str,
    include_revoked: bool = Query(False),
    user: CurrentUser = Depends(require_role(CCRole.superadmin)),
):
    """List all tokens issued for a program. Plaintext is never returned."""
    code = code.strip().upper()
    with get_connection() as conn:
        cur = conn.cursor()
        prog = _resolve_program(cur, code)
        sql = """
            SELECT id, label, token_prefix, issued_at, issued_by,
                   expires_at, revoked_at, last_used_at, last_used_ip
              FROM odyssey_api_tokens
             WHERE program_id = %s
        """
        if not include_revoked:
            sql += " AND revoked_at IS NULL"
        sql += " ORDER BY issued_at DESC, id DESC"
        cur.execute(sql, (prog["id"],))
        rows = cur.fetchall()

    return [
        TokenSummaryOut(
            id=r[0],
            label=r[1],
            token_prefix=r[2],
            issued_at=_iso(r[3]) or "",
            issued_by=r[4],
            expires_at=_iso(r[5]),
            revoked_at=_iso(r[6]),
            last_used_at=_iso(r[7]),
            last_used_ip=r[8],
        )
        for r in rows
    ]


@router.delete("/{code}/tokens/{token_id}")
def revoke_token(
    code: str,
    token_id: int,
    user: CurrentUser = Depends(require_role(CCRole.superadmin)),
):
    """Revoke a token. Subsequent calls bearing the token will 401."""
    code = code.strip().upper()
    with get_connection() as conn:
        cur = conn.cursor()
        prog = _resolve_program(cur, code)
        cur.execute(
            """
            UPDATE odyssey_api_tokens
               SET revoked_at = NOW(), revoked_by = %s
             WHERE id = %s AND program_id = %s AND revoked_at IS NULL
             RETURNING id
            """,
            (user.user_id, token_id, prog["id"]),
        )
        row = cur.fetchone()
        conn.commit()
    if row is None:
        raise HTTPException(status_code=404, detail="Token not found or already revoked.")

    try_log_mutation(
        user, "delete", "odyssey_api_tokens", str(token_id),
        metadata={"origin": "programs_token_revoke", "program_code": code},
    )
    logger.info("Odyssey token revoked: program=%s id=%d by=%s", code, token_id, user.user_id)
    return {"revoked": True, "token_id": token_id}


# ---------------------------------------------------------------------------
# Dataset preview -- dry-run for staff verification
# ---------------------------------------------------------------------------

@router.get("/{code}/preview")
def preview_dataset(
    code: str,
    dataset: str = Query(..., regex="^(electricity-payment|meter-metrics)$"),
    from_: str = Query(..., alias="from"),
    to: str = Query(...),
    page: int = Query(1, ge=1),
    page_size: int = Query(50, ge=1, le=500),
    user: CurrentUser = Depends(require_role(CCRole.superadmin)),
) -> Dict[str, Any]:
    """Same payload an Odyssey caller would receive, but reachable with the
    employee JWT instead of a bearer token. Useful before running the
    Odyssey validator -- staff can sanity-check the rows without having to
    handle production secrets.
    """
    from odyssey_api import (  # local import to avoid mutual import cycles
        ODYSSEY_MAX_WINDOW_HOURS,  # noqa: F401  (used implicitly via _validate_window)
        _format_meter_metric_record,
        _format_payment_record,
        _meter_metrics_query,
        _parse_iso,
        _payment_query,
        _validate_window,
    )

    code = code.strip().upper()
    with get_connection() as conn:
        cur = conn.cursor()
        prog = _resolve_program(cur, code)

    frm_dt = _parse_iso("from", from_)
    to_dt = _parse_iso("to", to)
    _validate_window(frm_dt, to_dt)
    offset = (page - 1) * page_size

    if dataset == "electricity-payment":
        from country_config import COUNTRY, get_currency_for_site
        sql, params = _payment_query(prog["id"], frm_dt, to_dt, page_size, offset)
        with get_connection() as conn:
            import psycopg2.extras
            cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
            cur.execute(sql, params)
            rows = cur.fetchall()
        total = int(rows[0]["total_count"]) if rows else 0
        data = [
            _format_payment_record(r, currency=get_currency_for_site(r.get("site_id") or "") or COUNTRY.currency)
            for r in rows
        ]
    else:
        sql, params = _meter_metrics_query(prog["id"], frm_dt, to_dt, page_size, offset)
        with get_connection() as conn:
            import psycopg2.extras
            cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
            cur.execute(sql, params)
            rows = cur.fetchall()
        total = int(rows[0]["total_count"]) if rows else 0
        data = [_format_meter_metric_record(r) for r in rows]

    return {
        "dataset": dataset,
        "program": code,
        "country": prog["country_code"],
        "from": frm_dt.isoformat(),
        "to": to_dt.isoformat(),
        "page": page,
        "page_size": page_size,
        "total": total,
        "count": len(data),
        "data": data,
        "preview": True,
    }


# ---------------------------------------------------------------------------
# Connections claim export (Phase 3)
# ---------------------------------------------------------------------------

# (Display header, parsingHeaders ID, source mapping)
# Source mapping is one of:
#   ('row', column_name)   -- read straight from the SQL row dict
#   ('const', value)       -- constant
#   ('callable', fn)       -- fn(row) -> value
#
# Headers + parsingHeader IDs are derived from
# docs/uef_zedsi_claim_template.xlsx, sheets ``Connections`` (row 1) and
# ``parsingHeaders`` (row 1). Keep order in lock-step with the template so
# Odyssey ingestion accepts the file unchanged.
_CONN_COLUMNS: List[Dict[str, Any]] = [
    {"display": "Submittable",                                                          "parsing": "submittable",                  "source": ("const", "TRUE")},
    {"display": "Verification Status",                                                  "parsing": "verificationStatus",           "source": ("const", "")},
    {"display": "Submission Status",                                                    "parsing": "submissionStatus",             "source": ("const", "Not Submitted")},
    {"display": "Milestone",                                                            "parsing": "systemConfiguration",          "source": ("row", "claim_milestone")},
    {"display": "Customer Name",                                                        "parsing": "customer.name",                "source": ("callable", lambda r: " ".join(s for s in (r.get("first_name"), r.get("middle_name"), r.get("last_name")) if s).strip())},
    {"display": "Smart Meter Serial Number",                                            "parsing": "remoteId",                     "source": ("row", "meter_serial")},
    {"display": "Customer ID",                                                          "parsing": "customer.governmentIdNumber",  "source": ("row", "national_id")},
    {"display": "Customer Gender",                                                      "parsing": "customer.gender",              "source": ("row", "gender")},
    {"display": "Customer Category",                                                    "parsing": "customer.simpleCategory",      "source": ("row", "simple_category")},
    {"display": "Customer Type",                                                        "parsing": "customer.type",                "source": ("row", "customer_type")},
    {"display": "Country",                                                              "parsing": "customer.locationState",       "source": ("callable", lambda r: r.get("_country_name") or "")},
    {"display": "Location (Province)",                                                  "parsing": "customer.locationDistrict",    "source": ("row", "district")},
    {"display": "Location (Address)",                                                   "parsing": "customer.locationAddress",     "source": ("row", "street_address")},
    {"display": "Customer Phone Number",                                                "parsing": "customer.phoneNumber",         "source": ("callable", lambda r: r.get("phone") or r.get("cell_phone_1") or "")},
    {"display": "Latitude",                                                             "parsing": "customer.latitude",            "source": ("row", "gps_lat")},
    {"display": "Longitude",                                                            "parsing": "customer.longitude",           "source": ("row", "gps_lon")},
    {"display": "Date Connected",                                                       "parsing": "connected",                    "source": ("callable", lambda r: r["date_service_connected"].isoformat() if r.get("date_service_connected") else "")},
    {"display": "Previous Energy Source",                                               "parsing": "energySource",                 "source": ("row", "previous_energy_source")},
    {"display": "Average Monthly Electricity Revenue from Customer (Kwacha)",           "parsing": "integer5",                     "source": ("callable", lambda r: round(float(r.get("avg_monthly_revenue") or 0), 2))},
    {"display": "Average Monthly Energy Consumption per Customer (kWh/Month)",          "parsing": "dailyEnergyAvailable",         "source": ("callable", lambda r: round(float(r.get("avg_monthly_kwh") or 0), 2))},
    {"display": "Indicate if your firm deployed this PUE to this customer",             "parsing": "salesMisc1",                   "source": ("callable", lambda r: (r.get("pue_equipment") or {}).get("primary_deployer", ""))},
    {"display": "Primary PUE Equipment Connected",                                      "parsing": "productType",                  "source": ("callable", lambda r: (r.get("pue_equipment") or {}).get("primary_type", ""))},
    {"display": "Primary PUE Equipment Brand/Manufacturer",                             "parsing": "model",                        "source": ("callable", lambda r: (r.get("pue_equipment") or {}).get("primary_brand", ""))},
    {"display": "Primray PUE Equipment serial number",                                  "parsing": "remoteControl",                "source": ("callable", lambda r: (r.get("pue_equipment") or {}).get("primary_serial", ""))},
    {"display": "Primary PUE Equipment Wattage/Load (kW) ",                             "parsing": "capacity",                     "source": ("callable", lambda r: (r.get("pue_equipment") or {}).get("primary_kw", ""))},
    {"display": "Primary Total CAPEX Cost of PUE Equipment ($)",                        "parsing": "integer3",                     "source": ("callable", lambda r: (r.get("pue_equipment") or {}).get("primary_capex_usd", ""))},
    {"display": "Secondary PUE Equipment",                                              "parsing": "salesMisc2",                   "source": ("callable", lambda r: (r.get("pue_equipment") or {}).get("secondary_type", ""))},
    {"display": " Secondary PUE Equipment serial number",                               "parsing": "serialNumber",                 "source": ("callable", lambda r: (r.get("pue_equipment") or {}).get("secondary_serial", ""))},
    {"display": "Secondary PUE Equipment Wattage/Load (kW)",                            "parsing": "decimal1",                     "source": ("callable", lambda r: (r.get("pue_equipment") or {}).get("secondary_kw", ""))},
    {"display": "Tertiary PUE Equipment Connected",                                     "parsing": "salesMisc3",                   "source": ("callable", lambda r: (r.get("pue_equipment") or {}).get("tertiary_type", ""))},
    {"display": "Tertiary PUE Equipment Serial Number",                                 "parsing": "text10",                       "source": ("callable", lambda r: (r.get("pue_equipment") or {}).get("tertiary_serial", ""))},
    {"display": "Tertiary PUE Equipment Wattage/Load (kW) ",                            "parsing": "decimal2",                     "source": ("callable", lambda r: (r.get("pue_equipment") or {}).get("tertiary_kw", ""))},
    {"display": "Which demand stimulation activity did the customer benefit from (Select Impact 1):",  "parsing": "text1",  "source": ("callable", lambda r: (r.get("pue_equipment") or {}).get("impact_1", ""))},
    {"display": "If Other, please specify:",                                            "parsing": "text2",                        "source": ("callable", lambda r: (r.get("pue_equipment") or {}).get("impact_other", ""))},
    {"display": "Which demand stimulation activity did the customer benefit from (Select Impact 2):",  "parsing": "text3",  "source": ("callable", lambda r: (r.get("pue_equipment") or {}).get("impact_2", ""))},
    {"display": "Which demand stimulation activity did the customer benefit from (Select Impact 3):",  "parsing": "text4",  "source": ("callable", lambda r: (r.get("pue_equipment") or {}).get("impact_3", ""))},
    {"display": "Submitted Date",                                                       "parsing": "submitted",                    "source": ("const", "")},
    {"display": "Approved",                                                             "parsing": "approved",                     "source": ("const", "")},
    {"display": "Paid",                                                                 "parsing": "paid",                         "source": ("const", "")},
    {"display": "Upload Date",                                                          "parsing": "created",                      "source": ("const", "")},
]


def _country_name(code: Optional[str]) -> str:
    if not code:
        return ""
    from country_config import _REGISTRY  # type: ignore[attr-defined]
    cfg = _REGISTRY.get(code.strip().upper())
    return cfg.name if cfg else code


def _resolve_cell(row: Dict[str, Any], source: tuple) -> Any:
    kind = source[0]
    if kind == "row":
        return row.get(source[1])
    if kind == "const":
        return source[1]
    if kind == "callable":
        try:
            return source[1](row)
        except Exception:  # noqa: BLE001 -- a single bad row shouldn't fail the whole export
            return ""
    return ""


@router.get("/{code}/connections.xlsx")
def export_connections_xlsx(
    code: str,
    milestone: Optional[str] = Query(
        None,
        description="If set, only export memberships with this claim_milestone (e.g. 'Milestone 1').",
    ),
    user: CurrentUser = Depends(require_role(CCRole.superadmin)),
):
    """Generate a Connections claim XLSX in the exact Odyssey upload format.

    Mirrors the column order and ``parsingHeaders`` row of
    ``docs/uef_zedsi_claim_template.xlsx`` so the file can be uploaded to
    Odyssey without manual fix-up. Each row is one in-program account.

    Per-customer averages (revenue / kWh) are computed over the trailing
    90 days from ``transactions`` and ``hourly_consumption``.
    """
    try:
        from openpyxl import Workbook
    except ImportError:
        raise HTTPException(status_code=500, detail="openpyxl not installed.")

    code = code.strip().upper()

    with get_connection() as conn:
        cur = conn.cursor()
        prog = _resolve_program(cur, code)

        where = ["pm.program_id = %s"]
        params: List[Any] = [prog["id"]]
        if milestone:
            where.append("pm.claim_milestone = %s")
            params.append(milestone)

        sql = f"""
            WITH avg90 AS (
                SELECT account_number,
                       AVG(transaction_amount) FILTER (WHERE is_payment) * 30
                           / NULLIF(GREATEST(EXTRACT(EPOCH FROM (
                               COALESCE(MAX(transaction_date), NOW()) -
                               COALESCE(MIN(transaction_date), NOW())
                             )) / 86400, 1), 0)        AS avg_monthly_revenue,
                       0::float                         AS _placeholder
                  FROM transactions
                 WHERE transaction_date >= NOW() - INTERVAL '90 days'
                 GROUP BY account_number
            ),
            kwh90 AS (
                SELECT account_number,
                       SUM(kwh) / 3.0 AS avg_monthly_kwh   -- 90 days = 3 months
                  FROM hourly_consumption
                 WHERE reading_hour >= NOW() - INTERVAL '90 days'
                 GROUP BY account_number
            ),
            active_meter AS (
                SELECT DISTINCT ON (m.account_number)
                       m.account_number, m.meter_id
                  FROM meters m
                 ORDER BY m.account_number, (m.status = 'active') DESC, m.meter_id
            )
            SELECT pm.account_number,
                   pm.claim_milestone,
                   pm.pue_equipment,
                   c.first_name, c.middle_name, c.last_name,
                   c.gender, c.simple_category, c.customer_type,
                   c.national_id, c.district, c.street_address,
                   c.phone, c.cell_phone_1,
                   c.gps_lat, c.gps_lon,
                   c.date_service_connected,
                   c.previous_energy_source,
                   a.community,
                   am.meter_id AS meter_serial,
                   COALESCE(av.avg_monthly_revenue, 0) AS avg_monthly_revenue,
                   COALESCE(kw.avg_monthly_kwh, 0)     AS avg_monthly_kwh
              FROM program_memberships pm
              JOIN accounts a  ON a.account_number = pm.account_number
              LEFT JOIN customers c     ON c.id = a.customer_id
              LEFT JOIN active_meter am ON am.account_number = pm.account_number
              LEFT JOIN avg90 av        ON av.account_number = pm.account_number
              LEFT JOIN kwh90 kw        ON kw.account_number = pm.account_number
             WHERE {' AND '.join(where)}
             ORDER BY a.community, pm.account_number
        """
        cur.execute(sql, params)
        col_names = [d[0] for d in cur.description]
        rows: List[Dict[str, Any]] = [dict(zip(col_names, r)) for r in cur.fetchall()]

    # Resolve country name per row from the bound program (single value, but
    # we set it per-row for compatibility with future cross-country programs).
    country_name = _country_name(prog.get("country_code"))
    for r in rows:
        r["_country_name"] = country_name

    wb = Workbook()
    ws = wb.active
    ws.title = "Connections"

    # Row 1: human-readable headers (matches xlsx template exactly).
    for col_idx, spec in enumerate(_CONN_COLUMNS, 1):
        cell = ws.cell(row=1, column=col_idx, value=spec["display"])
        cell.font = cell.font.copy(bold=True)

    # Row 2 onwards: data
    for row_idx, row in enumerate(rows, 2):
        for col_idx, spec in enumerate(_CONN_COLUMNS, 1):
            value = _resolve_cell(row, spec["source"])
            if value is not None and not isinstance(value, (str, int, float, bool)):
                value = str(value)
            ws.cell(row=row_idx, column=col_idx, value=value)

    # Hidden parsingHeaders sheet so Odyssey internal IDs travel with the file.
    ws_parse = wb.create_sheet("parsingHeaders")
    for col_idx, spec in enumerate(_CONN_COLUMNS, 1):
        ws_parse.cell(row=1, column=col_idx, value=spec["parsing"])
    ws_parse.sheet_state = "hidden"

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    fname = f"{code}_connections{('_' + milestone.replace(' ', '_')) if milestone else ''}.xlsx"
    logger.info(
        "Connections export: program=%s milestone=%s rows=%d by=%s",
        code, milestone or "*", len(rows), user.user_id,
    )
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={fname}"},
    )
