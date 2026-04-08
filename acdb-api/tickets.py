"""
O&M Maintenance Ticket endpoints.

Stores corrective (fault) maintenance records and mirrors tickets from
uGridPlan.  Supports full CRUD, filtering, and Excel export matching the
team's "O&M Corrective Maintenance Report" format.

  POST   /api/tickets             — create a ticket
  GET    /api/tickets             — list tickets (paginated, filterable)
  GET    /api/tickets/export      — download Excel report
  GET    /api/tickets/{id}        — single ticket
  PATCH  /api/tickets/{id}        — update ticket fields
"""

import io
import logging
from datetime import datetime, timezone
from typing import Optional

from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from pydantic import BaseModel

from customer_api import get_connection
from middleware import CurrentUser, require_employee

logger = logging.getLogger("cc-api.tickets")

router = APIRouter(prefix="/api/tickets", tags=["tickets"])

ALL_COLUMNS = [
    "id", "ugp_ticket_id", "source", "phone", "customer_id",
    "account_number", "site_code", "fault_description",
    "category", "priority", "reported_by", "created_at",
    "ticket_name", "failure_time", "services_affected",
    "troubleshooting_steps", "cause_of_fault", "precautions",
    "restoration_time", "resolution_approach", "duration",
    "status", "updated_at", "resolved_by",
]

SELECT_ALL = ", ".join(ALL_COLUMNS)


def _row_to_dict(row):
    t = dict(zip(ALL_COLUMNS, row))
    for k in ("created_at", "failure_time", "restoration_time", "updated_at"):
        if t.get(k) and hasattr(t[k], "isoformat"):
            t[k] = t[k].isoformat()
    return t


# ---------------------------------------------------------------------------
# Create
# ---------------------------------------------------------------------------

class TicketCreate(BaseModel):
    ugp_ticket_id: Optional[str] = None
    source: str = "portal"
    phone: Optional[str] = None
    customer_id: Optional[int] = None
    account_number: Optional[str] = None
    site_code: Optional[str] = None
    fault_description: Optional[str] = None
    category: Optional[str] = None
    priority: Optional[str] = None
    reported_by: Optional[str] = None
    ticket_name: Optional[str] = None
    failure_time: Optional[str] = None
    services_affected: Optional[str] = None
    troubleshooting_steps: Optional[str] = None
    cause_of_fault: Optional[str] = None
    precautions: Optional[str] = None
    restoration_time: Optional[str] = None
    resolution_approach: Optional[str] = None
    duration: Optional[str] = None
    status: str = "open"
    resolved_by: Optional[str] = None


@router.post("")
def create_ticket(body: TicketCreate, user: CurrentUser = Depends(require_employee)):
    ugp_id = body.ugp_ticket_id or ""
    with get_connection() as conn:
        cur = conn.cursor()
        cur.execute(
            """
            INSERT INTO wa_tickets
                (ugp_ticket_id, source, phone, customer_id, account_number,
                 site_code, fault_description, category, priority, reported_by,
                 ticket_name, failure_time, services_affected,
                 troubleshooting_steps, cause_of_fault, precautions,
                 restoration_time, resolution_approach, duration, status,
                 resolved_by)
            VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
            RETURNING id, created_at
            """,
            (
                ugp_id, body.source, body.phone, body.customer_id,
                body.account_number, body.site_code, body.fault_description,
                body.category, body.priority, body.reported_by,
                body.ticket_name, body.failure_time or None,
                body.services_affected, body.troubleshooting_steps,
                body.cause_of_fault, body.precautions,
                body.restoration_time or None, body.resolution_approach,
                body.duration, body.status, body.resolved_by,
            ),
        )
        row = cur.fetchone()
        conn.commit()
        logger.info("Ticket created: id=%s site=%s", row[0], body.site_code)
        return {
            "status": "ok",
            "id": row[0],
            "ugp_ticket_id": ugp_id,
            "created_at": row[1].isoformat() if row[1] else None,
        }


# ---------------------------------------------------------------------------
# Export  (must be before /{ticket_ref} to avoid route conflict)
# ---------------------------------------------------------------------------

REPORT_HEADERS = [
    "Ticket Name", "Site", "Failure Time", "Fault Description",
    "Service(s) Affected", "Troubleshooting Steps", "Cause of Fault",
    "Precautions", "Restoration Time", "Resolution Approach", "Duration",
    "Status", "Reported By", "Resolved By", "Category", "Priority",
]


@router.get("/export")
def export_tickets(
    site_code: Optional[str] = Query(None),
    status: Optional[str] = Query(None),
    quarter: Optional[str] = Query(None, description="Quarter filter, e.g. 2026-Q1"),
    user: CurrentUser = Depends(require_employee),
):
    """Download maintenance report as Excel (.xlsx)."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        raise HTTPException(status_code=500, detail="openpyxl not installed")

    clauses = []
    params: list = []
    if site_code:
        clauses.append("site_code = %s")
        params.append(site_code.upper())
    if status:
        clauses.append("status = %s")
        params.append(status)
    if quarter:
        try:
            year, q = quarter.split("-Q")
            q_num = int(q)
            start_month = (q_num - 1) * 3 + 1
            end_month = start_month + 3
            start_date = "{}-{:02d}-01".format(int(year), start_month)
            if end_month > 12:
                end_date = "{}-01-01".format(int(year) + 1)
            else:
                end_date = "{}-{:02d}-01".format(int(year), end_month)
            clauses.append("COALESCE(failure_time, created_at) >= %s")
            clauses.append("COALESCE(failure_time, created_at) < %s")
            params.append(start_date)
            params.append(end_date)
        except (ValueError, IndexError):
            pass

    where = ("WHERE " + " AND ".join(clauses)) if clauses else ""

    with get_connection() as conn:
        cur = conn.cursor()
        cur.execute(
            "SELECT {} FROM wa_tickets {} ORDER BY COALESCE(failure_time, created_at) ASC".format(
                SELECT_ALL, where
            ),
            params,
        )
        rows = [_row_to_dict(r) for r in cur.fetchall()]

    wb = Workbook()
    ws = wb.active
    ws.title = "Maintenance Log"

    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell_align = Alignment(vertical="top", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    title_text = "O&M Corrective (Fault) Maintenance Report"
    if quarter:
        title_text += " — {}".format(quarter)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(REPORT_HEADERS))
    title_cell = ws.cell(row=1, column=1, value=title_text)
    title_cell.font = Font(bold=True, size=14, color="2F5496")
    title_cell.alignment = Alignment(horizontal="center")

    for col_idx, header in enumerate(REPORT_HEADERS, 1):
        cell = ws.cell(row=3, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    def _fmt_dt(iso_str):
        if not iso_str:
            return ""
        try:
            dt = datetime.fromisoformat(str(iso_str))
            return dt.strftime("%d-%b-%Y %H:%M")
        except (ValueError, TypeError):
            return str(iso_str)

    for row_idx, t in enumerate(rows, 4):
        values = [
            t.get("ticket_name") or t.get("fault_description", "")[:60] or "",
            t.get("site_code") or "",
            _fmt_dt(t.get("failure_time") or t.get("created_at")),
            t.get("fault_description") or "",
            t.get("services_affected") or "",
            t.get("troubleshooting_steps") or "",
            t.get("cause_of_fault") or "",
            t.get("precautions") or "",
            _fmt_dt(t.get("restoration_time")),
            t.get("resolution_approach") or "",
            t.get("duration") or "",
            (t.get("status") or "open").title(),
            t.get("reported_by") or "",
            t.get("resolved_by") or "",
            t.get("category") or "",
            t.get("priority") or "",
        ]
        for col_idx, val in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.alignment = cell_align
            cell.border = thin_border

    col_widths = [22, 10, 18, 35, 22, 35, 25, 25, 18, 25, 12, 10, 15, 15, 14, 10]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[ws.cell(row=3, column=i).column_letter].width = w

    ws.auto_filter.ref = "A3:{}{}".format(
        ws.cell(row=3, column=len(REPORT_HEADERS)).column_letter,
        3 + len(rows),
    )

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    fname = "Maintenance_Log"
    if quarter:
        fname += "_{}".format(quarter.replace("-", "_"))
    fname += ".xlsx"

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename={}".format(fname)},
    )


# ---------------------------------------------------------------------------
# List (paginated)
# ---------------------------------------------------------------------------

@router.get("")
def list_tickets(
    limit: int = Query(50, ge=1, le=500),
    offset: int = Query(0, ge=0),
    site_code: Optional[str] = Query(None),
    account_number: Optional[str] = Query(None),
    status: Optional[str] = Query(None),
    search: Optional[str] = Query(None),
    user: CurrentUser = Depends(require_employee),
):
    clauses = []
    params: list = []

    if site_code:
        clauses.append("site_code = %s")
        params.append(site_code.upper())
    if account_number:
        clauses.append("account_number = %s")
        params.append(account_number)
    if status:
        clauses.append("status = %s")
        params.append(status)
    if search:
        clauses.append(
            "(ticket_name ILIKE %s OR fault_description ILIKE %s "
            "OR site_code ILIKE %s OR account_number ILIKE %s "
            "OR reported_by ILIKE %s OR cause_of_fault ILIKE %s)"
        )
        s = "%{}%".format(search)
        params.extend([s, s, s, s, s, s])

    where = ("WHERE " + " AND ".join(clauses)) if clauses else ""

    with get_connection() as conn:
        cur = conn.cursor()

        count_params = list(params)
        cur.execute(
            "SELECT COUNT(*) FROM wa_tickets {}".format(where),
            count_params,
        )
        total = cur.fetchone()[0]

        params.extend([limit, offset])
        cur.execute(
            "SELECT {} FROM wa_tickets {} ORDER BY COALESCE(failure_time, created_at) DESC LIMIT %s OFFSET %s".format(
                SELECT_ALL, where
            ),
            params,
        )
        tickets = [_row_to_dict(r) for r in cur.fetchall()]
        return {"tickets": tickets, "total": total, "count": len(tickets)}


# ---------------------------------------------------------------------------
# Get single
# ---------------------------------------------------------------------------

@router.get("/{ticket_ref}")
def get_ticket(
    ticket_ref: str,
    user: CurrentUser = Depends(require_employee),
):
    with get_connection() as conn:
        cur = conn.cursor()
        if ticket_ref.isdigit():
            cur.execute(
                "SELECT {} FROM wa_tickets WHERE id = %s".format(SELECT_ALL),
                (int(ticket_ref),),
            )
        else:
            cur.execute(
                "SELECT {} FROM wa_tickets WHERE ugp_ticket_id = %s".format(SELECT_ALL),
                (ticket_ref,),
            )
        row = cur.fetchone()
        if not row:
            raise HTTPException(status_code=404, detail="Ticket not found: {}".format(ticket_ref))
        return _row_to_dict(row)


# ---------------------------------------------------------------------------
# Update (partial)
# ---------------------------------------------------------------------------

class TicketUpdate(BaseModel):
    ticket_name: Optional[str] = None
    fault_description: Optional[str] = None
    failure_time: Optional[str] = None
    services_affected: Optional[str] = None
    troubleshooting_steps: Optional[str] = None
    cause_of_fault: Optional[str] = None
    precautions: Optional[str] = None
    restoration_time: Optional[str] = None
    resolution_approach: Optional[str] = None
    duration: Optional[str] = None
    status: Optional[str] = None
    resolved_by: Optional[str] = None
    site_code: Optional[str] = None
    category: Optional[str] = None
    priority: Optional[str] = None
    reported_by: Optional[str] = None
    account_number: Optional[str] = None
    phone: Optional[str] = None


UPDATABLE = {
    "ticket_name", "fault_description", "failure_time", "services_affected",
    "troubleshooting_steps", "cause_of_fault", "precautions",
    "restoration_time", "resolution_approach", "duration", "status",
    "resolved_by", "site_code", "category", "priority", "reported_by",
    "account_number", "phone",
}


@router.patch("/{ticket_id}")
def update_ticket(
    ticket_id: int,
    body: TicketUpdate,
    user: CurrentUser = Depends(require_employee),
):
    changes = {k: v for k, v in body.dict(exclude_unset=True).items() if k in UPDATABLE}
    if not changes:
        raise HTTPException(status_code=400, detail="No fields to update")

    changes["updated_at"] = datetime.now(timezone.utc).isoformat()

    sets = ["{} = %s".format(k) for k in changes]
    vals = list(changes.values()) + [ticket_id]

    with get_connection() as conn:
        cur = conn.cursor()
        cur.execute(
            "UPDATE wa_tickets SET {} WHERE id = %s RETURNING id".format(", ".join(sets)),
            vals,
        )
        if not cur.fetchone():
            raise HTTPException(status_code=404, detail="Ticket not found")
        conn.commit()

    return {"status": "ok", "id": ticket_id, "updated_fields": list(changes.keys())}
