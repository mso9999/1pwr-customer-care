"""
O&M Quarterly Report data endpoints.

Auto-generates analytics from PostgreSQL data to mirror the figures in
the SMP Operations & Maintenance Quarterly Report:
  - Customer statistics per site (total, active, new per quarter)
  - Customer connection growth over time (quarterly)
  - Consumption per site per quarter (kWh)
  - Sales/revenue per site per quarter (LSL)
  - Cumulative consumption and sales trends
  - Generation vs consumption per site
  - Average consumption per customer trends
  - Site overview (concession list with districts)
"""

import io
import json
import logging
import math
import os
from collections import defaultdict
from datetime import date, datetime, time, timedelta, timezone
from typing import Any, Dict, List, Optional, Set, Tuple

from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse

from models import CurrentUser
from middleware import require_employee

logger = logging.getLogger("acdb-api.om-report")

router = APIRouter(prefix="/api/om-report", tags=["om-report"])

from country_config import (
    SITE_ABBREV, KNOWN_SITES, SITE_DISTRICTS,
    COUNTRY, CURRENCY, CURRENCY_SYMBOL,
)


def _matches_customer_type(ctype: str, filter_type: str) -> bool:
    """Check if a customer type matches a filter, treating HH as aggregate of HH1+HH2+HH3."""
    ct = ctype.upper()
    ft = filter_type.upper()
    if ft == "HH":
        return ct.startswith("HH")
    return ct == ft


def _normalize_power_kw_value(raw_kw: Any, customer_type: str = "", source: str = "") -> Optional[float]:
    """Normalize mixed W/kW power values to kW for charts and exports.

    `meter_readings.power_kw` is intended to store kW, but some upstream writers
    have historically inserted watts. We can normalize the known/proven paths
    (`iot`) and apply a conservative guard for obvious household/thundercloud
    spikes until historical rows are backfilled.
    """
    if raw_kw is None:
        return None
    try:
        kw_val = float(raw_kw)
    except (ValueError, TypeError):
        return None
    if not math.isfinite(kw_val):
        return None

    ctype = str(customer_type or "").strip().upper()
    src = str(source or "").strip().lower()

    if src == "iot":
        return kw_val / 1000.0
    if kw_val > 20 and (src == "thundercloud" or ctype.startswith("HH")):
        return kw_val / 1000.0
    return kw_val


# SQL fragment to build account -> customer_type from the customers table
_ACCT_CTYPE_SQL = """
    SELECT a.account_number, c.customer_type
    FROM accounts a
    JOIN customers c ON a.customer_id = c.id
    WHERE c.customer_type IS NOT NULL AND c.customer_type <> ''
"""


def _get_connection():
    from customer_api import get_connection
    return get_connection()


def _extract_site(account_number: str) -> str:
    """Extract site code from the last 3 chars of account number.

    Only returns a value if it matches a known site code, to prevent
    meter serial suffixes (e.g., '7E7', '3F0') from polluting charts.
    """
    if not account_number:
        return ""
    candidate = account_number.strip()[-3:].upper()
    if candidate in KNOWN_SITES:
        return candidate
    return ""


def _coerce_export_timestamp(raw_val: Any) -> Optional[datetime]:
    """Normalize DB/date/string values to a naive datetime."""
    if raw_val is None:
        return None
    if isinstance(raw_val, datetime):
        return raw_val.replace(tzinfo=None) if raw_val.tzinfo else raw_val
    if isinstance(raw_val, date):
        return datetime.combine(raw_val, time.min)
    if isinstance(raw_val, str):
        raw = raw_val.strip()
        if not raw:
            return None
        for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d", "%m/%d/%Y", "%d/%m/%Y"):
            try:
                text = raw[:19] if fmt == "%Y-%m-%d %H:%M:%S" else raw
                return datetime.strptime(text, fmt)
            except ValueError:
                continue
    return None


def _date_to_quarter(dt) -> str:
    """Convert a date/datetime to 'YYYY QN' string."""
    if dt is None:
        return ""
    if isinstance(dt, str):
        for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d", "%m/%d/%Y", "%d/%m/%Y"):
            try:
                dt = datetime.strptime(dt.strip(), fmt)
                break
            except (ValueError, AttributeError):
                continue
        else:
            return ""
    try:
        q = (dt.month - 1) // 3 + 1
        return f"{dt.year} Q{q}"
    except (AttributeError, TypeError):
        return ""


def _day_key(val) -> Optional[str]:
    """Normalise a date/datetime/string to 'YYYY-MM-DD' for unique-day tracking."""
    if val is None:
        return None
    if hasattr(val, 'strftime'):
        return val.strftime('%Y-%m-%d')
    if isinstance(val, str):
        return val[:10]
    return None


# ---------------------------------------------------------------------------
# 1. Portfolio Overview
# ---------------------------------------------------------------------------

@router.get("/overview")
def report_overview(user: CurrentUser = Depends(require_employee)):
    """Summary statistics for the report header."""
    with _get_connection() as conn:
        cursor = conn.cursor()

        cursor.execute(
            "SELECT community, date_service_terminated FROM customers "
            "WHERE community IS NOT NULL AND community <> ''"
        )
        total_customers = 0
        terminated = 0
        seen_sites: set = set()
        for row in cursor.fetchall():
            comm = str(row[0] or "").strip().upper()
            if comm not in KNOWN_SITES:
                continue
            total_customers += 1
            seen_sites.add(comm)
            if row[1] is not None:
                terminated += 1

        active_customers = total_customers - terminated
        sites = sorted(seen_sites)

        total_kwh = 0.0
        total_lsl = 0.0
        txn_source = "transactions"

        try:
            cursor.execute("SELECT SUM(amount_lsl) FROM monthly_transactions")
            row = cursor.fetchone()
            if row and row[0] is not None:
                total_lsl = float(row[0])
                txn_source = "monthly_transactions"
        except Exception:
            pass

        try:
            cursor.execute("SELECT SUM(kwh) FROM monthly_consumption")
            row = cursor.fetchone()
            if row and row[0] is not None:
                total_kwh = float(row[0])
        except Exception:
            pass

        if total_kwh == 0 or (total_lsl == 0 and txn_source == "transactions"):
            try:
                cursor.execute(
                    "SELECT SUM(kwh_value), SUM(transaction_amount) "
                    "FROM transactions"
                )
                row = cursor.fetchone()
                if row:
                    if total_kwh == 0 and row[0] is not None:
                        total_kwh = float(row[0])
                    if total_lsl == 0 and row[1] is not None:
                        total_lsl = float(row[1])
            except Exception:
                pass

        return {
            "total_customers": total_customers,
            "active_customers": active_customers,
            "terminated_customers": terminated,
            "total_sites": len(sites),
            "sites": sites,
            "total_mwh": round(total_kwh / 1000, 2),
            "total_lsl_thousands": round(total_lsl / 1000, 2),
            "data_sources": {
                "revenue": txn_source,
            },
        }


# ---------------------------------------------------------------------------
# 2. Customer Statistics per Site (Figure 14)
# ---------------------------------------------------------------------------

@router.get("/customer-stats")
def customer_stats_by_site(
    quarter: Optional[str] = Query(None, description="Quarter in YYYY QN format, e.g. '2025 Q4'"),
    user: CurrentUser = Depends(require_employee),
):
    """Customer counts per community: total, active, and new in the specified quarter."""
    with _get_connection() as conn:
        cursor = conn.cursor()

        cursor.execute(
            "SELECT community, customer_id_legacy, "
            "date_service_connected, date_service_terminated "
            "FROM customers WHERE community IS NOT NULL"
        )
        rows = cursor.fetchall()

        sites: Dict[str, Dict[str, int]] = defaultdict(lambda: {"total": 0, "active": 0, "new": 0})

        for row in rows:
            concession = str(row[0] or "").strip().upper()
            if not concession or concession not in KNOWN_SITES:
                continue

            connected_date = row[2]
            terminated_date = row[3]

            sites[concession]["total"] += 1

            if terminated_date is None:
                sites[concession]["active"] += 1

            if quarter and connected_date:
                cq = _date_to_quarter(connected_date)
                if cq == quarter:
                    sites[concession]["new"] += 1

        result = []
        for name in sorted(sites.keys()):
            data = sites[name]
            result.append({
                "concession": name,
                "total": data["total"],
                "active": data["active"],
                "new": data["new"],
                "activation_rate": round(data["active"] / data["total"] * 100, 1) if data["total"] > 0 else 0,
            })

        totals = {
            "total": sum(s["total"] for s in result),
            "active": sum(s["active"] for s in result),
            "new": sum(s["new"] for s in result),
        }

        return {"sites": result, "totals": totals, "quarter": quarter}


# ---------------------------------------------------------------------------
# 3. Customer Growth Over Time (Figure 15)
# ---------------------------------------------------------------------------

@router.get("/customer-growth")
def customer_growth(user: CurrentUser = Depends(require_employee)):
    """Quarterly customer connection growth since first site commissioned."""
    with _get_connection() as conn:
        cursor = conn.cursor()

        cursor.execute(
            "SELECT date_service_connected FROM customers "
            "WHERE date_service_connected IS NOT NULL"
        )
        rows = cursor.fetchall()

        quarterly: Dict[str, int] = defaultdict(int)
        for row in rows:
            q = _date_to_quarter(row[0])
            if q:
                quarterly[q] += 1

        sorted_quarters = sorted(quarterly.keys())
        cumulative = 0
        result = []
        for q in sorted_quarters:
            new_count = quarterly[q]
            cumulative += new_count
            result.append({
                "quarter": q,
                "new_customers": new_count,
                "cumulative": cumulative,
            })

        return {"growth": result, "total": cumulative}


@router.get("/customer-growth-by-site")
def customer_growth_by_site(
    site: Optional[str] = Query(None, description="Filter by site code (e.g. KET)"),
    user: CurrentUser = Depends(require_employee),
):
    """
    Per-site quarterly customer growth based on first transaction date,
    cross-referenced with the customers table to exclude orphaned/test accounts.

    Uses MIN(transaction_date) per account as the "active since" date.
    Only counts accounts that exist in the customers table (via accounts join).
    Site comes from customers.community (authoritative), not account suffix.
    """
    with _get_connection() as conn:
        cursor = conn.cursor()

        cursor.execute(
            "SELECT c.community, a.account_number, MIN(t.transaction_date) AS first_txn "
            "FROM transactions t "
            "JOIN accounts a ON t.account_number = a.account_number "
            "JOIN customers c ON a.customer_id = c.id "
            "WHERE t.transaction_date IS NOT NULL "
            "  AND c.community IS NOT NULL "
            "GROUP BY c.community, a.account_number"
        )
        rows = cursor.fetchall()

        site_quarterly: Dict[str, Dict[str, int]] = defaultdict(lambda: defaultdict(int))
        for community, _acct, first_txn in rows:
            s = (community or "").strip().upper()
            if not s or s not in KNOWN_SITES:
                continue
            if site and s != site.upper():
                continue
            q = _date_to_quarter(first_txn)
            if q:
                site_quarterly[s][q] += 1

        result = {}
        for s in sorted(site_quarterly):
            sorted_q = sorted(site_quarterly[s].keys())
            cum = 0
            quarters = []
            for q in sorted_q:
                new = site_quarterly[s][q]
                cum += new
                quarters.append({"quarter": q, "new": new, "cumulative": cum})
            result[s] = {
                "name": SITE_ABBREV.get(s, s),
                "total": cum,
                "quarters": quarters,
            }

        return {"sites": result, "source": "MIN(transaction_date) JOIN customers"}


# ---------------------------------------------------------------------------
# 4. Consumption by Site per Quarter (Figures 5, 12)
# ---------------------------------------------------------------------------

@router.get("/consumption-by-site")
def consumption_by_site(
    quarter: Optional[str] = Query(None, description="Filter to specific quarter"),
    user: CurrentUser = Depends(require_employee),
):
    """kWh consumption per site, optionally filtered by quarter."""
    with _get_connection() as conn:
        cursor = conn.cursor()

        try:
            cursor.execute(
                "SELECT account_number, transaction_date, kwh_value "
                "FROM transactions"
            )
            rows = cursor.fetchall()
        except Exception as e:
            logger.warning("Failed to query transactions for consumption: %s", e)
            return {"sites": [], "total_kwh": 0, "error": "No account history data found"}

        if not rows:
            return {"sites": [], "total_kwh": 0, "error": "No account history data found"}

        site_quarter: Dict[str, Dict[str, float]] = defaultdict(lambda: defaultdict(float))
        site_totals: Dict[str, float] = defaultdict(float)

        for row in rows:
            acct = str(row[0] or "").strip()
            site = _extract_site(acct)
            if not site or len(site) < 2:
                continue

            kwh = float(row[2] or 0)
            q = _date_to_quarter(row[1]) if row[1] else "Unknown"

            if quarter and q != quarter:
                continue

            site_quarter[site][q] += kwh
            site_totals[site] += kwh

        per_site = []
        for site_code in sorted(site_totals.keys()):
            quarters_data = {q: round(v, 2) for q, v in sorted(site_quarter[site_code].items())}
            per_site.append({
                "site": site_code,
                "name": SITE_ABBREV.get(site_code, site_code),
                "total_kwh": round(site_totals[site_code], 2),
                "quarters": quarters_data,
            })

        return {
            "sites": per_site,
            "total_kwh": round(sum(site_totals.values()), 2),
            "source_table": "transactions",
            "quarter_filter": quarter,
        }


# ---------------------------------------------------------------------------
# 5. Sales by Site per Quarter (Figure 6)
# ---------------------------------------------------------------------------

@router.get("/sales-by-site")
def sales_by_site(
    quarter: Optional[str] = Query(None),
    user: CurrentUser = Depends(require_employee),
):
    """LSL revenue per site, optionally filtered by quarter.

    Data source priority:
      1. monthly_transactions (SparkMeter, includes corrections)
      2. transactions (raw history, fallback)
    """
    with _get_connection() as conn:
        cursor = conn.cursor()

        rows: List[tuple] = []
        source_table = ""

        try:
            cursor.execute(
                "SELECT account_number, year_month, amount_lsl, community "
                "FROM monthly_transactions"
            )
            raw = cursor.fetchall()
            if raw:
                for r in raw:
                    acct = str(r[0] or "").strip()
                    ym = str(r[1] or "").strip()
                    lsl = float(r[2] or 0)
                    community = str(r[3] or "").strip().upper()
                    if acct and ym and lsl > 0:
                        try:
                            y, m = int(ym[:4]), int(ym[5:7])
                            dt = datetime(y, m, 15)
                        except (ValueError, IndexError):
                            continue
                        rows.append((acct, dt, lsl, community))
                if rows:
                    source_table = "monthly_transactions"
        except Exception as e:
            logger.warning("monthly_transactions query failed: %s", e)

        if not rows:
            try:
                cursor.execute(
                    "SELECT account_number, transaction_date, transaction_amount "
                    "FROM transactions"
                )
                raw = cursor.fetchall()
                if raw:
                    for r in raw:
                        acct = str(r[0] or "").strip()
                        lsl = float(r[2] or 0)
                        rows.append((acct, r[1], lsl, ""))
                    if rows:
                        source_table = "transactions"
            except Exception as e:
                logger.warning("Failed to query transactions for sales: %s", e)

        if not rows:
            return {"sites": [], "total_lsl": 0, "error": "No transaction data found"}

        site_quarter: Dict[str, Dict[str, float]] = defaultdict(
            lambda: defaultdict(float)
        )
        site_totals: Dict[str, float] = defaultdict(float)

        for acct, dt_or_ym, lsl, community in rows:
            site = (community.upper() if community else "") or _extract_site(acct)
            if not site or site not in KNOWN_SITES:
                continue
            q = _date_to_quarter(dt_or_ym) if dt_or_ym else "Unknown"
            if quarter and q != quarter:
                continue
            site_quarter[site][q] += lsl
            site_totals[site] += lsl

        per_site = []
        for site_code in sorted(site_totals.keys()):
            quarters_data = {
                q: round(v, 2)
                for q, v in sorted(site_quarter[site_code].items())
            }
            per_site.append({
                "site": site_code,
                "name": SITE_ABBREV.get(site_code, site_code),
                "total_lsl": round(site_totals[site_code], 2),
                "quarters": quarters_data,
            })

        return {
            "sites": per_site,
            "total_lsl": round(sum(site_totals.values()), 2),
            "source_table": source_table,
            "quarter_filter": quarter,
        }


# ---------------------------------------------------------------------------
# 6. Cumulative Consumption & Sales Trends (Figures 3, 4)
# ---------------------------------------------------------------------------

@router.get("/cumulative-trends")
def cumulative_trends(user: CurrentUser = Depends(require_employee)):
    """Quarterly cumulative consumption (kWh) and sales (LSL) over time."""
    with _get_connection() as conn:
        cursor = conn.cursor()

        try:
            cursor.execute(
                "SELECT transaction_date, kwh_value, transaction_amount "
                "FROM transactions"
            )
            rows = cursor.fetchall()
        except Exception as e:
            logger.warning("Failed to query transactions for cumulative: %s", e)
            return {"trends": [], "error": "No date column found in account history"}

        if not rows:
            return {"trends": [], "error": "No date column found in account history"}

        quarterly_kwh: Dict[str, float] = defaultdict(float)
        quarterly_lsl: Dict[str, float] = defaultdict(float)

        for row in rows:
            q = _date_to_quarter(row[0])
            if not q:
                continue
            quarterly_kwh[q] += float(row[1] or 0)
            quarterly_lsl[q] += float(row[2] or 0)

        sorted_quarters = sorted(set(quarterly_kwh.keys()) | set(quarterly_lsl.keys()))
        cum_kwh = 0.0
        cum_lsl = 0.0
        result = []
        for q in sorted_quarters:
            kwh = quarterly_kwh.get(q, 0)
            lsl = quarterly_lsl.get(q, 0)
            cum_kwh += max(kwh, 0)
            cum_lsl += max(lsl, 0)
            result.append({
                "quarter": q,
                "kwh": round(kwh, 2),
                "lsl": round(lsl, 2),
                "cumulative_kwh": round(cum_kwh, 2),
                "cumulative_lsl": round(cum_lsl, 2),
            })

        return {"trends": result, "source_table": "transactions"}


# ---------------------------------------------------------------------------
# 7. Average Consumption per Customer Trend (Figures 8, 9)
# ---------------------------------------------------------------------------

@router.get("/avg-consumption-trend")
def avg_consumption_trend(user: CurrentUser = Depends(require_employee)):
    """Average daily consumption and sales per customer per quarter."""
    with _get_connection() as conn:
        cursor = conn.cursor()

        cursor.execute(
            "SELECT date_service_connected FROM customers "
            "WHERE date_service_connected IS NOT NULL"
        )
        cust_rows = cursor.fetchall()

        quarterly_new: Dict[str, int] = defaultdict(int)
        for row in cust_rows:
            q = _date_to_quarter(row[0])
            if q:
                quarterly_new[q] += 1

        all_quarters = sorted(quarterly_new.keys())
        cum_customers: Dict[str, int] = {}
        cum = 0
        for q in all_quarters:
            cum += quarterly_new[q]
            cum_customers[q] = cum

        try:
            cursor.execute(
                "SELECT transaction_date, kwh_value, transaction_amount "
                "FROM transactions"
            )
            rows = cursor.fetchall()
        except Exception as e:
            logger.warning("Failed to query transactions for avg trend: %s", e)
            return {"trends": [], "error": "No data found"}

        if not rows:
            return {"trends": [], "error": "No data found"}

        quarterly_kwh: Dict[str, float] = defaultdict(float)
        quarterly_lsl: Dict[str, float] = defaultdict(float)
        quarterly_days: Dict[str, set] = defaultdict(set)

        for row in rows:
            q = _date_to_quarter(row[0])
            if not q:
                continue
            quarterly_kwh[q] += float(row[1] or 0)
            quarterly_lsl[q] += float(row[2] or 0)
            dk = _day_key(row[0])
            if dk:
                quarterly_days[q].add(dk)

        sorted_q = sorted(set(quarterly_kwh.keys()) & set(cum_customers.keys()))
        result = []
        for q in sorted_q:
            customers = cum_customers.get(q, 1)
            days = len(quarterly_days.get(q, set())) or 90
            kwh = quarterly_kwh.get(q, 0)
            lsl = quarterly_lsl.get(q, 0)

            result.append({
                "quarter": q,
                "customers": customers,
                "total_kwh": round(kwh, 2),
                "total_lsl": round(lsl, 2),
                "avg_daily_kwh_per_customer": round(kwh / (customers * days), 4) if customers > 0 else 0,
                "avg_daily_lsl_per_customer": round(lsl / (customers * days), 4) if customers > 0 else 0,
            })

        return {"trends": result, "source_table": "transactions"}


# ---------------------------------------------------------------------------
# 8. Site Overview with Districts (Tables 1, 2, 3)
# ---------------------------------------------------------------------------

@router.get("/site-overview")
def site_overview(user: CurrentUser = Depends(require_employee)):
    """List of all communities with customer counts and district info."""
    with _get_connection() as conn:
        cursor = conn.cursor()

        cursor.execute(
            "SELECT community, COUNT(*) AS cnt "
            "FROM customers "
            "WHERE community IS NOT NULL AND community <> '' "
            "GROUP BY community "
            "ORDER BY community"
        )
        rows = cursor.fetchall()

        sites = []
        for row in rows:
            name = str(row[0]).strip().upper()
            if name not in KNOWN_SITES:
                continue
            count = row[1]
            sites.append({
                "concession": name,
                "abbreviation": name,
                "district": SITE_DISTRICTS.get(name, ""),
                "customer_count": count,
            })

        return {"sites": sites}


# ---------------------------------------------------------------------------
# 9. Load Curves by Customer Type
# ---------------------------------------------------------------------------

@router.get("/load-curves-by-type")
def load_curves_by_type(
    quarter: Optional[str] = Query(None, description="Filter to quarter, e.g. '2025 Q4'"),
    user: CurrentUser = Depends(require_employee),
):
    """
    Average daily consumption per customer type.
    Joins meters table (customer type + account number) with transactions.
    """
    with _get_connection() as conn:
        cursor = conn.cursor()

        # 1. Build account -> customer_type mapping from customers table
        acct_type: Dict[str, str] = {}
        cursor.execute(_ACCT_CTYPE_SQL)
        for row in cursor.fetchall():
            acct = str(row[0] or "").strip()
            ctype = str(row[1] or "").strip()
            if acct and ctype:
                acct_type[acct] = ctype

        if not acct_type:
            return {
                "curves": [],
                "quarterly": [],
                "note": "No customer type data found.",
            }

        # 2. Query transactions
        try:
            cursor.execute(
                "SELECT account_number, transaction_date, kwh_value, "
                "transaction_amount FROM transactions"
            )
            rows = cursor.fetchall()
        except Exception as e:
            logger.warning("Failed to query transactions for load curves: %s", e)
            return {"curves": [], "quarterly": [], "error": "No account history data found"}

        if not rows:
            return {"curves": [], "quarterly": [], "error": "No account history data found"}

        type_totals: Dict[str, Dict[str, Any]] = defaultdict(
            lambda: {"kwh": 0.0, "lsl": 0.0, "customers": set(), "days": set()}
        )
        type_quarter: Dict[str, Dict[str, Dict[str, float]]] = defaultdict(
            lambda: defaultdict(lambda: {"kwh": 0.0, "lsl": 0.0})
        )

        for row in rows:
            acct = str(row[0] or "").strip()
            ctype = acct_type.get(acct)
            if not ctype:
                continue

            kwh = float(row[2] or 0)
            lsl = float(row[3] or 0)
            q = _date_to_quarter(row[1]) if row[1] else "Unknown"

            if quarter and q != quarter:
                continue

            type_totals[ctype]["kwh"] += kwh
            type_totals[ctype]["lsl"] += lsl
            type_totals[ctype]["customers"].add(acct)
            dk = _day_key(row[1])
            if dk:
                type_totals[ctype]["days"].add(dk)

            type_quarter[ctype][q]["kwh"] += kwh
            type_quarter[ctype][q]["lsl"] += lsl

        curves = []
        for ctype in sorted(type_totals.keys()):
            data = type_totals[ctype]
            n_customers = len(data["customers"])
            n_days = len(data["days"]) or 90
            curves.append({
                "type": ctype,
                "total_kwh": round(data["kwh"], 2),
                "total_lsl": round(data["lsl"], 2),
                "customer_count": n_customers,
                "avg_daily_kwh": round(data["kwh"] / n_days, 4) if n_days > 0 else 0,
                "avg_daily_kwh_per_customer": round(
                    data["kwh"] / (n_customers * n_days), 4
                ) if n_customers > 0 and n_days > 0 else 0,
            })

        all_quarters = sorted(
            set(q for tq in type_quarter.values() for q in tq.keys())
        )
        quarterly = []
        for q in all_quarters:
            entry: Dict[str, Any] = {"quarter": q}
            for ctype in sorted(type_totals.keys()):
                entry[ctype] = round(type_quarter[ctype].get(q, {}).get("kwh", 0), 2)
            quarterly.append(entry)

        return {
            "curves": curves,
            "quarterly": quarterly,
            "customer_types": sorted(type_totals.keys()),
            "total_typed_customers": len(acct_type),
            "source_table": "transactions",
            "meter_source": "meters",
            "quarter_filter": quarter,
        }


# ---------------------------------------------------------------------------
# 10. 24-Hour Daily Load Profiles by Customer Type
# ---------------------------------------------------------------------------
#
# Uses meter_readings (10-minute interval readings: reading_time, power_kw,
# meter_id) joined with meters (meter_id -> customer type) to build average
# hourly power curves for each customer type.

@router.get("/daily-load-profiles")
def daily_load_profiles(
    site: Optional[str] = Query(None, description="Filter to site code (e.g. MAK)"),
    customer_type: Optional[str] = Query(None, description="Filter to customer type (e.g. HH, SME)"),
    user: CurrentUser = Depends(require_employee),
):
    """
    Average 24-hour load profiles by customer type.

    Data source priority:
      1. meter_readings (10-min interval power_kw) — LS via ThunderCloud
      2. hourly_consumption (hourly kWh from Koios CSV) — BN fallback
    """
    with _get_connection() as conn:
        cursor = conn.cursor()

        from country_config import UTC_OFFSET_HOURS

        # 1. Build account → customer_type from customers table, then
        #    extend to meter_id via meters.account_number
        acct_type: Dict[str, str] = {}
        meter_type: Dict[str, str] = {}

        site_filter_sql = ""
        site_params: tuple = ()
        if site:
            site_filter_sql = " AND a.account_number LIKE %s"
            site_params = (f"%{site.upper()}",)

        cursor.execute(
            "SELECT a.account_number, c.customer_type "
            "FROM accounts a "
            "JOIN customers c ON a.customer_id = c.id "
            "WHERE c.customer_type IS NOT NULL AND c.customer_type <> ''"
            + site_filter_sql,
            site_params,
        )
        for row in cursor.fetchall():
            acct = str(row[0] or "").strip()
            ctype = str(row[1] or "").strip()
            if customer_type and not _matches_customer_type(ctype, customer_type):
                continue
            if acct:
                acct_type[acct] = ctype

        # Map meter_ids to customer types via account_number
        if acct_type:
            cursor.execute("SELECT meter_id, account_number FROM meters")
            for row in cursor.fetchall():
                mid = str(row[0] or "").strip()
                acct = str(row[1] or "").strip()
                if mid and acct in acct_type:
                    meter_type[mid] = acct_type[acct]

        if not meter_type and not acct_type:
            return {
                "profiles": [],
                "chart_data": [],
                "customer_types": [],
                "note": "No customer type data found in meters table.",
            }

        # 2. Try meter_readings first (10-min interval power data)
        type_hour_kw: Dict[str, Dict[int, List[float]]] = defaultdict(
            lambda: defaultdict(list)
        )
        type_meter_count: Dict[str, set] = defaultdict(set)
        total_readings = 0
        data_source = "meter_readings"

        try:
            cursor.execute(
                "SELECT meter_id, reading_time, power_kw, account_number, source "
                "FROM meter_readings "
                "WHERE power_kw IS NOT NULL" + (
                    " AND community = %s" if site else ""
                ),
                (site.upper(),) if site else (),
            )

            for row in cursor.fetchall():
                mid = str(row[0] or "").strip()
                ctype = meter_type.get(mid)
                if not ctype:
                    acct = str(row[3] or "").strip()
                    ctype = acct_type.get(acct) if acct else None
                if not ctype:
                    continue
                dt = row[1]
                kw = row[2]
                source_name = row[4]
                if dt is None or kw is None:
                    continue
                kw_val = _normalize_power_kw_value(kw, ctype, source_name)
                if kw_val is None:
                    continue
                try:
                    if hasattr(dt, 'hour'):
                        local_dt = dt.replace(tzinfo=None) + timedelta(hours=UTC_OFFSET_HOURS) if hasattr(dt, 'tzinfo') and dt.tzinfo else dt + timedelta(hours=UTC_OFFSET_HOURS)
                        hour = local_dt.hour
                    elif isinstance(dt, str):
                        hour = (int(dt.split(" ")[1].split(":")[0]) + UTC_OFFSET_HOURS) % 24
                    else:
                        continue
                except (IndexError, ValueError, AttributeError):
                    continue
                type_hour_kw[ctype][hour].append(kw_val)
                type_meter_count[ctype].add(mid)
                total_readings += 1

        except Exception as e:
            logger.warning("meter_readings query failed: %s", e)

        # 3. Also pull hourly_consumption for accounts not already covered
        #    by meter_readings. This is the primary data source for most LS
        #    accounts and the sole source for BN.
        mr_accounts = set()
        for accts in type_meter_count.values():
            mr_accounts.update(accts)

        hc_source = False
        try:
            cursor.execute(
                "SELECT account_number, reading_hour, kwh FROM hourly_consumption "
                "WHERE kwh IS NOT NULL AND kwh > 0"
                + (" AND community = %s" if site else ""),
                (site.upper(),) if site else (),
            )
            for row in cursor.fetchall():
                acct = str(row[0] or "").strip()
                if acct in mr_accounts:
                    continue
                ctype = acct_type.get(acct)
                if not ctype:
                    continue
                dt = row[1]
                kwh = row[2]
                if dt is None or kwh is None:
                    continue
                try:
                    kw_val = float(kwh)
                except (ValueError, TypeError):
                    continue
                try:
                    if hasattr(dt, 'hour'):
                        local_dt = dt.replace(tzinfo=None) + timedelta(hours=UTC_OFFSET_HOURS) if hasattr(dt, 'tzinfo') and dt.tzinfo else dt + timedelta(hours=UTC_OFFSET_HOURS)
                        hour = local_dt.hour
                    elif isinstance(dt, str):
                        hour = (int(dt.split(" ")[1].split(":")[0]) + UTC_OFFSET_HOURS) % 24
                    else:
                        continue
                except (IndexError, ValueError, AttributeError):
                    continue
                type_hour_kw[ctype][hour].append(kw_val)
                type_meter_count[ctype].add(acct)
                total_readings += 1
                hc_source = True
        except Exception as e:
            logger.warning("hourly_consumption query failed: %s", e)

        if hc_source:
            data_source = ("meter_readings+hourly_consumption"
                           if data_source == "meter_readings"
                           else "hourly_consumption")

        if not type_hour_kw:
            return {
                "profiles": [],
                "chart_data": [],
                "customer_types": [],
                "note": "No meter reading data found.",
            }

        # 4. Build 24-hour profiles
        profiles = []
        for ctype in sorted(type_hour_kw.keys()):
            hourly = []
            for h in range(24):
                readings = type_hour_kw[ctype].get(h, [])
                avg_kw = sum(readings) / len(readings) if readings else 0
                hourly.append({
                    "hour": h,
                    "avg_kw": round(avg_kw, 4),
                    "readings": len(readings),
                })

            profiles.append({
                "type": ctype,
                "meter_count": len(type_meter_count[ctype]),
                "hourly": hourly,
                "peak_hour": max(range(24), key=lambda h: sum(type_hour_kw[ctype].get(h, [0])) / max(len(type_hour_kw[ctype].get(h, [1])), 1)),
                "peak_kw": round(max(
                    sum(type_hour_kw[ctype].get(h, [0])) / max(len(type_hour_kw[ctype].get(h, [1])), 1)
                    for h in range(24)
                ), 4),
            })

        chart_data = []
        for h in range(24):
            point: Dict[str, Any] = {"hour": f"{h:02d}:00"}
            for ctype in sorted(type_hour_kw.keys()):
                readings = type_hour_kw[ctype].get(h, [])
                point[ctype] = round(sum(readings) / len(readings), 4) if readings else 0
            chart_data.append(point)

        return {
            "profiles": profiles,
            "chart_data": chart_data,
            "customer_types": sorted(type_hour_kw.keys()),
            "total_readings": total_readings,
            "data_source": data_source,
            "site_filter": site,
            "customer_type_filter": customer_type,
        }


# ---------------------------------------------------------------------------
# 11. ARPU (Average Revenue Per User) Time Series
# ---------------------------------------------------------------------------

@router.get("/arpu")
def arpu_time_series(user: CurrentUser = Depends(require_employee)):
    """
    Quarterly ARPU: total revenue / cumulative customer base per quarter.

    "Active customers" = all distinct account numbers that have ever
    transacted up to and including the quarter.  This produces a
    monotonically-increasing customer count that reflects the growing
    customer base, and divides quarterly revenue by that base.

    Data source priority:
      1. monthly_transactions (SparkMeter portfolio data, includes manual corrections)
      2. transactions (raw history, fallback)
    """

    with _get_connection() as conn:
        cursor = conn.cursor()

        txn_rows: List[tuple] = []
        source_table = ""

        try:
            cursor.execute(
                "SELECT account_number, year_month, amount_lsl, community "
                "FROM monthly_transactions"
            )
            raw = cursor.fetchall()
            if raw:
                for row in raw:
                    acct = str(row[0] or "").strip()
                    ym = str(row[1] or "").strip()
                    lsl = float(row[2] or 0)
                    community = str(row[3] or "").strip().upper()
                    if not acct or not ym or lsl <= 0:
                        continue
                    try:
                        y, m = int(ym[:4]), int(ym[5:7])
                        dt = datetime(y, m, 15)
                    except (ValueError, IndexError):
                        continue
                    txn_rows.append((acct, dt, lsl, community))
                if txn_rows:
                    source_table = "monthly_transactions"
        except Exception as e:
            logger.warning("Failed to query monthly_transactions for ARPU: %s", e)

        if not txn_rows:
            try:
                cursor.execute(
                    "SELECT account_number, transaction_date, transaction_amount "
                    "FROM transactions"
                )
                raw = cursor.fetchall()
                if raw:
                    txn_rows = [
                        (str(r[0] or "").strip(), r[1], float(r[2] or 0), "")
                        for r in raw
                    ]
                    source_table = "transactions"
            except Exception:
                pass

        if txn_rows:
            q_revenue: Dict[str, float] = defaultdict(float)
            q_site_revenue: Dict[str, Dict[str, float]] = defaultdict(
                lambda: defaultdict(float)
            )
            acct_first_quarter: Dict[str, str] = {}
            acct_site: Dict[str, str] = {}

            for row in txn_rows:
                acct = str(row[0] or "").strip()
                q = _date_to_quarter(row[1])
                lsl = float(row[2] or 0)
                community = str(row[3] or "").strip().upper() if len(row) > 3 else ""
                if not q or not acct:
                    continue
                site = community or _extract_site(acct)

                q_revenue[q] += lsl
                if site and len(site) >= 2:
                    q_site_revenue[q][site] += lsl

                if acct not in acct_first_quarter or q < acct_first_quarter[acct]:
                    acct_first_quarter[acct] = q
                    if site and len(site) >= 2:
                        acct_site[acct] = site

            all_quarters = sorted(q_revenue.keys())
            cumulative_all: set = set()
            cumulative_by_site: Dict[str, set] = defaultdict(set)

            result = []
            for q in all_quarters:
                for acct, first_q in acct_first_quarter.items():
                    if first_q <= q:
                        cumulative_all.add(acct)
                        site = acct_site.get(acct, "")
                        if site:
                            cumulative_by_site[site].add(acct)

                revenue = q_revenue[q]
                active = len(cumulative_all)
                arpu = round(revenue / active, 2) if active > 0 else 0

                per_site = {}
                for site_code in sorted(q_site_revenue[q]):
                    site_rev = q_site_revenue[q][site_code]
                    site_custs = len(cumulative_by_site.get(site_code, set()))
                    per_site[site_code] = {
                        "name": SITE_ABBREV.get(site_code, site_code),
                        "revenue": round(site_rev, 2),
                        "customers": site_custs,
                        "arpu": round(site_rev / site_custs, 2) if site_custs > 0 else 0,
                    }

                result.append({
                    "quarter": q,
                    "total_revenue": round(revenue, 2),
                    "active_customers": active,
                    "arpu": arpu,
                    "per_site": per_site,
                })

            all_site_codes = sorted(
                set(code for entry in result for code in entry["per_site"])
            )

            return {
                "arpu": result,
                "site_codes": all_site_codes,
                "site_names": {c: SITE_ABBREV.get(c, c) for c in all_site_codes},
                "source_table": source_table,
            }

        return {"arpu": [], "site_codes": [], "error": "No account history data found"}


# ---------------------------------------------------------------------------
# 12. Monthly ARPU Time Series
# ---------------------------------------------------------------------------

def _date_to_month(dt) -> str:
    """Convert a date/datetime to 'YYYY-MM' string."""
    if dt is None:
        return ""
    if isinstance(dt, str):
        for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d", "%m/%d/%Y", "%d/%m/%Y"):
            try:
                dt = datetime.strptime(dt.strip(), fmt)
                break
            except (ValueError, AttributeError):
                continue
        else:
            return ""
    try:
        return f"{dt.year}-{dt.month:02d}"
    except (AttributeError, TypeError):
        return ""


@router.get("/monthly-arpu")
def monthly_arpu_time_series(user: CurrentUser = Depends(require_employee)):
    """
    Monthly ARPU: total revenue / cumulative customer base per month.

    Data source priority:
      1. monthly_transactions (SparkMeter portfolio, includes corrections)
      2. transactions (raw history, fallback)
    """
    with _get_connection() as conn:
        cursor = conn.cursor()

        txn_rows: List[tuple] = []
        source_table = ""

        try:
            cursor.execute(
                "SELECT account_number, year_month, amount_lsl, community "
                "FROM monthly_transactions"
            )
            raw = cursor.fetchall()
            if raw:
                for row in raw:
                    acct = str(row[0] or "").strip()
                    ym = str(row[1] or "").strip()
                    lsl = float(row[2] or 0)
                    community = str(row[3] or "").strip().upper()
                    if acct and ym and lsl > 0:
                        txn_rows.append((acct, ym, lsl, community))
                if txn_rows:
                    source_table = "monthly_transactions"
        except Exception as e:
            logger.warning("monthly_transactions query failed: %s", e)

        if not txn_rows:
            try:
                cursor.execute(
                    "SELECT account_number, transaction_date, transaction_amount "
                    "FROM transactions"
                )
                raw = cursor.fetchall()
                if raw:
                    for r in raw:
                        acct = str(r[0] or "").strip()
                        m = _date_to_month(r[1])
                        lsl = float(r[2] or 0)
                        if acct and m:
                            txn_rows.append((acct, m, lsl, ""))
                    if txn_rows:
                        source_table = "transactions"
            except Exception:
                pass

        if not txn_rows:
            return {"monthly_arpu": [], "site_codes": [], "error": "No transaction data found"}

        m_revenue: Dict[str, float] = defaultdict(float)
        m_site_revenue: Dict[str, Dict[str, float]] = defaultdict(
            lambda: defaultdict(float)
        )
        acct_first_month: Dict[str, str] = {}
        acct_site: Dict[str, str] = {}

        for acct, m, lsl, community in txn_rows:
            site = (community.upper() if community else "") or _extract_site(acct)
            m_revenue[m] += lsl
            if site and len(site) >= 2:
                m_site_revenue[m][site] += lsl
            if acct not in acct_first_month or m < acct_first_month[acct]:
                acct_first_month[acct] = m
                if site and len(site) >= 2:
                    acct_site[acct] = site

        all_months = sorted(m_revenue.keys())
        cumulative_all: set = set()
        cumulative_by_site: Dict[str, set] = defaultdict(set)

        result = []
        for m in all_months:
            for acct, first_m in acct_first_month.items():
                if first_m <= m:
                    cumulative_all.add(acct)
                    site = acct_site.get(acct, "")
                    if site:
                        cumulative_by_site[site].add(acct)

            revenue = m_revenue[m]
            active = len(cumulative_all)
            arpu = round(revenue / active, 2) if active > 0 else 0

            per_site = {}
            for site_code in sorted(m_site_revenue[m]):
                site_rev = m_site_revenue[m][site_code]
                site_custs = len(cumulative_by_site.get(site_code, set()))
                per_site[site_code] = {
                    "name": SITE_ABBREV.get(site_code, site_code),
                    "revenue": round(site_rev, 2),
                    "customers": site_custs,
                    "arpu": round(site_rev / site_custs, 2) if site_custs > 0 else 0,
                }

            result.append({
                "month": m,
                "quarter": _date_to_quarter_from_month(m),
                "total_revenue": round(revenue, 2),
                "active_customers": active,
                "arpu": arpu,
                "per_site": per_site,
            })

        all_site_codes = sorted(
            set(code for entry in result for code in entry["per_site"])
        )

        return {
            "monthly_arpu": result,
            "site_codes": all_site_codes,
            "site_names": {c: SITE_ABBREV.get(c, c) for c in all_site_codes},
            "source_table": source_table,
        }


def _date_to_quarter_from_month(month_str: str) -> str:
    """Convert 'YYYY-MM' to 'YYYY QN'."""
    try:
        y, m = month_str.split("-")
        q = (int(m) - 1) // 3 + 1
        return f"{y} Q{q}"
    except (ValueError, AttributeError):
        return ""


# ---------------------------------------------------------------------------
# 13. Average Consumption by Tenure (months since first transaction)
# ---------------------------------------------------------------------------

@router.get("/consumption-by-tenure")
def consumption_by_tenure(
    user: CurrentUser = Depends(require_employee),
):
    """
    Average monthly kWh consumption as a function of tenure (months since
    first reading), segmented by customer type (HH, SME, etc.)
    with +/- 1 standard deviation bands.

    Data source: monthly_consumption table only (actual meter readings).
    This excludes transaction/vending data to avoid conflating purchased kWh
    with consumed kWh.

    Customer type is resolved from the meters table first, then from the
    static JSON mapping (meter_customer_types.json) as a fallback.
    """

    # -- Load meter -> customer-type mapping (JSON fallback) --
    _data_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), "data")
    _type_map_path = os.path.join(_data_dir, "meter_customer_types.json")
    try:
        with open(_type_map_path, "r") as f:
            json_type_map: Dict[str, str] = json.load(f)
    except Exception as e:
        logger.error("Cannot load meter_customer_types.json: %s", e)
        json_type_map = {}

    norm_map: Dict[str, str] = {}
    for mid, ctype in json_type_map.items():
        norm_map[mid.upper().replace("_", "-")] = ctype

    def _parse_dt(dt) -> Optional[datetime]:
        return _coerce_export_timestamp(dt)

    def _lookup_type(meter_id: str) -> Optional[str]:
        if not meter_id:
            return None
        key = meter_id.strip().upper().replace("_", "-")
        return norm_map.get(key)

    with _get_connection() as conn:
        cursor = conn.cursor()

        # -- Build account -> customer_type from customers table --
        acct_type: Dict[str, str] = {}
        cursor.execute(_ACCT_CTYPE_SQL)
        for row in cursor.fetchall():
            acct = str(row[0] or "").strip()
            ctype = str(row[1] or "").strip()
            if acct and ctype:
                acct_type[acct] = ctype

        # -- Fetch meter -> account mapping --
        cursor.execute("SELECT meter_id, account_number FROM meters")
        all_meter_rows = cursor.fetchall()

        meter_type_map: Dict[str, str] = {}
        meter_to_acct: Dict[str, str] = {}

        for row in all_meter_rows:
            mid = str(row[0] or "").strip()
            acct = str(row[1] or "").strip()

            if mid and acct:
                meter_to_acct[mid] = acct
                meter_to_acct[mid.upper()] = acct

            ctype = acct_type.get(acct, "")
            if mid and ctype:
                meter_type_map[mid] = ctype
                meter_type_map[mid.upper()] = ctype

        # Enrich acct_type for accounts without a direct type, via JSON
        for row in all_meter_rows:
            mid = str(row[0] or "").strip()
            acct = str(row[1] or "").strip()
            if not acct or acct in acct_type:
                continue
            ctype_json = _lookup_type(mid)
            if ctype_json:
                acct_type[acct] = ctype_json

        acct_type_lower: Dict[str, str] = {k.lower(): v for k, v in acct_type.items()}

        def _resolve_type(acct: str, meterid: str = "") -> Optional[str]:
            """Multi-strategy customer type resolution."""
            ct = acct_type.get(acct) or acct_type_lower.get(acct.lower())
            if ct:
                return ct
            if meterid:
                ct = meter_type_map.get(meterid) or meter_type_map.get(meterid.upper())
                if ct:
                    return ct
                ct = _lookup_type(meterid)
                if ct:
                    return ct
            ct = meter_type_map.get(acct) or meter_type_map.get(acct.upper())
            if ct:
                return ct
            return _lookup_type(acct)

        parsed_rows: List[tuple] = []
        acct_first_txn: Dict[str, datetime] = {}
        debug_info: Dict[str, Any] = {"acct_type_map_size": len(acct_type)}

        # -- monthly_consumption (actual meter readings) --
        consumption_rows: list = []

        try:
            cursor.execute(
                "SELECT account_number, year_month, kwh, meter_id "
                "FROM monthly_consumption"
            )
            consumption_rows = cursor.fetchall()
        except Exception as e:
            logger.warning("monthly_consumption query failed: %s", e)

        cons_matched = 0
        cons_unmatched = 0
        cons_added = 0

        for row in consumption_rows:
            raw_acct = str(row[0] or "").strip()
            ym = str(row[1] or "").strip()
            try:
                kwh = float(row[2] or 0)
            except (ValueError, TypeError):
                continue
            meterid = str(row[3] or "").strip() if len(row) > 3 else ""
            if not raw_acct or not ym or kwh <= 0:
                continue

            acct = (meter_to_acct.get(raw_acct)
                    or meter_to_acct.get(raw_acct.upper())
                    or raw_acct)
            ctype = _resolve_type(acct, meterid or raw_acct)
            if not ctype:
                cons_unmatched += 1
                continue
            cons_matched += 1

            try:
                y, m = int(ym[:4]), int(ym[5:7])
                dt = datetime(y, m, 1)
            except (ValueError, IndexError):
                continue

            parsed_rows.append((acct, ctype, dt, kwh))
            cons_added += 1
            if acct not in acct_first_txn or dt < acct_first_txn[acct]:
                acct_first_txn[acct] = dt

        debug_info["consumption"] = {
            "total_rows": len(consumption_rows),
            "matched": cons_matched,
            "unmatched": cons_unmatched,
            "added": cons_added,
        }
        debug_info["total_unique_accounts"] = len(acct_first_txn)

        data_source = "metered"

        if not acct_first_txn:
            return {
                "chart_data": [], "customer_types": [],
                "data_source": data_source,
                "error": "No account data found",
                "debug": debug_info,
            }

        # -- Aggregate by type -> tenure_month -> acct --
        type_tenure_acct: Dict[str, Dict[int, Dict[str, float]]] = defaultdict(
            lambda: defaultdict(lambda: defaultdict(float))
        )

        for acct, ctype, txn_dt, kwh in parsed_rows:
            if kwh <= 0:
                continue
            first_dt = acct_first_txn[acct]
            tenure_months = (
                (txn_dt.year - first_dt.year) * 12
                + (txn_dt.month - first_dt.month)
            )
            if tenure_months < 0:
                continue
            type_tenure_acct[ctype][tenure_months][acct] += kwh

        all_types = sorted(type_tenure_acct.keys())
        if not all_types:
            return {
                "chart_data": [], "customer_types": [],
                "data_source": data_source,
                "error": "No typed tenure data after aggregation",
                "debug": debug_info,
            }

        now = datetime.now()
        type_acct_tenure: Dict[str, Dict[str, int]] = defaultdict(dict)
        max_tenure = 0
        for acct, ctype, _dt, _kwh in parsed_rows:
            if acct in type_acct_tenure.get(ctype, {}):
                continue
            first_dt = acct_first_txn[acct]
            tenure = (now.year - first_dt.year) * 12 + (now.month - first_dt.month)
            if tenure < 0:
                continue
            type_acct_tenure[ctype][acct] = tenure
            if tenure > max_tenure:
                max_tenure = tenure

        chart_data = []
        last_valid_t = 0
        for t in range(max_tenure + 1):
            point: Dict[str, Any] = {"tenure_month": t}
            has_any_data = False
            for ctype in all_types:
                n_eligible = sum(
                    1 for ten in type_acct_tenure.get(ctype, {}).values()
                    if ten >= t
                )
                acct_kwh = type_tenure_acct[ctype].get(t, {})
                raw_values = sorted(acct_kwh.values())
                if len(raw_values) >= 4:
                    q1_idx = len(raw_values) // 4
                    q3_idx = 3 * len(raw_values) // 4
                    q1 = raw_values[q1_idx]
                    q3 = raw_values[q3_idx]
                    iqr = q3 - q1
                    lo = q1 - 1.5 * iqr
                    hi = q3 + 1.5 * iqr
                    values = [v for v in raw_values if lo <= v <= hi]
                else:
                    values = raw_values
                if len(values) < 3:
                    point[ctype] = None
                    point[f"{ctype}_upper"] = None
                    point[f"{ctype}_lower"] = None
                    point[f"{ctype}_n"] = n_eligible
                    point[f"{ctype}_nd"] = len(raw_values)
                else:
                    has_any_data = True
                    n_data = len(values)
                    mean = sum(values) / n_data
                    if n_data > 1:
                        variance = sum((v - mean) ** 2 for v in values) / n_data
                        sd = math.sqrt(variance)
                    else:
                        sd = 0.0
                    point[ctype] = round(mean, 2)
                    point[f"{ctype}_upper"] = round(mean + sd, 2)
                    point[f"{ctype}_lower"] = round(max(mean - sd, 0), 2)
                    point[f"{ctype}_n"] = n_eligible
                    point[f"{ctype}_nd"] = len(raw_values)
                    point[f"{ctype}_nf"] = len(raw_values) - n_data
                    point[f"{ctype}_min"] = round(min(values), 2)
                    point[f"{ctype}_max"] = round(max(values), 2)
            chart_data.append(point)
            if has_any_data:
                last_valid_t = t

        chart_data = chart_data[: last_valid_t + 1]
        max_tenure = last_valid_t

        type_stats = []
        for ctype in all_types:
            all_accts: set = set()
            total_kwh = 0.0
            for t_data in type_tenure_acct[ctype].values():
                all_accts.update(t_data.keys())
                total_kwh += sum(t_data.values())
            max_t = (
                max(type_tenure_acct[ctype].keys())
                if type_tenure_acct[ctype]
                else 0
            )
            type_stats.append({
                "type": ctype,
                "customer_count": len(all_accts),
                "total_kwh": round(total_kwh, 2),
                "max_tenure_months": max_t,
            })

        debug_info["type_acct_tenure_counts"] = {
            ct: len(accts) for ct, accts in type_acct_tenure.items()
        }

        return {
            "chart_data": chart_data,
            "customer_types": all_types,
            "type_stats": type_stats,
            "max_tenure_months": max_tenure,
            "total_accounts_matched": len(acct_first_txn),
            "data_source": data_source,
            "segmentation": "customer_type",
            "mapping_size": len(json_type_map),
            "debug": debug_info,
        }


# ---------------------------------------------------------------------------
# 14. Meter Data Export (for CDF building)
# ---------------------------------------------------------------------------

@router.get("/meter-export")
def meter_data_export(
    customer_type: Optional[str] = Query(None, description="Filter by customer type (e.g. HH, SME, SCH)"),
    site: Optional[str] = Query(None, description="Filter by site code (e.g. MAK)"),
    start_date: Optional[str] = Query(None, description="Start date YYYY-MM-DD"),
    end_date: Optional[str] = Query(None, description="End date YYYY-MM-DD"),
    user: CurrentUser = Depends(require_employee),
):
    """
    Export raw meter readings from meter_readings for CDF generation.

    Returns timestamped kW readings joined with the meters table to
    include customer type and site.  Designed for batch consumption by
    the uGridPlan 8760 CDF builder script.

    Response: {readings: [{timestamp, kw, customer_type, site, meterid}, ...], meta: {...}}
    """
    with _get_connection() as conn:
        cursor = conn.cursor()

        valid_start_date = None
        valid_end_date = None
        start_dt: Optional[datetime] = None
        end_exclusive_dt: Optional[datetime] = None
        if start_date:
            try:
                start_dt = datetime.strptime(start_date, "%Y-%m-%d")
                valid_start_date = start_date
            except ValueError:
                logger.warning("meter-export ignoring invalid start_date: %s", start_date)
        if end_date:
            try:
                end_dt = datetime.strptime(end_date, "%Y-%m-%d")
                end_exclusive_dt = end_dt + timedelta(days=1)
                valid_end_date = end_date
            except ValueError:
                logger.warning("meter-export ignoring invalid end_date: %s", end_date)

        # -- 1. Mirror daily_load_profiles() type resolution --
        # Build account -> customer_type from customers first, then extend to
        # meter_id via meters.account_number. This lets raw rows fall back to
        # account_number when the stored meter_id does not exactly match the
        # registry key.
        acct_type: Dict[str, str] = {}
        meter_type: Dict[str, str] = {}
        acct_site: Dict[str, str] = {}
        acct_meter: Dict[str, str] = {}
        meter_site: Dict[str, str] = {}

        site_filter_sql = ""
        site_params: tuple = ()
        if site:
            site_filter_sql = " AND a.account_number LIKE %s"
            site_params = (f"%{site.upper()}",)

        cursor.execute(
            "SELECT a.account_number, c.customer_type "
            "FROM accounts a "
            "JOIN customers c ON a.customer_id = c.id "
            "WHERE c.customer_type IS NOT NULL AND c.customer_type <> ''"
            + site_filter_sql,
            site_params,
        )
        for row in cursor.fetchall():
            acct = str(row[0] or "").strip()
            ctype = str(row[1] or "").strip().upper()
            if customer_type and not _matches_customer_type(ctype, customer_type):
                continue
            if acct:
                acct_type[acct] = ctype

        cursor.execute(
            "SELECT meter_id, account_number, community, role, status "
            "FROM meters "
            "ORDER BY account_number, "
            "CASE WHEN role = 'primary' THEN 0 ELSE 1 END, "
            "CASE WHEN status = 'active' THEN 0 ELSE 1 END, "
            "meter_id"
        )
        for row in cursor.fetchall():
            mid = str(row[0] or "").strip()
            acct = str(row[1] or "").strip()
            community = str(row[2] or "").strip().upper()
            if acct and community and acct not in acct_site:
                acct_site[acct] = community
            if acct and mid and acct not in acct_meter:
                acct_meter[acct] = mid
            if mid and community:
                meter_site[mid] = community
            if mid and acct in acct_type:
                meter_type[mid] = acct_type[acct]

        if not meter_type and not acct_type:
            return {"readings": [], "meta": {"error": "No typed customer data found"}}

        # -- 2. Query meter_readings --
        sql = (
            "SELECT meter_id, reading_time, power_kw, account_number, source, community "
            "FROM meter_readings "
            "WHERE power_kw IS NOT NULL"
        )
        params: List[Any] = []

        if site:
            sql += " AND community = %s"
            params.append(site.upper())
        if valid_start_date:
            sql += " AND reading_time >= %s::timestamp"
            params.append(valid_start_date)
        if valid_end_date:
            sql += " AND reading_time < (%s::date + 1)::timestamp"
            params.append(valid_end_date)

        try:
            cursor.execute(sql, params) if params else cursor.execute(sql)
        except Exception as e:
            logger.warning("meter-export: query failed: %s", e)
            return {"readings": [], "meta": {"error": str(e)}}

        # -- 3. Stream results with Python-side filtering --
        readings: List[Dict[str, Any]] = []
        raw_accounts_covered: Set[str] = set()
        skipped = 0
        resolved_by_meter_id = 0
        resolved_by_account = 0
        source_rows = {"meter_readings": 0, "hourly_consumption": 0}

        for row in cursor.fetchall():
            mid = str(row[0] or "").strip()
            acct = str(row[3] or "").strip()

            ctype = meter_type.get(mid)
            if ctype:
                resolved_by_meter_id += 1
            elif acct:
                ctype = acct_type.get(acct)
                if ctype:
                    resolved_by_account += 1

            if not ctype:
                skipped += 1
                continue

            community = (
                meter_site.get(mid)
                or acct_site.get(acct)
                or str(row[5] or "").strip().upper()
                or _extract_site(acct)
            )

            if customer_type and not _matches_customer_type(ctype, customer_type):
                continue

            dt_val = row[1]
            kw_val = row[2]
            source_name = row[4]
            if dt_val is None or kw_val is None:
                continue

            kw_float = _normalize_power_kw_value(kw_val, ctype, source_name)
            if kw_float is None:
                continue

            ts = _coerce_export_timestamp(dt_val)
            if ts is None:
                continue

            if start_dt and ts < start_dt:
                continue
            if end_exclusive_dt and ts >= end_exclusive_dt:
                continue

            readings.append({
                "timestamp": ts.strftime("%Y-%m-%d %H:%M:%S"),
                "kw": round(kw_float, 4),
                "customer_type": ctype,
                "site": community,
                "meterid": mid,
                "source_table": "meter_readings",
                "source": source_name,
            })
            source_rows["meter_readings"] += 1
            if acct:
                raw_accounts_covered.add(acct)

        skipped_hourly_no_type = 0
        skipped_hourly_covered = 0
        hourly_accounts_used: Set[str] = set()

        try:
            hourly_sql = (
                "SELECT account_number, meter_id, reading_hour, kwh, community, source "
                "FROM hourly_consumption "
                "WHERE kwh IS NOT NULL AND kwh > 0"
            )
            hourly_params: List[Any] = []
            if site:
                hourly_sql += " AND community = %s"
                hourly_params.append(site.upper())
            if valid_start_date:
                hourly_sql += " AND reading_hour >= %s::timestamp"
                hourly_params.append(valid_start_date)
            if valid_end_date:
                hourly_sql += " AND reading_hour < (%s::date + 1)::timestamp"
                hourly_params.append(valid_end_date)
            cursor.execute(hourly_sql, hourly_params)
            for row in cursor.fetchall():
                acct = str(row[0] or "").strip()
                if not acct:
                    skipped_hourly_no_type += 1
                    continue
                if acct in raw_accounts_covered:
                    skipped_hourly_covered += 1
                    continue

                ctype = acct_type.get(acct)
                if not ctype:
                    skipped_hourly_no_type += 1
                    continue

                dt_val = row[2]
                kwh_val = row[3]
                if dt_val is None or kwh_val is None:
                    continue
                try:
                    kw_float = float(kwh_val)
                except (ValueError, TypeError):
                    continue
                if not math.isfinite(kw_float) or kw_float <= 0:
                    continue

                ts = _coerce_export_timestamp(dt_val)
                if ts is None:
                    continue

                if start_dt and ts < start_dt:
                    continue
                if end_exclusive_dt and ts >= end_exclusive_dt:
                    continue

                meterid = str(row[1] or "").strip() or acct_meter.get(acct) or acct
                community = (
                    acct_site.get(acct)
                    or meter_site.get(meterid)
                    or str(row[4] or "").strip().upper()
                    or _extract_site(acct)
                )
                source_name = str(row[5] or "").strip().lower()

                readings.append({
                    "timestamp": ts.strftime("%Y-%m-%d %H:%M:%S"),
                    "kw": round(kw_float, 4),
                    "customer_type": ctype,
                    "site": community,
                    "meterid": meterid,
                    "source_table": "hourly_consumption",
                    "source": source_name,
                })
                source_rows["hourly_consumption"] += 1
                hourly_accounts_used.add(acct)
        except Exception as e:
            logger.warning("meter-export hourly fallback query failed: %s", e)

        # -- 4. Summary --
        type_counts: Dict[str, int] = defaultdict(int)
        site_counts: Dict[str, int] = defaultdict(int)
        for r in readings:
            type_counts[r["customer_type"]] += 1
            site_counts[r["site"]] += 1

        return {
            "readings": readings,
            "meta": {
                "total_readings": len(readings),
                "skipped_no_type": skipped,
                "skipped_hourly_no_type": skipped_hourly_no_type,
                "skipped_hourly_already_covered": skipped_hourly_covered,
                "meter_source": "meters+accounts",
                "customer_types": dict(type_counts),
                "sites": dict(site_counts),
                "source_rows": source_rows,
                "type_resolution": {
                    "meter_id": resolved_by_meter_id,
                    "account_number": resolved_by_account,
                },
                "raw_accounts_covered": len(raw_accounts_covered),
                "hourly_fallback_accounts": len(hourly_accounts_used),
                "filters": {
                    "customer_type": customer_type,
                    "site": site,
                    "start_date": start_date,
                    "end_date": end_date,
                },
            },
        }


# ---------------------------------------------------------------------------
# 15. Check Meter vs Primary Meter Comparison
# ---------------------------------------------------------------------------

def _build_check_meter_comparison(conn, days: int) -> Dict[str, Any]:
    """Build the check-meter comparison payload shared by JSON and export views."""
    cursor = conn.cursor()
    from country_config import UTC_OFFSET_HOURS

    cursor.execute("""
        SELECT m_check.account_number,
               m_check.meter_id   AS check_meter_id,
               m_primary.meter_id AS primary_meter_id
        FROM meters m_check
        JOIN meters m_primary
          ON m_primary.account_number = m_check.account_number
         AND m_primary.role = 'primary'
         AND m_primary.status = 'active'
        WHERE m_check.role = 'check'
          AND m_check.status = 'active'
          AND m_check.account_number IS NOT NULL
          AND m_check.account_number <> ''
    """)
    pairs: List[Dict[str, Any]] = []
    pair_accounts: List[str] = []
    for row in cursor.fetchall():
        pairs.append({
            "account": row[0],
            "check_meter_id": row[1],
            "primary_meter_id": row[2],
        })
        pair_accounts.append(row[0])

    if not pair_accounts:
        return {"pairs": [], "time_series": [], "days": days, "cutoff": "", "note": "No check meter pairs found"}

    if days > 0:
        cutoff = datetime.utcnow() - timedelta(days=days)
    else:
        placeholders_iot = ",".join(["%s"] * len(pair_accounts))
        cursor.execute(
            f"SELECT MIN(reading_hour) FROM hourly_consumption "
            f"WHERE account_number IN ({placeholders_iot}) AND source = 'iot'",
            tuple(pair_accounts),
        )
        row = cursor.fetchone()
        cutoff = row[0] if row and row[0] else datetime.utcnow() - timedelta(days=30)

    cutoff_for_compare = cutoff
    if isinstance(cutoff_for_compare, datetime):
        if cutoff_for_compare.tzinfo is None:
            cutoff_for_compare = cutoff_for_compare.replace(tzinfo=timezone.utc)
        visible_cutoff_key = (cutoff_for_compare + timedelta(hours=UTC_OFFSET_HOURS)).strftime("%Y-%m-%dT%H:%M:%S")
    else:
        visible_cutoff_key = str(cutoff_for_compare)

    # Pull a short lookback before the visible cutoff so we can detect the first
    # iot hour after long outages / reconnects inside the requested window.
    query_start = cutoff - timedelta(hours=24)
    placeholders = ",".join(["%s"] * len(pair_accounts))
    cursor.execute(
        f"SELECT account_number, meter_id, reading_hour, kwh, source "
        f"FROM hourly_consumption "
        f"WHERE account_number IN ({placeholders}) AND reading_hour >= %s "
        f"ORDER BY reading_hour",
        (*pair_accounts, query_start),
    )

    hour_data: Dict[str, Dict[str, Dict[str, Optional[float]]]] = defaultdict(
        lambda: defaultdict(lambda: {"sm": None, "1m": None})
    )
    iot_hour_meta: Dict[str, Dict[str, Dict[str, Any]]] = defaultdict(dict)
    for row in cursor.fetchall():
        acct = str(row[0] or "").strip()
        meter_id = str(row[1] or "").strip()
        hour = row[2]
        kwh = float(row[3]) if row[3] is not None else None
        source = str(row[4] or "").lower()

        if hasattr(hour, "strftime"):
            if hour.tzinfo is None:
                hour = hour.replace(tzinfo=timezone.utc)
            local_hour = hour + timedelta(hours=UTC_OFFSET_HOURS)
            hour_key = local_hour.strftime("%Y-%m-%dT%H:%M:%S")
        else:
            hour_key = str(hour)

        if source in ("thundercloud", "koios"):
            existing = hour_data[hour_key][acct]["sm"]
            hour_data[hour_key][acct]["sm"] = (existing or 0) + (kwh or 0)
        elif source == "iot":
            existing = hour_data[hour_key][acct]["1m"]
            hour_data[hour_key][acct]["1m"] = (existing or 0) + (kwh or 0)
            meta = iot_hour_meta[acct].setdefault(
                hour_key,
                {"utc_hour": hour, "meter_ids": set()},
            )
            meta["meter_ids"].add(meter_id)

    excluded_iot_hours: Dict[str, Set[str]] = defaultdict(set)
    for acct, hour_meta in iot_hour_meta.items():
        prev_hour: Optional[datetime] = None
        prev_meter_ids: Optional[Tuple[str, ...]] = None
        ordered = sorted(
            hour_meta.items(),
            key=lambda item: item[1]["utc_hour"] if item[1]["utc_hour"] is not None else item[0],
        )
        for hour_key, meta in ordered:
            utc_hour = meta["utc_hour"]
            meter_ids = tuple(sorted(str(mid) for mid in meta["meter_ids"] if mid))
            has_gap = (
                prev_hour is not None
                and utc_hour is not None
                and utc_hour - prev_hour > timedelta(hours=2)
            )
            meter_changed = (
                prev_meter_ids is not None
                and meter_ids
                and meter_ids != prev_meter_ids
            )
            mixed_meter_hour = len(meter_ids) > 1
            if has_gap or meter_changed or mixed_meter_hour:
                excluded_iot_hours[acct].add(hour_key)
            if utc_hour is not None:
                prev_hour = utc_hour
            if meter_ids:
                prev_meter_ids = meter_ids

    sorted_hours = sorted(hour_data.keys())
    time_series: List[Dict[str, Any]] = []
    visible_hours = [hour_key for hour_key in sorted_hours if hour_key >= visible_cutoff_key]
    for hour_key in sorted_hours:
        if hour_key < visible_cutoff_key:
            continue
        point: Dict[str, Any] = {"reading_hour": hour_key}
        for pair in pairs:
            acct = pair["account"]
            vals = hour_data[hour_key].get(acct, {"sm": None, "1m": None})
            sm_val = vals["sm"]
            m1_val = vals["1m"]
            if hour_key in excluded_iot_hours.get(acct, set()):
                m1_val = None
            point[f"{acct}_sm"] = round(sm_val, 4) if sm_val is not None else None
            point[f"{acct}_1m"] = round(m1_val, 4) if m1_val is not None else None
        time_series.append(point)

    for pair in pairs:
        acct = pair["account"]
        deviations: List[float] = []
        sm_vals: List[float] = []
        m1_vals: List[float] = []
        for hour_key in visible_hours:
            vals = hour_data[hour_key].get(acct, {"sm": None, "1m": None})
            sm = vals["sm"]
            m1 = vals["1m"]
            if hour_key in excluded_iot_hours.get(acct, set()):
                m1 = None
            if sm is not None and m1 is not None and sm > 0:
                deviations.append((m1 - sm) / sm * 100)
                sm_vals.append(sm)
                m1_vals.append(m1)

        n = len(deviations)
        total_sm = sum(sm_vals)
        total_1m = sum(m1_vals)

        if n > 0:
            mean_dev = sum(deviations) / n
            stddev_dev = (
                math.sqrt(sum((d - mean_dev) ** 2 for d in deviations) / n)
                if n > 1
                else 0
            )
            mean_sm = total_sm / n
            mean_1m = total_1m / n
        else:
            mean_dev = stddev_dev = mean_sm = mean_1m = 0.0

        total_dev_pct = (
            (total_1m - total_sm) / total_sm * 100 if total_sm > 0 else 0
        )

        pair["stats"] = {
            "total_deviation_pct": round(total_dev_pct, 2),
            "mean_deviation_pct": round(mean_dev, 2),
            "stddev_deviation_pct": round(stddev_dev, 2),
            "mean_sm_kwh": round(mean_sm, 4),
            "mean_1m_kwh": round(mean_1m, 4),
            "n_matched_hours": n,
            "total_sm_kwh": round(total_sm, 2),
            "total_1m_kwh": round(total_1m, 2),
            "excluded_1m_hours": len([
                hour_key for hour_key in visible_hours
                if hour_key in excluded_iot_hours.get(acct, set())
            ]),
        }

    check_meter_ids = [p["check_meter_id"] for p in pairs]
    health_map: Dict[str, Dict[str, Any]] = {}
    if check_meter_ids:
        ph = ",".join(["%s"] * len(check_meter_ids))
        cursor.execute(
            f"SELECT meter_id, account_number, last_seen_at, last_sample_time, firmware_version "
            f"FROM prototype_meter_state WHERE meter_id IN ({ph})",
            tuple(str(m) for m in check_meter_ids),
        )
        now = datetime.now(timezone.utc)
        for row in cursor.fetchall():
            meter_id = str(row[0]).strip()
            acct = str(row[1]).strip()
            last_seen_db = row[2]
            raw_ts = str(row[3] or "").strip()
            last_seen = None
            hours_ago = None

            if last_seen_db:
                last_seen = last_seen_db
                if last_seen.tzinfo is None:
                    last_seen = last_seen.replace(tzinfo=timezone.utc)
                hours_ago = round((now - last_seen).total_seconds() / 3600, 1)
            elif raw_ts and len(raw_ts) >= 12:
                try:
                    last_seen = datetime(
                        int(raw_ts[:4]), int(raw_ts[4:6]),
                        int(raw_ts[6:8]), int(raw_ts[8:10]),
                        int(raw_ts[10:12]),
                        tzinfo=timezone.utc,
                    )
                    hours_ago = round((now - last_seen).total_seconds() / 3600, 1)
                except (ValueError, IndexError):
                    pass
            fw_ver = None
            try:
                if len(row) > 4 and row[4] is not None:
                    fw_ver = str(row[4]).strip() or None
            except (IndexError, TypeError):
                fw_ver = None
            health_map[acct] = {
                "meter_id": meter_id,
                "last_seen_utc": last_seen.astimezone(timezone.utc).strftime("%Y-%m-%dT%H:%M:%S") if last_seen else None,
                "hours_since_report": hours_ago,
                "firmware_version": fw_ver,
                "status": (
                    "online" if hours_ago is not None and hours_ago < 2
                    else "stale" if hours_ago is not None and hours_ago < 6
                    else "offline"
                ),
            }

    for pair in pairs:
        pair["health"] = health_map.get(pair["account"], {
            "meter_id": pair["check_meter_id"],
            "last_seen_utc": None,
            "hours_since_report": None,
            "firmware_version": None,
            "status": "unknown",
        })

    total_excluded = sum(
        len([hour_key for hour_key in visible_hours if hour_key in excluded_iot_hours.get(pair["account"], set())])
        for pair in pairs
    )

    result = {
        "pairs": pairs,
        "time_series": time_series,
        "days": days,
        "cutoff": cutoff.strftime("%Y-%m-%dT%H:%M:%S"),
    }
    if total_excluded:
        result["note"] = (
            f"Excluded {total_excluded} 1Meter hourly points after long gaps, "
            f"meter changes, or mixed-meter hours."
        )
    return result


def _style_export_sheet(ws) -> None:
    """Apply light formatting to export worksheets."""
    for cell in ws[1]:
        cell.font = cell.font.copy(bold=True)
    ws.freeze_panes = "A2"
    max_row = min(ws.max_row, 250)
    for col_idx in range(1, ws.max_column + 1):
        max_len = 0
        for row_idx in range(1, max_row + 1):
            val = ws.cell(row=row_idx, column=col_idx).value
            if val is not None:
                max_len = max(max_len, len(str(val)))
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = min(max_len + 2, 40)


def _export_check_meter_comparison_xlsx(data: Dict[str, Any], days: int) -> StreamingResponse:
    """Generate an XLSX workbook for offline check-meter analysis."""
    try:
        from openpyxl import Workbook
    except ImportError:
        raise HTTPException(status_code=500, detail="openpyxl not installed; XLSX export unavailable")

    wb = Workbook()

    meta_ws = wb.active
    meta_ws.title = "meta"
    meta_ws.append(["field", "value"])
    meta_ws.append(["generated_at_utc", datetime.utcnow().strftime("%Y-%m-%dT%H:%M:%S")])
    meta_ws.append(["days_param", days])
    meta_ws.append(["cutoff", data.get("cutoff") or ""])
    meta_ws.append(["pair_count", len(data.get("pairs") or [])])
    meta_ws.append(["time_series_points", len(data.get("time_series") or [])])
    if data.get("note"):
        meta_ws.append(["note", data["note"]])
    _style_export_sheet(meta_ws)

    summary_ws = wb.create_sheet("summary")
    summary_ws.append([
        "account",
        "primary_meter_id",
        "check_meter_id",
        "health_status",
        "firmware_version",
        "last_seen_utc",
        "hours_since_report",
        "matched_hours",
        "excluded_1m_hours",
        "total_sm_kwh",
        "total_1m_kwh",
        "total_deviation_pct",
        "mean_sm_kwh",
        "mean_1m_kwh",
        "mean_deviation_pct",
        "stddev_deviation_pct",
    ])
    for pair in data.get("pairs", []):
        stats = pair.get("stats", {})
        health = pair.get("health", {})
        summary_ws.append([
            pair.get("account"),
            pair.get("primary_meter_id"),
            pair.get("check_meter_id"),
            health.get("status"),
            health.get("firmware_version"),
            health.get("last_seen_utc"),
            health.get("hours_since_report"),
            stats.get("n_matched_hours"),
            stats.get("excluded_1m_hours"),
            stats.get("total_sm_kwh"),
            stats.get("total_1m_kwh"),
            stats.get("total_deviation_pct"),
            stats.get("mean_sm_kwh"),
            stats.get("mean_1m_kwh"),
            stats.get("mean_deviation_pct"),
            stats.get("stddev_deviation_pct"),
        ])
    _style_export_sheet(summary_ws)

    wide_ws = wb.create_sheet("hourly_wide")
    wide_headers = ["reading_hour"]
    for pair in data.get("pairs", []):
        acct = pair.get("account")
        wide_headers.extend([f"{acct}_sm", f"{acct}_1m"])
    wide_ws.append(wide_headers)
    for point in data.get("time_series", []):
        row = [point.get("reading_hour")]
        for pair in data.get("pairs", []):
            acct = pair.get("account")
            row.extend([point.get(f"{acct}_sm"), point.get(f"{acct}_1m")])
        wide_ws.append(row)
    _style_export_sheet(wide_ws)

    long_ws = wb.create_sheet("hourly_long")
    long_ws.append([
        "reading_hour",
        "account",
        "primary_meter_id",
        "check_meter_id",
        "sm_kwh",
        "one_meter_kwh",
        "deviation_pct",
    ])
    for point in data.get("time_series", []):
        reading_hour = point.get("reading_hour")
        for pair in data.get("pairs", []):
            acct = pair.get("account")
            sm_val = point.get(f"{acct}_sm")
            m1_val = point.get(f"{acct}_1m")
            deviation_pct = None
            if sm_val not in (None, 0) and m1_val is not None:
                deviation_pct = round((m1_val - sm_val) / sm_val * 100, 4)
            long_ws.append([
                reading_hour,
                acct,
                pair.get("primary_meter_id"),
                pair.get("check_meter_id"),
                sm_val,
                m1_val,
                deviation_pct,
            ])
    _style_export_sheet(long_ws)

    filename_suffix = "since_firmware_update" if days == 0 else f"last_{days}_days"
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": (
                f"attachment; filename=check_meter_comparison_{filename_suffix}.xlsx"
            )
        },
    )


@router.get("/check-meter-comparison")
def check_meter_comparison(
    days: int = Query(0, description="Number of days of history (0 = since first check-meter reading)"),
    user: CurrentUser = Depends(require_employee),
):
    """
    Hourly time series comparison of SparkMeter (primary) vs 1Meter (check)
    for every account that has both meter roles installed. Returns aligned
    time series plus per-pair deviation statistics.
    """
    with _get_connection() as conn:
        return _build_check_meter_comparison(conn, days)


@router.get("/check-meter-comparison/export")
def export_check_meter_comparison(
    days: int = Query(0, description="Number of days of history (0 = since first check-meter reading)"),
    user: CurrentUser = Depends(require_employee),
):
    """Download the check-meter comparison dataset as an Excel workbook."""
    with _get_connection() as conn:
        data = _build_check_meter_comparison(conn, days)
    return _export_check_meter_comparison_xlsx(data, days)


# ---------------------------------------------------------------------------
# Onboarding Pipeline Report
# ---------------------------------------------------------------------------

@router.get("/api/om-report/pipeline")
def onboarding_pipeline(
    site: Optional[str] = Query(None),
    user: CurrentUser = Depends(require_employee),
):
    """Aggregate commissioning step counts into a funnel.

    Returns counts at each stage: registered -> connection_fee_paid -> ... -> customer_commissioned.
    """
    from customer_api import get_connection

    steps = [
        "connection_fee_paid",
        "readyboard_fee_paid",
        "readyboard_tested",
        "readyboard_installed",
        "airdac_connected",
        "meter_installed",
        "customer_commissioned",
    ]

    with get_connection() as conn:
        cur = conn.cursor()

        site_clause = ""
        params: list = []
        if site:
            site_clause = "WHERE m.community = %s"
            params = [site]

        cur.execute(f"""
            SELECT count(*) FROM customers c
            LEFT JOIN accounts a ON a.customer_id = c.id
            LEFT JOIN meters m ON m.account_number = a.account_number
            {site_clause}
        """, params)
        total_registered = cur.fetchone()[0]

        funnel = [{"stage": "registered", "count": total_registered}]

        for step in steps:
            cur.execute(f"""
                SELECT count(*) FROM customers c
                LEFT JOIN accounts a ON a.customer_id = c.id
                LEFT JOIN meters m ON m.account_number = a.account_number
                WHERE c.{step} = true
                {("AND m.community = %s" if site else "")}
            """, [site] if site else [])
            funnel.append({"stage": step, "count": cur.fetchone()[0]})

        sites_list = []
        cur.execute("""
            SELECT DISTINCT m.community FROM meters m
            WHERE m.community IS NOT NULL AND m.community != ''
            ORDER BY m.community
        """)
        sites_list = [r[0] for r in cur.fetchall()]

        return {"funnel": funnel, "sites": sites_list}
