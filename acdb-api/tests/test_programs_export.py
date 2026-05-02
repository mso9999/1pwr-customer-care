"""Tests for the Connections claim export.

Verifies that ``programs._CONN_COLUMNS`` matches the column order and
``parsingHeaders`` row of the canonical Odyssey upload template at
``docs/uef_zedsi_claim_template.xlsx``. Drift here would cause Odyssey to
reject uploads silently, so this is a guard test.
"""

import os
import sys
import types
import unittest
import warnings

# Stub customer_api before importing programs (same pattern as
# test_odyssey_api -- avoids loading the heavy customer_api import chain
# which pulls xhtml2pdf etc.).
_stub = types.ModuleType("customer_api")
_stub.get_connection = lambda: None  # type: ignore[attr-defined]
sys.modules.setdefault("customer_api", _stub)

# Stub middleware -- only require_role attribute is referenced at import time.
def _require_role_stub(*_a, **_k):
    def _dep():
        return None
    return _dep
_mw = types.ModuleType("middleware")
_mw.require_role = _require_role_stub  # type: ignore[attr-defined]
sys.modules.setdefault("middleware", _mw)

# Stub mutations
_mu = types.ModuleType("mutations")
_mu.try_log_mutation = lambda *a, **k: None  # type: ignore[attr-defined]
sys.modules.setdefault("mutations", _mu)

import programs  # noqa: E402

REPO_ROOT = os.path.abspath(os.path.join(os.path.dirname(__file__), "..", ".."))
TEMPLATE = os.path.join(REPO_ROOT, "docs", "uef_zedsi_claim_template.xlsx")


class TestConnectionsTemplateAlignment(unittest.TestCase):
    def setUp(self):
        warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")

    def _load_template_headers(self):
        import openpyxl

        wb = openpyxl.load_workbook(TEMPLATE, data_only=False)
        ws = wb["Connections"]
        ws_parse = wb["parsingHeaders"]
        display = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
        parsing = [c.value for c in next(ws_parse.iter_rows(min_row=1, max_row=1))]
        return display, parsing

    def test_column_count_matches_template(self):
        display, parsing = self._load_template_headers()
        self.assertEqual(
            len(display), len(programs._CONN_COLUMNS),
            f"Display header count drift: template has {len(display)} cols, "
            f"_CONN_COLUMNS has {len(programs._CONN_COLUMNS)}.",
        )
        self.assertEqual(len(parsing), len(programs._CONN_COLUMNS))

    def test_display_headers_in_order(self):
        display, _ = self._load_template_headers()
        for i, (template_h, ours) in enumerate(zip(display, programs._CONN_COLUMNS)):
            self.assertEqual(
                template_h, ours["display"],
                f"Display header drift at column {i + 1}: "
                f"template={template_h!r} ours={ours['display']!r}",
            )

    def test_parsing_headers_in_order(self):
        _, parsing = self._load_template_headers()
        for i, (template_p, ours) in enumerate(zip(parsing, programs._CONN_COLUMNS)):
            self.assertEqual(
                template_p, ours["parsing"],
                f"parsingHeader drift at column {i + 1}: "
                f"template={template_p!r} ours={ours['parsing']!r}",
            )


class TestResolveCell(unittest.TestCase):
    def test_row_source(self):
        out = programs._resolve_cell({"x": 42}, ("row", "x"))
        self.assertEqual(out, 42)

    def test_const_source(self):
        out = programs._resolve_cell({}, ("const", "Hello"))
        self.assertEqual(out, "Hello")

    def test_callable_source(self):
        out = programs._resolve_cell({"a": 1, "b": 2}, ("callable", lambda r: r["a"] + r["b"]))
        self.assertEqual(out, 3)

    def test_callable_swallows_errors(self):
        out = programs._resolve_cell({}, ("callable", lambda r: r["missing"]))
        self.assertEqual(out, "")


if __name__ == "__main__":
    unittest.main()
