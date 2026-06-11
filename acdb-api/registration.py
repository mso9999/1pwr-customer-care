"""
Customer Registration Module
=============================
Account number generation, portal registration, and Excel bulk import.

Replaces VBA: ExcelImport.bas, programs.bas (assignaccno)

Endpoints:
  POST /api/customers/register         — Create single customer
  POST /api/customers/bulk-import      — Upload Excel for bulk registration
  GET  /api/customers/next-account     — Preview next account number for a site
"""

import io
import logging
import re
from datetime import datetime
from typing import Any, Dict, List, Optional

from fastapi import APIRouter, Depends, File, HTTPException, Query, UploadFile
from pydantic import BaseModel, Field

from country_config import COUNTRY
from country_fees import get_country_fees
from middleware import require_employee, require_role
from models import CurrentUser
from mutations import log_mutation, try_log_mutation
from sparkmeter_customer import create_sparkmeter_customer

logger = logging.getLogger("cc-api.registration")

router = APIRouter(prefix="/api/customers", tags=["registration"])


def _get_connection():
    from customer_api import get_connection
    return get_connection()


# ---------------------------------------------------------------------------
# Account number generation
# ---------------------------------------------------------------------------

def generate_account_number(conn, community: str) -> str:
    """Generate the next sequential account number for a community.

    Format: NNNNXXX (e.g., 0042MAS)
    Uses PostgreSQL's next_account_number() function defined in schema.
    """
    cursor = conn.cursor()
    comm = community.upper()
    cursor.execute("SELECT next_account_number(%s)", (comm,))
    candidate = str(cursor.fetchone()[0] or "").strip().upper()

    # Defensive collision handling: if account_sequence drift exists in DB,
    # next_account_number() can return an already-used code. Walk forward
    # until we find a free account number.
    for _ in range(50):
        cursor.execute(
            "SELECT 1 FROM accounts WHERE account_number = %s LIMIT 1",
            (candidate,),
        )
        if not cursor.fetchone():
            return candidate

        m = re.match(r"^(\d{4})([A-Z]{2,4})$", candidate)
        if not m:
            raise HTTPException(
                status_code=500,
                detail=f"Invalid generated account number format: {candidate}",
            )
        next_seq = int(m.group(1)) + 1
        candidate = f"{next_seq:04d}{comm}"

    raise HTTPException(
        status_code=500,
        detail=f"Could not allocate unique account number for site {comm}",
    )


# ---------------------------------------------------------------------------
# Models
# ---------------------------------------------------------------------------

class CustomerCreateRequest(BaseModel):
    first_name: str = Field(..., min_length=1)
    middle_name: Optional[str] = None
    gender: Optional[str] = None
    last_name: str = Field(..., min_length=1)
    community: str = Field(..., min_length=2, max_length=10)
    phone: Optional[str] = None
    cell_phone_1: Optional[str] = None
    cell_phone_2: Optional[str] = None
    email: Optional[str] = None
    national_id: Optional[str] = None
    plot_number: Optional[str] = None
    street_address: Optional[str] = None
    city: Optional[str] = None
    district: Optional[str] = None
    customer_type: Optional[str] = None
    gps_lat: Optional[float] = None
    gps_lon: Optional[float] = None
    date_service_connected: Optional[str] = None
    meter_id: Optional[str] = None
    acquires_1pwr_readyboard: bool = False
    # Optional EXISTING account number (legacy ACCDB accounts that were never imported
    # but already carry payments). When provided, CC uses it instead of auto-generating;
    # creating the account adopts any transactions already keyed to that number.
    account_number: Optional[str] = None


class BulkImportResult(BaseModel):
    total_rows: int
    imported: int
    skipped: int
    errors: List[Dict[str, Any]]


VALID_CUSTOMER_TYPES = {
    "HH1", "HH2", "HH3",
    "SME", "CHU", "SCP", "SCH", "HC", "PWH", "GOV", "COM", "IND",
    "REL", "AGR", "CLI", "PUE", "HCF", "OTH", "OTHER",
}

VALID_GENDERS = {
    "MALE": "Male",
    "M": "Male",
    "FEMALE": "Female",
    "F": "Female",
}


def _infer_customer_type(explicit_value: Optional[str], plot_number: Optional[str]) -> Optional[str]:
    """Resolve a canonical stored customer type without persisting aggregate HH."""
    explicit = str(explicit_value or "").strip().upper()
    if explicit in VALID_CUSTOMER_TYPES:
        return explicit
    if explicit == "HH":
        explicit = ""

    plot = str(plot_number or "").strip().upper()
    if plot:
        for code in sorted(VALID_CUSTOMER_TYPES | {"HH"}, key=len, reverse=True):
            if re.search(rf"(?:^|[\s_]){re.escape(code)}(?:[\s_]|$)", plot):
                if code == "HH":
                    return None
                return code

    return None


def _normalize_phone_for_storage(raw: Optional[str]) -> Optional[str]:
    digits = "".join(c for c in str(raw or "") if c.isdigit())
    if not digits:
        return None

    if digits.startswith("00"):
        digits = digits[2:]
    if digits.startswith(COUNTRY.dial_code) and len(digits) > len(COUNTRY.dial_code):
        digits = digits[len(COUNTRY.dial_code):]
    if digits.startswith("0") and len(digits) > 8:
        digits = digits[1:]

    return digits or None


def _normalize_gender_for_storage(raw: Optional[str]) -> Optional[str]:
    value = str(raw or "").strip()
    if not value:
        return None

    normalized = VALID_GENDERS.get(value.upper())
    if not normalized:
        raise HTTPException(status_code=400, detail="gender must be Male or Female when provided")

    return normalized


# ---------------------------------------------------------------------------
# Endpoints
# ---------------------------------------------------------------------------

@router.post("/register")
def register_customer(
    req: CustomerCreateRequest,
    user: CurrentUser = Depends(require_employee),
):
    """Register a new customer.

    Account number is auto-generated unless ``account_number`` is supplied (legacy
    ACCDB accounts known to the field team — see O&M request 2026-06-10). Manual
    numbers are validated for format, site match, and uniqueness.
    """
    with _get_connection() as conn:
        cursor = conn.cursor()
        try:
            community = req.community.upper()

            manual_account = (req.account_number or "").strip().upper()
            if manual_account:
                if not re.match(r"^\d{4}[A-Z]{2,4}$", manual_account):
                    raise HTTPException(
                        status_code=400,
                        detail=(
                            f"Invalid account number '{manual_account}' — expected "
                            "4 digits + site code (e.g. 0286SHG)"
                        ),
                    )
                if not manual_account.endswith(community):
                    raise HTTPException(
                        status_code=400,
                        detail=(
                            f"Account number '{manual_account}' does not match the "
                            f"selected site '{community}'"
                        ),
                    )
                cursor.execute(
                    "SELECT 1 FROM accounts WHERE account_number = %s LIMIT 1",
                    (manual_account,),
                )
                if cursor.fetchone():
                    raise HTTPException(
                        status_code=409,
                        detail=f"Account number '{manual_account}' already exists",
                    )
                account_number = manual_account
            else:
                account_number = generate_account_number(conn, community)

            resolved_customer_type = _infer_customer_type(req.customer_type, req.plot_number)
            gender = _normalize_gender_for_storage(req.gender)
            phone = _normalize_phone_for_storage(req.phone)
            cell_phone_1 = _normalize_phone_for_storage(req.cell_phone_1)
            cell_phone_2 = _normalize_phone_for_storage(req.cell_phone_2)

            # Insert customer
            cursor.execute("""
                INSERT INTO customers (
                    first_name, middle_name, gender, last_name, community, phone, cell_phone_1,
                    cell_phone_2, email, national_id, plot_number,
                    street_address, city, district, customer_type,
                    gps_lat, gps_lon, date_service_connected, is_active,
                    created_by, updated_by
                ) VALUES (
                    %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s,
                    TRUE, %s, %s
                ) RETURNING id, customer_id_legacy
            """, (
                req.first_name, req.middle_name, gender, req.last_name, community,
                phone, cell_phone_1, cell_phone_2,
                req.email, req.national_id, req.plot_number,
                req.street_address, req.city, req.district,
                resolved_customer_type, req.gps_lat, req.gps_lon, req.date_service_connected,
                user.user_id, user.user_id,
            ))

            row = cursor.fetchone()
            customer_pg_id = row[0]
            customer_legacy_id = row[1]

            fees = get_country_fees(conn)
            conn_fee_amt = float(fees["connection_fee_amount"])
            rb_fee_amt = (
                float(fees["readyboard_fee_amount"])
                if req.acquires_1pwr_readyboard
                else 0.0
            )
            cursor.execute(
                """
                UPDATE customers
                   SET acquires_1pwr_readyboard = %s,
                       fee_debt_connection_remaining = %s,
                       fee_debt_readyboard_remaining = %s
                 WHERE id = %s
                """,
                (bool(req.acquires_1pwr_readyboard), conn_fee_amt, rb_fee_amt, customer_pg_id),
            )

            # Extract sequence number from account number
            seq = int(account_number[:4])

            # Create account record
            cursor.execute("""
                INSERT INTO accounts (
                    account_number, customer_id, meter_id, community,
                    account_sequence, created_by
                ) VALUES (%s, %s, %s, %s, %s, %s)
            """, (account_number, customer_pg_id, req.meter_id, community, seq, user.user_id))

            customer_values = {
                "id": customer_pg_id,
                "customer_id_legacy": customer_legacy_id,
                "first_name": req.first_name,
                "middle_name": req.middle_name,
                "gender": gender,
                "last_name": req.last_name,
                "community": community,
                "phone": phone,
                "cell_phone_1": cell_phone_1,
                "cell_phone_2": cell_phone_2,
                "email": req.email,
                "national_id": req.national_id,
                "plot_number": req.plot_number,
                "street_address": req.street_address,
                "city": req.city,
                "district": req.district,
                "customer_type": resolved_customer_type,
                "gps_lat": req.gps_lat,
                "gps_lon": req.gps_lon,
                "date_service_connected": req.date_service_connected,
                "is_active": True,
                "created_by": user.user_id,
                "updated_by": user.user_id,
            }
            account_values = {
                "account_number": account_number,
                "customer_id": customer_pg_id,
                "meter_id": req.meter_id,
                "community": community,
                "account_sequence": seq,
                "created_by": user.user_id,
                "manual_account_number": bool(manual_account),
            }
            log_mutation(
                user,
                "create",
                "customers",
                str(customer_pg_id),
                new_values=customer_values,
                conn=conn,
            )
            log_mutation(
                user,
                "create",
                "accounts",
                account_number,
                new_values=account_values,
                conn=conn,
            )
            conn.commit()
        except HTTPException:
            conn.rollback()
            raise
        except Exception as e:
            conn.rollback()
            raise HTTPException(status_code=400, detail=f"Customer registration failed: {e}")

        logger.info(
            "Customer registered: %s %s -> %s by %s",
            req.first_name, req.last_name, account_number, user.user_id,
        )

        # Claim any parked merchant-line payments that referenced this account number
        # before it existed (see merchant_unmatched.py). Own transaction; best-effort.
        claimed_payments: list = []
        try:
            from merchant_unmatched import claim_unmatched_for_account

            claimed_payments = claim_unmatched_for_account(conn, account_number)
            if claimed_payments:
                conn.commit()
            else:
                conn.rollback()
        except Exception as e:
            logger.warning("Unmatched-payment claim failed for %s: %s", account_number, e)
            try:
                conn.rollback()
            except Exception:
                pass

        sm_result = None
        try:
            full_name = f"{req.first_name} {req.last_name}".strip()
            phone = _normalize_phone_for_storage(req.phone) or _normalize_phone_for_storage(req.cell_phone_1)
            sm_result = create_sparkmeter_customer(
                account_number=account_number,
                name=full_name,
                meter_serial=req.meter_id,
                phone=phone,
            )
            if sm_result.success and not sm_result.skipped:
                logger.info("SM customer synced: %s -> %s", account_number, sm_result.platform)
            elif sm_result.skipped:
                logger.info("SM customer sync deferred for %s: %s", account_number, sm_result.error)
            else:
                logger.warning("SM customer sync failed for %s: %s", account_number, sm_result.error)
        except Exception as e:
            logger.error("SM customer sync exception for %s: %s", account_number, e)

        response: dict = {
            "account_number": account_number,
            "customer_id": customer_pg_id,
            "customer_id_legacy": customer_legacy_id,
            "first_name": req.first_name,
            "last_name": req.last_name,
            "community": community,
        }
        if claimed_payments:
            response["claimed_payments"] = claimed_payments
        if sm_result:
            response["sm_sync"] = {
                "success": sm_result.success,
                "platform": sm_result.platform,
                "skipped": sm_result.skipped,
            }
            if sm_result.sm_customer_id:
                response["sm_sync"]["sm_customer_id"] = sm_result.sm_customer_id
            if sm_result.error and not sm_result.success:
                response["sm_sync"]["error"] = sm_result.error

        return response


@router.get("/next-account")
def preview_next_account(
    community: str = Query(..., min_length=2, max_length=10),
    user: CurrentUser = Depends(require_employee),
):
    """Preview the next account number that would be generated for a site."""
    with _get_connection() as conn:
        account_number = generate_account_number(conn, community.upper())
        # Don't commit — this is a preview, not a reservation
        conn.rollback()
        return {"community": community.upper(), "next_account_number": account_number}


@router.post("/bulk-import", response_model=BulkImportResult)
async def bulk_import_customers(
    file: UploadFile = File(...),
    community: str = Query(..., min_length=2, max_length=10),
    user: CurrentUser = Depends(require_role(["superadmin", "onm_team"])),
):
    """Bulk import customers from an Excel file.

    Expected columns: first_name, last_name, phone, customer_type,
                      gender,
                      plot_number, national_id, gps_lat, gps_lon
    """
    if not file.filename.endswith((".xlsx", ".xls")):
        raise HTTPException(status_code=400, detail="File must be .xlsx or .xls")

    try:
        import openpyxl
    except ImportError:
        raise HTTPException(status_code=500, detail="openpyxl not installed")

    contents = await file.read()
    wb = openpyxl.load_workbook(io.BytesIO(contents), read_only=True)
    ws = wb.active

    # Read header row
    headers = [str(cell.value or "").strip().lower().replace(" ", "_")
               for cell in next(ws.iter_rows(min_row=1, max_row=1))]

    required = {"first_name", "last_name"}
    if not required.issubset(set(headers)):
        raise HTTPException(
            status_code=400,
            detail=f"Excel must have columns: {', '.join(required)}. Found: {headers}",
        )

    community_upper = community.upper()
    imported = 0
    skipped = 0
    errors = []
    imported_accounts: list[str] = []
    total_rows = 0

    with _get_connection() as conn:
        cursor = conn.cursor()

        for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            total_rows = row_num - 1
            row_dict = dict(zip(headers, row))

            first_name = str(row_dict.get("first_name", "")).strip()
            last_name = str(row_dict.get("last_name", "")).strip()

            if not first_name or not last_name:
                skipped += 1
                continue

            cursor.execute("SAVEPOINT bulk_import_row")
            try:
                account_number = generate_account_number(conn, community_upper)
                seq = int(account_number[:4])
                gender = _normalize_gender_for_storage(
                    str(row_dict.get("gender", "") or "").strip() or None
                )

                cursor.execute("""
                    INSERT INTO customers (
                        first_name, gender, last_name, community, phone,
                        national_id, plot_number, customer_type,
                        gps_lat, gps_lon, is_active,
                        created_by
                    ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, TRUE, %s)
                    RETURNING id, customer_id_legacy
                """, (
                    # Preserve explicit HH1/HH2/HH3/etc. when present, but do not
                    # persist aggregate HH as though it were an atomic type.
                    first_name, gender, last_name, community_upper,
                    _normalize_phone_for_storage(str(row_dict.get("phone", "") or "").strip() or None),
                    str(row_dict.get("national_id", "") or "").strip() or None,
                    str(row_dict.get("plot_number", "") or "").strip() or None,
                    _infer_customer_type(
                        str(row_dict.get("customer_type", "") or "").strip() or None,
                        str(row_dict.get("plot_number", "") or "").strip() or None,
                    ),
                    float(row_dict["gps_lat"]) if row_dict.get("gps_lat") else None,
                    float(row_dict["gps_lon"]) if row_dict.get("gps_lon") else None,
                    user.user_id,
                ))
                customer_pg_row = cursor.fetchone()
                customer_pg_id = customer_pg_row[0]
                customer_legacy_id = customer_pg_row[1]

                fees = get_country_fees(conn)
                cursor.execute(
                    """
                    UPDATE customers
                       SET acquires_1pwr_readyboard = false,
                           fee_debt_connection_remaining = %s,
                           fee_debt_readyboard_remaining = 0
                     WHERE id = %s
                    """,
                    (float(fees["connection_fee_amount"]), customer_pg_id),
                )

                cursor.execute("""
                    INSERT INTO accounts (
                        account_number, customer_id, community,
                        account_sequence, created_by
                    ) VALUES (%s, %s, %s, %s, %s)
                """, (account_number, customer_pg_id, community_upper, seq, user.user_id))

                customer_values = {
                    "id": customer_pg_id,
                    "customer_id_legacy": customer_legacy_id,
                    "first_name": first_name,
                    "gender": gender,
                    "last_name": last_name,
                    "community": community_upper,
                    "phone": _normalize_phone_for_storage(str(row_dict.get("phone", "") or "").strip() or None),
                    "national_id": str(row_dict.get("national_id", "") or "").strip() or None,
                    "plot_number": str(row_dict.get("plot_number", "") or "").strip() or None,
                    "customer_type": _infer_customer_type(
                        str(row_dict.get("customer_type", "") or "").strip() or None,
                        str(row_dict.get("plot_number", "") or "").strip() or None,
                    ),
                    "gps_lat": float(row_dict["gps_lat"]) if row_dict.get("gps_lat") else None,
                    "gps_lon": float(row_dict["gps_lon"]) if row_dict.get("gps_lon") else None,
                    "created_by": user.user_id,
                }
                account_values = {
                    "account_number": account_number,
                    "customer_id": customer_pg_id,
                    "community": community_upper,
                    "account_sequence": seq,
                    "created_by": user.user_id,
                }
                log_mutation(
                    user,
                    "create",
                    "customers",
                    str(customer_pg_id),
                    new_values=customer_values,
                    conn=conn,
                )
                log_mutation(
                    user,
                    "create",
                    "accounts",
                    account_number,
                    new_values=account_values,
                    conn=conn,
                )
                cursor.execute("RELEASE SAVEPOINT bulk_import_row")
                imported += 1
                imported_accounts.append(account_number)

                try:
                    phone_raw = str(row_dict.get("phone", "") or "").strip() or None
                    sm_r = create_sparkmeter_customer(
                        account_number=account_number,
                        name=f"{first_name} {last_name}".strip(),
                        phone=_normalize_phone_for_storage(phone_raw),
                    )
                    if not sm_r.success and not sm_r.skipped:
                        logger.warning("Bulk SM sync failed for %s: %s", account_number, sm_r.error)
                except Exception as e:
                    logger.error("Bulk SM sync exception for %s: %s", account_number, e)
            except Exception as e:
                cursor.execute("ROLLBACK TO SAVEPOINT bulk_import_row")
                cursor.execute("RELEASE SAVEPOINT bulk_import_row")
                errors.append({
                    "row": row_num,
                    "error": e.detail if isinstance(e, HTTPException) else str(e),
                })

        conn.commit()

    wb.close()
    if imported > 0:
        try_log_mutation(
            user,
            "bulk_import",
            "customers",
            community_upper,
            new_values={
                "community": community_upper,
                "filename": file.filename,
                "imported": imported,
                "skipped": skipped,
                "error_count": len(errors),
                "account_numbers": imported_accounts,
            },
            metadata={"total_rows": total_rows},
        )
    logger.info(
        "Bulk import: %d imported, %d skipped, %d errors from %s by %s",
        imported, skipped, len(errors), file.filename, user.user_id,
    )

    return BulkImportResult(
        total_rows=total_rows,
        imported=imported,
        skipped=skipped,
        errors=errors[:20],  # Cap error list
    )
