"""
CAPEX Import Module
===================

Parses the 1PWR Financial Model (FM) workbook for CAPEX per site
and stores it in the financial_metrics table.

Provides:
  - ``parse_fm_workbook(path)`` — Extract CAPEX per site from the FM workbook
  - ``POST /api/admin/import-capex`` — Upload FM workbook and import CAPEX
"""

from __future__ import annotations

import io
import logging
from datetime import date
from typing import Any, Dict, List, Optional

from fastapi import APIRouter, Depends, File, UploadFile
from models import CurrentUser, CCRole
from middleware import require_role

logger = logging.getLogger("cc-api.capex")

router = APIRouter(prefix="/api/admin", tags=["capex"])


def parse_fm_workbook(file_bytes: bytes) -> List[Dict[str, Any]]:
    """Parse the FM workbook and extract CAPEX per site.

    Expected structure: a sheet with site codes and CAPEX columns.
    Returns a list of dicts with site_code, capex_deployed_usd, period.
    """
    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(file_bytes), data_only=True)
    results: List[Dict[str, Any]] = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        # Look for rows that look like site data (3-letter site code in first column)
        for row in ws.iter_rows(min_row=1, values_only=False):
            first_cell = row[0].value
            if first_cell and isinstance(first_cell, str) and len(first_cell) == 3 and first_cell.isupper():
                site_code = first_cell.strip()
                # Try to find CAPEX columns — scan header row or use heuristics
                # This is intentionally flexible since FM workbook structure varies
                capex_values: List[float] = []
                for cell in row[1:]:
                    val = cell.value
                    if isinstance(val, (int, float)) and val > 1000:
                        capex_values.append(float(val))

                if capex_values:
                    results.append({
                        "site_code": site_code,
                        "capex_deployed_usd": max(capex_values),
                        "period": f"{date.today().year}-01",
                        "source": "fm_workbook",
                    })

    logger.info("CAPEX import: parsed %d sites from FM workbook", len(results))
    return results


def store_capex(records: List[Dict[str, Any]]) -> int:
    """Store CAPEX records in financial_metrics table."""
    from customer_api import get_connection

    with get_connection() as conn:
        stored = 0
        with conn.cursor() as cur:
            for rec in records:
                cur.execute(
                    """
                    INSERT INTO financial_metrics
                        (site_code, period, capex_deployed_usd, capex_cumulative_usd, source, synced_at)
                    VALUES (%s, %s, %s, %s, %s, NOW())
                    ON CONFLICT (site_code, period, source) DO UPDATE SET
                        capex_deployed_usd = EXCLUDED.capex_deployed_usd,
                        capex_cumulative_usd = EXCLUDED.capex_cumulative_usd,
                        synced_at = NOW()
                    """,
                    (rec["site_code"], rec["period"], rec["capex_deployed_usd"],
                     rec.get("capex_cumulative_usd"), rec["source"]),
                )
                stored += 1
        conn.commit()

    logger.info("CAPEX import: stored %d records", stored)
    return stored


@router.post("/import-capex")
async def import_capex(
    file: UploadFile = File(...),
    user: CurrentUser = Depends(require_role(CCRole.superadmin)),
) -> Dict[str, Any]:
    """Upload an FM workbook and import CAPEX data. Superadmin only."""
    contents = await file.read()
    records = parse_fm_workbook(contents)
    stored = store_capex(records)
    return {
        "file": file.filename,
        "records_parsed": len(records),
        "records_stored": stored,
    }
