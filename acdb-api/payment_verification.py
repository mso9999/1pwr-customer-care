"""
Payment verification workflow for 1PWR Customer Care Portal.

Provides endpoints for verifying/rejecting payments (connection fees,
readyboard fees, etc.) that require finance team approval.
"""

import io
import logging
from datetime import datetime, timezone
from typing import List, Optional

from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from pydantic import BaseModel

from models import CurrentUser
from middleware import require_employee
from customer_api import get_connection

logger = logging.getLogger("cc-api.verification")

router = APIRouter(prefix="/api/payment-verification", tags=["payment-verification"])


class VerificationAction(BaseModel):
    ids: List[int]
    action: str  # "verify" or "reject"
    note: Optional[str] = None


@router.get("/pending")
def list_pending(
    status: str = Query("pending"),
    payment_type: Optional[str] = Query(None),
    account_number: Optional[str] = Query(None),
    limit: int = Query(100, le=500),
    offset: int = Query(0, ge=0),
    user: CurrentUser = Depends(require_employee),
):
    clauses = ["pv.status = %s"]
    params: list = [status]

    if payment_type:
        clauses.append("pv.payment_type = %s")
        params.append(payment_type)
    if account_number:
        clauses.append("pv.account_number = %s")
        params.append(account_number)

    where = "WHERE " + " AND ".join(clauses)

    with get_connection() as conn:
        cur = conn.cursor()
        cur.execute(f"""
            SELECT pv.*,
                   c.first_name, c.last_name
            FROM payment_verifications pv
            LEFT JOIN accounts a ON pv.account_number = a.account_number
            LEFT JOIN customers c ON a.customer_id = c.id
            {where}
            ORDER BY pv.created_at DESC
            LIMIT %s OFFSET %s
        """, params + [limit, offset])
        cols = [d[0] for d in cur.description]
        rows = [dict(zip(cols, r)) for r in cur.fetchall()]

        cur.execute(f"""
            SELECT count(*) FROM payment_verifications pv {where}
        """, params)
        total = cur.fetchone()[0]

        return {"verifications": rows, "total": total}


@router.post("/verify")
def verify_payments(body: VerificationAction, user: CurrentUser = Depends(require_employee)):
    if body.action not in ("verify", "reject"):
        raise HTTPException(400, "Action must be 'verify' or 'reject'")

    new_status = "verified" if body.action == "verify" else "rejected"

    with get_connection() as conn:
        cur = conn.cursor()
        cur.execute("""
            UPDATE payment_verifications
            SET status = %s, verified_by = %s, verified_at = NOW(), note = COALESCE(%s, note)
            WHERE id = ANY(%s) AND status = 'pending'
            RETURNING id
        """, (new_status, user.user_id, body.note, body.ids))
        updated = [r[0] for r in cur.fetchall()]
        conn.commit()

    return {"updated": len(updated), "ids": updated, "status": new_status}


@router.get("/export")
def export_verifications(
    status: str = Query("pending"),
    payment_type: Optional[str] = Query(None),
    user: CurrentUser = Depends(require_employee),
):
    """Export payment verifications to XLSX."""
    try:
        import openpyxl
    except ImportError:
        raise HTTPException(500, "openpyxl not installed on server")

    clauses = ["pv.status = %s"]
    params: list = [status]
    if payment_type:
        clauses.append("pv.payment_type = %s")
        params.append(payment_type)
    where = "WHERE " + " AND ".join(clauses)

    with get_connection() as conn:
        cur = conn.cursor()
        cur.execute(f"""
            SELECT pv.id, pv.created_at, pv.account_number,
                   c.first_name, c.last_name,
                   pv.payment_type, pv.amount, pv.status,
                   pv.verified_by, pv.verified_at, pv.note
            FROM payment_verifications pv
            LEFT JOIN accounts a ON pv.account_number = a.account_number
            LEFT JOIN customers c ON a.customer_id = c.id
            {where}
            ORDER BY pv.created_at DESC
        """, params)
        rows = cur.fetchall()
        cols = [d[0] for d in cur.description]

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Payment Verifications"

    headers = ["ID", "Date", "Account", "First Name", "Last Name",
               "Type", "Amount", "Status", "Verified By", "Verified At", "Note"]
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.font = openpyxl.styles.Font(bold=True)

    for row_idx, row in enumerate(rows, 2):
        for col_idx, val in enumerate(row, 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            if isinstance(val, datetime):
                cell.value = val.strftime("%Y-%m-%d %H:%M")
            else:
                cell.value = val

    for col in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 40)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"payment_verifications_{status}_{datetime.now().strftime('%Y%m%d')}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


def create_verification_entry(
    conn, transaction_id: int, account_number: str,
    payment_type: str, amount: float,
):
    """Insert a pending payment verification (called from payments.py or commission.py)."""
    cur = conn.cursor()
    cur.execute("""
        INSERT INTO payment_verifications
            (transaction_id, account_number, payment_type, amount)
        VALUES (%s, %s, %s, %s)
        RETURNING id
    """, (transaction_id, account_number, payment_type, amount))
    return cur.fetchone()[0]
