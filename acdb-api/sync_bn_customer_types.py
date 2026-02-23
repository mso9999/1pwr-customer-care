"""
Sync Benin customer types into the meters table.

Two-tier mapping:
  1. Census spreadsheets (primary): HH1/HH2/HH3 → HH, SME, CHU, SCH
     These are the authoritative 1PWR customer classifications from the
     Recensement Clients workbooks.
  2. Koios tariff names (fallback): For customers not in the census,
     the Koios tariff name is mapped to the nearest 1PWR classification.

Usage:
    python3 sync_bn_customer_types.py              # sync all sites
    python3 sync_bn_customer_types.py --dry-run     # preview without writing
"""
import argparse
import logging
import os
import re

import psycopg2
import requests

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    datefmt="%Y-%m-%d %H:%M:%S",
)
log = logging.getLogger("sync_bn_types")

DATABASE_URL = os.environ.get(
    "DATABASE_URL",
    "postgresql://cc_api:gKkYLkzYwSRPNoSwuC87YVqbzCmnhI4e@localhost:5432/onepower_bj",
)
KOIOS_BASE = "https://www.sparkmeter.cloud"
ORG_ID = "0123589c-7f1f-4eb4-8888-d8f8aa706ea4"

KOIOS_EMAIL = os.environ.get("KOIOS_WEB_EMAIL", "mso@1pwrafrica.com")
KOIOS_PASSWORD = os.environ.get("KOIOS_WEB_PASSWORD", "1PWRBN2026")

SITES = {
    "GBO": "a23c334e-33f7-473d-9ae3-9e631d5336e4",
    "SAM": "8f80b0a8-0502-4e26-9043-7152979360aa",
}

# --- Census type normalisation (tier 1) ---
# Collapse HH tiers into a single "HH" to match the LS classification.
CENSUS_NORMALIZE = {
    "HH1": "HH",
    "HH2": "HH",
    "HH3": "HH",
    "SME": "SME",
    "CHU": "CHU",
    "SCH": "SCH",
}

# --- Koios tariff → 1PWR type (tier 2 fallback) ---
TARIFF_TO_TYPE = {
    "Residentiel A": "HH",
    "Residentiel B": "HH",
    "PME": "SME",
    "Industriel": "SME",
    "INDUSTRIEL TEST": "SME",
    "Social": "CHU",
    "PWH AC": "PWH",
    "PWH METER": "PWH",
}

# Path to census spreadsheets (on the EC2, deployed alongside this script)
CENSUS_DIR = os.environ.get("CENSUS_DIR", os.path.join(os.path.dirname(__file__), "data", "bn_census"))


def load_census_mapping():
    """Load {account_number: customer_type} from local census XLSX files.

    Returns empty dict if openpyxl is unavailable or files are missing.
    """
    mapping = {}
    try:
        import openpyxl
    except ImportError:
        log.warning("openpyxl not installed — skipping census spreadsheets")
        return mapping

    census_dir = CENSUS_DIR
    if not os.path.isdir(census_dir):
        log.info("Census directory not found at %s — skipping tier-1", census_dir)
        return mapping

    for fname in os.listdir(census_dir):
        if not fname.lower().endswith(".xlsx") or fname.startswith("~"):
            continue
        path = os.path.join(census_dir, fname)
        try:
            wb = openpyxl.load_workbook(path, read_only=True)
            ws = wb.active
            headers = [str(c or "").strip().lower() for c in next(ws.iter_rows(max_row=1, values_only=True))]

            code_idx = next((i for i, h in enumerate(headers) if "concession" in h), None)
            type_idx = next((i for i, h in enumerate(headers) if "type" in h and ("hh" in h or "sme" in h)), None)
            if code_idx is None or type_idx is None:
                log.warning("  %s: could not find code/type columns — skipping", fname)
                wb.close()
                continue

            count = 0
            for row in ws.iter_rows(min_row=3, values_only=True):
                if len(row) <= max(code_idx, type_idx):
                    continue
                code = str(row[code_idx] or "").strip()
                raw_type = str(row[type_idx] or "").strip().upper()
                if not code or not raw_type:
                    continue
                ctype = CENSUS_NORMALIZE.get(raw_type, raw_type)
                mapping[code] = ctype
                count += 1

            log.info("  %s: %d customer types loaded", fname, count)
            wb.close()
        except Exception as e:
            log.warning("  %s: failed to read — %s", fname, e)

    return mapping


def koios_login(session):
    r = session.get(f"{KOIOS_BASE}/login", timeout=30)
    r.raise_for_status()
    csrf = re.search(r'name="csrf_token".*?value="([^"]+)"', r.text)
    if not csrf:
        raise RuntimeError("Could not find CSRF token on login page")
    r = session.post(
        f"{KOIOS_BASE}/login",
        data={"csrf_token": csrf.group(1), "email": KOIOS_EMAIL, "password": KOIOS_PASSWORD},
        timeout=30,
    )
    if r.status_code != 200 or "/login" in r.url:
        raise RuntimeError(f"Koios login failed: HTTP {r.status_code}")
    log.info("Koios web login successful")


def fetch_customers(session, site_id):
    r = session.get(
        f"{KOIOS_BASE}/sm/organizations/{ORG_ID}/customers",
        headers={"Accept": "application/json"},
        params={"page_size": 1000, "site_id": site_id},
        timeout=60,
    )
    r.raise_for_status()
    return r.json().get("customers", [])


def build_koios_mapping(all_customers):
    """Build {account_number: customer_type} from Koios tariff names."""
    mapping = {}
    for c in all_customers:
        code = (c.get("code") or "").strip()
        if not code:
            continue
        tariff = c.get("tariff")
        if not isinstance(tariff, dict):
            continue
        tname = (tariff.get("name") or "").strip()
        if not tname:
            continue
        ctype = TARIFF_TO_TYPE.get(tname, tname)
        mapping[code] = ctype
    return mapping


def main():
    parser = argparse.ArgumentParser(description="Sync BN customer types from census + Koios tariffs")
    parser.add_argument("--dry-run", action="store_true", help="Preview only, don't update DB")
    args = parser.parse_args()

    # --- Tier 1: Census spreadsheets ---
    census_map = load_census_mapping()
    if census_map:
        log.info("Tier 1 (census): %d customer types", len(census_map))

    # --- Tier 2: Koios tariff names ---
    session = requests.Session()
    log.info("Authenticating to Koios web UI...")
    koios_login(session)

    all_customers = []
    for site_code, site_id in sorted(SITES.items()):
        custs = fetch_customers(session, site_id)
        log.info("  %s: %d customers from Koios", site_code, len(custs))
        all_customers.extend(custs)

    koios_map = build_koios_mapping(all_customers)
    log.info("Tier 2 (Koios tariffs): %d customer types", len(koios_map))

    # --- Merge: census wins over Koios ---
    merged = dict(koios_map)
    merged.update(census_map)
    log.info("Merged mapping: %d customers", len(merged))

    from collections import Counter
    dist = Counter(merged.values())
    for t, n in dist.most_common():
        log.info("  %s: %d", t, n)

    if args.dry_run:
        log.info("DRY RUN — no database changes made")
        return

    conn = psycopg2.connect(DATABASE_URL)
    cur = conn.cursor()

    updated = 0
    skipped = 0
    not_found = 0

    for acct, ctype in merged.items():
        cur.execute(
            "UPDATE meters SET customer_type = %s WHERE account_number = %s AND "
            "(customer_type IS NULL OR customer_type = '' OR customer_type <> %s)",
            (ctype, acct, ctype),
        )
        if cur.rowcount > 0:
            updated += 1
        else:
            cur.execute("SELECT 1 FROM meters WHERE account_number = %s", (acct,))
            if cur.fetchone():
                skipped += 1
            else:
                not_found += 1

    conn.commit()
    cur.close()
    conn.close()

    log.info("Done: %d updated, %d already correct, %d not in meters table", updated, skipped, not_found)


if __name__ == "__main__":
    main()
