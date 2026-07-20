"""
Declarative Report Engine
=========================

Template-agnostic report definitions that compose KPI endpoint data
into Excel workbooks.  Each report is a declarative spec — the engine
fetches data from the investor analytics endpoints and renders sheets.

Report definition structure:
  {
    "id": "quarterly_investor_report",
    "title": "Quarterly Investor Report",
    "sheets": [
      {
        "name": "Portfolio Overview",
        "source": "asset-register",
        "columns": ["site_code", "full_name", "country", ...],
      },
      {
        "name": "KPI Time Series",
        "source": "kpis",
        "params": {"period": "quarter"},
        "columns": ["period", "total_connections", ...],
      },
    ]
  }
"""

from __future__ import annotations

import io
import logging
from datetime import date
from typing import Any, Dict, List, Optional

from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from pydantic import BaseModel

from models import CurrentUser, CCRole
from middleware import require_employee

logger = logging.getLogger("cc-api.report-engine")

router = APIRouter(prefix="/api/reports", tags=["reports"])


# ---------------------------------------------------------------------------
# Report definitions (declarative)
# ---------------------------------------------------------------------------

REPORT_DEFINITIONS: List[Dict[str, Any]] = [
    {
        "id": "quarterly_investor_report",
        "title": "Quarterly Investor Report",
        "description": "Portfolio overview, KPI time series, and asset register",
        "sheets": [
            {
                "name": "Portfolio Overview",
                "source": "asset-register",
                "columns": [
                    "site_code", "full_name", "country", "region", "status",
                    "commissioning_date", "pv_kwp", "battery_kwh", "thermal_kw",
                    "total_connections", "active_connections",
                    "hh_count", "sme_count", "ci_count",
                    "avg_tariff_usd_kwh", "system_availability_pct",
                    "concession_expiry", "metering_tech",
                ],
            },
            {
                "name": "KPI Time Series",
                "source": "kpis",
                "params": {"period": "quarter"},
                "columns": [
                    "period", "concession", "total_connections", "active_connections",
                    "new_connections", "energy_kwh", "revenue_usd",
                    "arpu_usd_month", "avg_tariff_usd_kwh", "productive_use_share",
                    "system_availability_pct",
                    "opex_usd", "opex_per_connection_usd",
                    "ebitda_usd", "ebitda_per_connection_usd",
                    "capex_deployed_usd", "capex_cumulative_usd", "capex_per_connection_usd",
                ],
            },
        ],
    },
    {
        "id": "monthly_operations_report",
        "title": "Monthly Operations Report",
        "description": "Monthly KPI time series with site breakdown",
        "sheets": [
            {
                "name": "Monthly KPIs",
                "source": "kpis",
                "params": {"period": "month"},
                "columns": [
                    "period", "concession", "total_connections", "active_connections",
                    "new_connections", "energy_kwh", "revenue_usd",
                    "arpu_usd_month", "avg_tariff_usd_kwh", "productive_use_share",
                    "system_availability_pct",
                ],
            },
            {
                "name": "Asset Register",
                "source": "asset-register",
                "columns": [
                    "site_code", "full_name", "country", "status",
                    "total_connections", "active_connections",
                    "avg_tariff_usd_kwh", "system_availability_pct",
                ],
            },
        ],
    },
    {
        "id": "site_detail_report",
        "title": "Site Detail Report",
        "description": "Per-site customer and transaction detail",
        "sheets": [
            {
                "name": "Asset Register",
                "source": "asset-register",
                "columns": [
                    "site_code", "full_name", "country", "region", "status",
                    "total_connections", "active_connections",
                    "hh_count", "sme_count", "ci_count",
                ],
            },
        ],
        "requires_concession": True,
        "extra_sheets": [
            {
                "name": "Customers",
                "source": "site-customers",
                "columns": [
                    "account_number", "customer_name", "customer_type",
                    "tariff_plan", "connection_date", "last_transaction_date", "status",
                ],
            },
            {
                "name": "Transactions",
                "source": "site-transactions",
                "columns": [
                    "account_number", "customer_name", "customer_type",
                    "timestamp", "kwh", "amount_local", "currency", "amount_usd",
                    "rate_used", "tariff_plan",
                ],
            },
        ],
    },
]


# ---------------------------------------------------------------------------
# Data fetchers
# ---------------------------------------------------------------------------

def _fetch_data(source: str, params: Dict[str, Any], concession: Optional[str] = None) -> List[Dict[str, Any]]:
    """Fetch data from investor analytics endpoints (internal calls)."""
    from investor_analytics import get_asset_register, get_kpis, get_site_customers, get_site_transactions
    from models import CurrentUser, UserType

    # Create a mock internal user for direct function calls
    internal_user = CurrentUser(
        user_type=UserType.employee,
        user_id="report-engine",
        role="superadmin",
        name="Report Engine",
    )

    if source == "asset-register":
        return get_asset_register(user=internal_user)
    elif source == "kpis":
        return get_kpis(
            period=params.get("period", "quarter"),
            start=params.get("start"),
            end=params.get("end"),
            concession=concession,
            user=internal_user,
        )
    elif source == "site-customers":
        if not concession:
            return []
        result = get_site_customers(
            concession=concession,
            status=None,
            customer_type=params.get("customer_type"),
            page=1,
            limit=500,
            user=internal_user,
        )
        return result.get("customers", [])
    elif source == "site-transactions":
        if not concession:
            return []
        result = get_site_transactions(
            concession=concession,
            date_from=params.get("date_from"),
            date_to=params.get("date_to"),
            customer_type=None,
            page=1,
            limit=500,
            user=internal_user,
        )
        return result.get("transactions", [])
    return []


# ---------------------------------------------------------------------------
# Excel renderer
# ---------------------------------------------------------------------------

def _render_xlsx(report: Dict[str, Any], data: Dict[str, List[Dict[str, Any]]]) -> bytes:
    """Render report data into an XLSX workbook using openpyxl."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    wb.remove(wb.active)

    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="2563eb", end_color="2563eb", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        bottom=Side(style="thin", color="CCCCCC"),
    )

    for sheet_def in report["sheets"]:
        sheet_name = sheet_def["name"][:31]  # Excel sheet name limit
        ws = wb.create_sheet(title=sheet_name)
        columns = sheet_def["columns"]
        source_key = sheet_def["name"]
        rows = data.get(source_key, [])

        # Header row
        for col_idx, col in enumerate(columns, 1):
            cell = ws.cell(row=1, column=col_idx, value=col.replace("_", " ").title())
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = thin_border

        # Data rows
        for row_idx, row_data in enumerate(rows, 2):
            for col_idx, col in enumerate(columns, 1):
                val = row_data.get(col)
                if val is not None and not isinstance(val, (str, int, float, bool)):
                    val = str(val)
                cell = ws.cell(row=row_idx, column=col_idx, value=val)
                cell.border = thin_border

        # Auto-width columns
        for col_idx in range(1, len(columns) + 1):
            max_len = len(str(columns[col_idx - 1]))
            for row_idx in range(2, min(len(rows) + 2, 100)):
                cell_val = ws.cell(row=row_idx, column=col_idx).value
                if cell_val:
                    max_len = max(max_len, len(str(cell_val)))
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 3, 40)

        # Freeze header row
        ws.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


# ---------------------------------------------------------------------------
# API endpoints
# ---------------------------------------------------------------------------

@router.get("/definitions")
def list_report_definitions(
    user: CurrentUser = Depends(require_employee),
) -> List[Dict[str, Any]]:
    """List available report definitions."""
    return [
        {
            "id": r["id"],
            "title": r["title"],
            "description": r.get("description", ""),
            "requires_concession": r.get("requires_concession", False),
            "sheet_count": len(r["sheets"]) + len(r.get("extra_sheets", [])),
        }
        for r in REPORT_DEFINITIONS
    ]


@router.get("/{report_id}/export")
def export_report(
    report_id: str,
    concession: Optional[str] = Query(None),
    period: Optional[str] = Query(None),
    start: Optional[str] = Query(None),
    end: Optional[str] = Query(None),
    user: CurrentUser = Depends(require_employee),
) -> StreamingResponse:
    """Generate and download an XLSX report."""
    report = next((r for r in REPORT_DEFINITIONS if r["id"] == report_id), None)
    if not report:
        raise HTTPException(status_code=404, detail=f"Report '{report_id}' not found")

    if report.get("requires_concession") and not concession:
        raise HTTPException(status_code=400, detail="This report requires a 'concession' parameter")

    # Build all sheets
    all_sheets = report["sheets"] + report.get("extra_sheets", [])
    data: Dict[str, List[Dict[str, Any]]] = {}

    for sheet_def in all_sheets:
        params = dict(sheet_def.get("params", {}))
        if period:
            params["period"] = period
        if start:
            params["start"] = start
        if end:
            params["end"] = end
        rows = _fetch_data(sheet_def["source"], params, concession)
        data[sheet_def["name"]] = rows

    xlsx_bytes = _render_xlsx({**report, "sheets": all_sheets}, data)

    filename = f"{report_id}_{date.today().isoformat()}.xlsx"
    return StreamingResponse(
        io.BytesIO(xlsx_bytes),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
