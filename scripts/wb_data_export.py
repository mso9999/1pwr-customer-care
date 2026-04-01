#!/usr/bin/env python3
"""
World Bank Data Export Script

Generates a comprehensive Excel workbook with 1PWR minigrid operational data
for World Bank energy team analysis of customer demand and energy usage patterns.

Usage:
    python scripts/wb_data_export.py [--api-url URL] [--output PATH] [--token TOKEN]

Options:
    --api-url URL     CC API base URL (default: http://localhost:8100)
    --output PATH     Output Excel file path (default: wb_demand_analysis_YYYYMMDD.xlsx)
    --token TOKEN     JWT token for API auth (or set CC_JWT_TOKEN env var)

Environment:
    CC_JWT_TOKEN      JWT token for API authentication
"""

__version__ = "1.0.0"

import argparse
import os
import sys
from datetime import datetime
from typing import Any, Dict, List, Optional

import requests

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill
    from openpyxl.utils import get_column_letter
except ImportError:
    print("ERROR: openpyxl is required. Install with: pip install openpyxl")
    sys.exit(1)


# Site code to full name mapping (Lesotho sites)
SITE_ABBREV = {
    "MAK": "Ha Makebe",
    "MAS": "Mashai",
    "SHG": "Sehonghong",
    "LEB": "Lebakeng",
    "SEH": "Sehlabathebe",
    "MAT": "Matsoaing",
    "TLH": "Tlhanyaku",
    "TOS": "Tosing",
    "SEB": "Sebapala",
    "RIB": "Ribaneng",
    "KET": "Ketane",
    "LSB": "Lets'eng-la-Baroa",
    "NKU": "Ha Nkau",
    "MET": "Methalaneng",
    "BOB": "Bobete",
    "MAN": "Manamaneng",
}


class APIClient:
    """Simple HTTP client for CC API with JWT auth."""

    def __init__(self, base_url: str, token: Optional[str] = None):
        self.base_url = base_url.rstrip("/")
        self.token = token
        self.session = requests.Session()
        if token:
            self.session.headers["Authorization"] = f"Bearer {token}"

    def get(self, endpoint: str, params: Optional[Dict] = None) -> Dict[str, Any]:
        """Make GET request to API endpoint."""
        url = f"{self.base_url}{endpoint}"
        resp = self.session.get(url, params=params, timeout=60)
        resp.raise_for_status()
        return resp.json()


def format_number(value: Any, decimals: int = 2) -> Any:
    """Format number with comma separators and decimal places."""
    if value is None:
        return None
    try:
        num = float(value)
        if decimals == 0:
            return int(round(num))
        return round(num, decimals)
    except (ValueError, TypeError):
        return value


def style_worksheet(ws, header_fill: Optional[PatternFill] = None):
    """Apply standard styling to worksheet: bold headers, freeze top row, auto-width."""
    if header_fill is None:
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    
    bold_white = Font(bold=True, color="FFFFFF")
    bold_black = Font(bold=True)
    
    # Style header row
    for cell in ws[1]:
        cell.font = bold_white if header_fill else bold_black
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Freeze top row
    ws.freeze_panes = "A2"
    
    # Auto-size columns based on content
    for col_idx in range(1, ws.max_column + 1):
        max_length = 0
        col_letter = get_column_letter(col_idx)
        
        for row_idx in range(1, min(ws.max_row + 1, 500)):  # Sample first 500 rows
            cell = ws.cell(row=row_idx, column=col_idx)
            if cell.value is not None:
                cell_len = len(str(cell.value))
                max_length = max(max_length, cell_len)
        
        # Set width with min/max bounds
        ws.column_dimensions[col_letter].width = min(max(max_length + 2, 10), 50)


def create_portfolio_overview_sheet(wb: Workbook, api: APIClient) -> int:
    """Sheet 1: Portfolio Overview from /api/om-report/overview."""
    print("Fetching portfolio overview...")
    
    ws = wb.active
    ws.title = "Portfolio Overview"
    
    try:
        data = api.get("/api/om-report/overview")
    except Exception as e:
        print(f"  WARNING: Failed to fetch overview: {e}")
        ws.append(["Error fetching data", str(e)])
        return 0
    
    # Write key metrics as a vertical table
    ws.append(["Metric", "Value"])
    ws.append(["Total Customers", format_number(data.get("total_customers"), 0)])
    ws.append(["Active Customers", format_number(data.get("active_customers"), 0)])
    ws.append(["Terminated Customers", format_number(data.get("terminated_customers"), 0)])
    ws.append(["Total Sites", format_number(data.get("total_sites"), 0)])
    ws.append(["Total MWh Consumed", format_number(data.get("total_mwh"), 2)])
    ws.append(["Total Revenue (LSL thousands)", format_number(data.get("total_lsl_thousands"), 2)])
    ws.append([])
    ws.append(["Sites", ", ".join(data.get("sites", []))])
    ws.append(["Revenue Data Source", data.get("data_sources", {}).get("revenue", "N/A")])
    
    style_worksheet(ws)
    
    row_count = ws.max_row - 1  # Exclude header
    print(f"  Portfolio Overview: {row_count} metrics")
    return row_count


def create_customer_growth_sheet(wb: Workbook, api: APIClient) -> int:
    """Sheet 2: Customer Growth from /api/om-report/customer-growth."""
    print("Fetching customer growth data...")
    
    ws = wb.create_sheet("Customer Growth")
    
    try:
        data = api.get("/api/om-report/customer-growth")
    except Exception as e:
        print(f"  WARNING: Failed to fetch customer growth: {e}")
        ws.append(["Error fetching data", str(e)])
        return 0
    
    growth = data.get("growth", [])
    if not growth:
        ws.append(["No customer growth data available"])
        print("  WARNING: No customer growth data")
        return 0
    
    # Headers
    ws.append(["Quarter", "New Customers", "Cumulative Customers", "Growth Rate (%)"])
    
    # Data rows with calculated growth rate
    prev_cumulative = 0
    for row in growth:
        cumulative = row.get("cumulative", 0)
        if prev_cumulative > 0:
            growth_rate = round((cumulative - prev_cumulative) / prev_cumulative * 100, 1)
        else:
            growth_rate = None  # First quarter has no previous
        
        ws.append([
            row.get("quarter"),
            format_number(row.get("new_customers"), 0),
            format_number(cumulative, 0),
            growth_rate
        ])
        prev_cumulative = cumulative
    
    style_worksheet(ws)
    
    row_count = len(growth)
    print(f"  Customer Growth: {row_count} quarters")
    
    # Validation: check monotonically increasing
    cumulatives = [r.get("cumulative", 0) for r in growth]
    if cumulatives != sorted(cumulatives):
        print("  WARNING: Cumulative customer counts are NOT monotonically increasing!")
    
    return row_count


def create_consumption_by_site_sheet(wb: Workbook, api: APIClient) -> int:
    """Sheet 3: Consumption by Site from /api/om-report/consumption-by-site."""
    print("Fetching consumption by site...")
    
    ws = wb.create_sheet("Consumption by Site")
    
    try:
        data = api.get("/api/om-report/consumption-by-site")
    except Exception as e:
        print(f"  WARNING: Failed to fetch consumption data: {e}")
        ws.append(["Error fetching data", str(e)])
        return 0
    
    sites = data.get("sites", [])
    if not sites:
        ws.append(["No consumption data available"])
        print("  WARNING: No consumption data")
        return 0
    
    # Collect all quarters across all sites
    all_quarters = set()
    for site in sites:
        all_quarters.update(site.get("quarters", {}).keys())
    sorted_quarters = sorted(all_quarters)
    
    # Headers: Site, Full Name, Total kWh, then each quarter
    headers = ["Site Code", "Site Name", "Total kWh"] + sorted_quarters
    ws.append(headers)
    
    # Data rows
    for site in sites:
        site_code = site.get("site", "")
        full_name = site.get("name") or SITE_ABBREV.get(site_code, site_code)
        row = [
            site_code,
            full_name,
            format_number(site.get("total_kwh"), 2)
        ]
        quarters = site.get("quarters", {})
        for q in sorted_quarters:
            row.append(format_number(quarters.get(q), 2))
        ws.append(row)
    
    # Add total row
    total_row = ["TOTAL", "", format_number(data.get("total_kwh"), 2)]
    for q in sorted_quarters:
        q_total = sum(s.get("quarters", {}).get(q, 0) for s in sites)
        total_row.append(format_number(q_total, 2))
    ws.append(total_row)
    
    style_worksheet(ws)
    
    row_count = len(sites)
    print(f"  Consumption by Site: {row_count} sites, {len(sorted_quarters)} quarters")
    return row_count


def create_avg_consumption_trend_sheet(wb: Workbook, api: APIClient) -> int:
    """Sheet 4: Average Consumption Trend from /api/om-report/avg-consumption-trend."""
    print("Fetching average consumption trend...")
    
    ws = wb.create_sheet("Average Consumption Trend")
    
    try:
        data = api.get("/api/om-report/avg-consumption-trend")
    except Exception as e:
        print(f"  WARNING: Failed to fetch avg consumption trend: {e}")
        ws.append(["Error fetching data", str(e)])
        return 0
    
    trends = data.get("trends", [])
    if not trends:
        ws.append(["No consumption trend data available"])
        print("  WARNING: No consumption trend data")
        return 0
    
    # Headers
    ws.append([
        "Quarter",
        "Customers",
        "Total kWh",
        "Total LSL",
        "Avg Daily kWh/Customer",
        "Avg Daily LSL/Customer"
    ])
    
    # Data rows
    for row in trends:
        ws.append([
            row.get("quarter"),
            format_number(row.get("customers"), 0),
            format_number(row.get("total_kwh"), 2),
            format_number(row.get("total_lsl"), 2),
            format_number(row.get("avg_daily_kwh_per_customer"), 4),
            format_number(row.get("avg_daily_lsl_per_customer"), 4)
        ])
    
    style_worksheet(ws)
    
    row_count = len(trends)
    print(f"  Average Consumption Trend: {row_count} quarters")
    return row_count


def create_load_curves_by_type_sheet(wb: Workbook, api: APIClient) -> int:
    """Sheet 5: Load Curves by Type from /api/om-report/load-curves-by-type."""
    print("Fetching load curves by customer type...")
    
    ws = wb.create_sheet("Load Curves by Type")
    
    try:
        data = api.get("/api/om-report/load-curves-by-type")
    except Exception as e:
        print(f"  WARNING: Failed to fetch load curves: {e}")
        ws.append(["Error fetching data", str(e)])
        return 0
    
    curves = data.get("curves", [])
    quarterly = data.get("quarterly", [])
    
    if not curves and not quarterly:
        ws.append(["No load curve data available"])
        print("  WARNING: No load curve data")
        return 0
    
    # Part 1: Summary by customer type
    ws.append(["SUMMARY BY CUSTOMER TYPE"])
    ws.append([
        "Customer Type",
        "Total kWh",
        "Total LSL",
        "Customer Count",
        "Avg Daily kWh (total)",
        "Avg Daily kWh/Customer"
    ])
    
    for curve in curves:
        ws.append([
            curve.get("type"),
            format_number(curve.get("total_kwh"), 2),
            format_number(curve.get("total_lsl"), 2),
            format_number(curve.get("customer_count"), 0),
            format_number(curve.get("avg_daily_kwh"), 4),
            format_number(curve.get("avg_daily_kwh_per_customer"), 4)
        ])
    
    # Blank row separator
    ws.append([])
    ws.append([])
    
    # Part 2: Quarterly breakdown
    if quarterly:
        customer_types = data.get("customer_types", [])
        ws.append(["QUARTERLY kWh BY CUSTOMER TYPE"])
        headers = ["Quarter"] + customer_types
        ws.append(headers)
        
        for q_data in quarterly:
            row = [q_data.get("quarter")]
            for ctype in customer_types:
                row.append(format_number(q_data.get(ctype), 2))
            ws.append(row)
    
    style_worksheet(ws)
    
    row_count = len(curves) + len(quarterly)
    print(f"  Load Curves by Type: {len(curves)} types, {len(quarterly)} quarters")
    return row_count


def create_daily_load_profiles_sheet(wb: Workbook, api: APIClient) -> int:
    """Sheet 6: Daily Load Profiles from /api/om-report/daily-load-profiles."""
    print("Fetching daily load profiles...")
    
    ws = wb.create_sheet("Daily Load Profiles")
    
    try:
        data = api.get("/api/om-report/daily-load-profiles")
    except Exception as e:
        print(f"  WARNING: Failed to fetch daily load profiles: {e}")
        ws.append(["Error fetching data", str(e)])
        return 0
    
    profiles = data.get("profiles", [])
    chart_data = data.get("chart_data", [])
    
    if not chart_data:
        ws.append(["No daily load profile data available"])
        print("  WARNING: No daily load profile data")
        return 0
    
    # Part 1: Summary by customer type with peak info
    ws.append(["PROFILE SUMMARY BY CUSTOMER TYPE"])
    ws.append(["Customer Type", "Meter Count", "Peak Hour", "Peak kW"])
    
    for profile in profiles:
        peak_hour = profile.get("peak_hour", 0)
        ws.append([
            profile.get("type"),
            format_number(profile.get("meter_count"), 0),
            f"{peak_hour:02d}:00",
            format_number(profile.get("peak_kw"), 4)
        ])
    
    ws.append([])
    ws.append([])
    
    # Part 2: 24-hour profile data
    customer_types = data.get("customer_types", [])
    ws.append(["24-HOUR AVERAGE POWER (kW) BY CUSTOMER TYPE"])
    headers = ["Hour"] + customer_types
    ws.append(headers)
    
    for point in chart_data:
        row = [point.get("hour")]
        for ctype in customer_types:
            row.append(format_number(point.get(ctype), 4))
        ws.append(row)
    
    style_worksheet(ws)
    
    # Validation: should have 24 rows for hours 0-23
    if len(chart_data) != 24:
        print(f"  WARNING: Expected 24 hourly rows, got {len(chart_data)}")
    
    # Check for evening peak (typically 18:00-21:00)
    peak_hours = []
    for profile in profiles:
        peak_hour = profile.get("peak_hour", 0)
        if 18 <= peak_hour <= 21:
            peak_hours.append(peak_hour)
    if peak_hours:
        print(f"  Peak hours in expected range (18:00-21:00): {peak_hours}")
    else:
        print("  INFO: Peak hours may not be in typical evening range (18:00-21:00)")
    
    row_count = len(chart_data)
    print(f"  Daily Load Profiles: {row_count} hours, {len(profiles)} customer types")
    return row_count


def create_arpu_sheet(wb: Workbook, api: APIClient) -> int:
    """Sheet 7: ARPU from /api/om-report/arpu and /api/om-report/monthly-arpu."""
    print("Fetching ARPU data...")
    
    ws = wb.create_sheet("ARPU")
    
    # Fetch quarterly ARPU
    try:
        quarterly_data = api.get("/api/om-report/arpu")
    except Exception as e:
        print(f"  WARNING: Failed to fetch quarterly ARPU: {e}")
        quarterly_data = {"arpu": []}
    
    # Fetch monthly ARPU
    try:
        monthly_data = api.get("/api/om-report/monthly-arpu")
    except Exception as e:
        print(f"  WARNING: Failed to fetch monthly ARPU: {e}")
        monthly_data = {"monthly_arpu": []}
    
    quarterly = quarterly_data.get("arpu", [])
    monthly = monthly_data.get("monthly_arpu", [])
    
    if not quarterly and not monthly:
        ws.append(["No ARPU data available"])
        print("  WARNING: No ARPU data")
        return 0
    
    # Part 1: Quarterly ARPU
    ws.append(["QUARTERLY ARPU (Average Revenue Per User)"])
    ws.append(["Quarter", "Total Revenue (LSL)", "Active Customers", "ARPU (LSL)"])
    
    for row in quarterly:
        ws.append([
            row.get("quarter"),
            format_number(row.get("total_revenue"), 2),
            format_number(row.get("active_customers"), 0),
            format_number(row.get("arpu"), 2)
        ])
    
    ws.append([])
    ws.append([])
    
    # Part 2: Monthly ARPU
    ws.append(["MONTHLY ARPU"])
    ws.append(["Month", "Quarter", "Total Revenue (LSL)", "Active Customers", "ARPU (LSL)"])
    
    for row in monthly:
        arpu = row.get("arpu", 0)
        ws.append([
            row.get("month"),
            row.get("quarter"),
            format_number(row.get("total_revenue"), 2),
            format_number(row.get("active_customers"), 0),
            format_number(arpu, 2)
        ])
        
        # Validation: ARPU should be reasonable for households (20-60 LSL/month)
        if arpu and (arpu < 5 or arpu > 200):
            print(f"  INFO: Month {row.get('month')} ARPU {arpu:.2f} LSL may be outside typical range")
    
    style_worksheet(ws)
    
    row_count = len(quarterly) + len(monthly)
    print(f"  ARPU: {len(quarterly)} quarters, {len(monthly)} months")
    return row_count


def create_revenue_by_site_sheet(wb: Workbook, api: APIClient) -> int:
    """Sheet 8: Revenue by Site from /api/om-report/sales-by-site."""
    print("Fetching revenue by site...")
    
    ws = wb.create_sheet("Revenue by Site")
    
    try:
        data = api.get("/api/om-report/sales-by-site")
    except Exception as e:
        print(f"  WARNING: Failed to fetch revenue data: {e}")
        ws.append(["Error fetching data", str(e)])
        return 0
    
    sites = data.get("sites", [])
    if not sites:
        ws.append(["No revenue data available"])
        print("  WARNING: No revenue data")
        return 0
    
    # Collect all quarters across all sites
    all_quarters = set()
    for site in sites:
        all_quarters.update(site.get("quarters", {}).keys())
    sorted_quarters = sorted(all_quarters)
    
    # Headers: Site, Full Name, Total LSL, then each quarter
    headers = ["Site Code", "Site Name", "Total LSL"] + sorted_quarters
    ws.append(headers)
    
    # Data rows
    for site in sites:
        site_code = site.get("site", "")
        full_name = site.get("name") or SITE_ABBREV.get(site_code, site_code)
        row = [
            site_code,
            full_name,
            format_number(site.get("total_lsl"), 2)
        ]
        quarters = site.get("quarters", {})
        for q in sorted_quarters:
            row.append(format_number(quarters.get(q), 2))
        ws.append(row)
    
    # Add total row
    total_row = ["TOTAL", "", format_number(data.get("total_lsl"), 2)]
    for q in sorted_quarters:
        q_total = sum(s.get("quarters", {}).get(q, 0) for s in sites)
        total_row.append(format_number(q_total, 2))
    ws.append(total_row)
    
    style_worksheet(ws)
    
    row_count = len(sites)
    print(f"  Revenue by Site: {row_count} sites, {len(sorted_quarters)} quarters")
    return row_count


def generate_workbook(api: APIClient, output_path: str):
    """Generate the complete Excel workbook with all 8 sheets."""
    print(f"\n{'='*60}")
    print("World Bank Data Export - 1PWR Lesotho Minigrid Analysis")
    print(f"{'='*60}\n")
    
    wb = Workbook()
    sheet_results = {}
    
    # Create each sheet
    sheet_results["Portfolio Overview"] = create_portfolio_overview_sheet(wb, api)
    sheet_results["Customer Growth"] = create_customer_growth_sheet(wb, api)
    sheet_results["Consumption by Site"] = create_consumption_by_site_sheet(wb, api)
    sheet_results["Average Consumption Trend"] = create_avg_consumption_trend_sheet(wb, api)
    sheet_results["Load Curves by Type"] = create_load_curves_by_type_sheet(wb, api)
    sheet_results["Daily Load Profiles"] = create_daily_load_profiles_sheet(wb, api)
    sheet_results["ARPU"] = create_arpu_sheet(wb, api)
    sheet_results["Revenue by Site"] = create_revenue_by_site_sheet(wb, api)
    
    # Save workbook
    wb.save(output_path)
    
    # Summary
    print(f"\n{'='*60}")
    print("EXPORT SUMMARY")
    print(f"{'='*60}")
    print(f"Output file: {output_path}")
    print(f"\nSheet row counts:")
    
    empty_sheets = []
    for sheet_name, count in sheet_results.items():
        status = "OK" if count > 0 else "EMPTY"
        print(f"  {sheet_name}: {count} rows [{status}]")
        if count == 0:
            empty_sheets.append(sheet_name)
    
    if empty_sheets:
        print(f"\nWARNING: {len(empty_sheets)} sheet(s) are empty: {', '.join(empty_sheets)}")
    
    print(f"\nTotal sheets: {len(sheet_results)}")
    print(f"{'='*60}\n")
    
    return len(empty_sheets) == 0


def main():
    parser = argparse.ArgumentParser(
        description="Generate World Bank data export from CC API",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog=__doc__
    )
    parser.add_argument(
        "--api-url",
        default=os.environ.get("CC_API_URL", "http://localhost:8100"),
        help="CC API base URL (default: http://localhost:8100 or CC_API_URL env)"
    )
    parser.add_argument(
        "--output",
        default=None,
        help="Output Excel file path (default: wb_demand_analysis_YYYYMMDD.xlsx)"
    )
    parser.add_argument(
        "--token",
        default=os.environ.get("CC_JWT_TOKEN"),
        help="JWT token for API auth (or set CC_JWT_TOKEN env var)"
    )
    
    args = parser.parse_args()
    
    # Default output filename with date
    if args.output is None:
        date_str = datetime.now().strftime("%Y%m%d")
        args.output = f"wb_demand_analysis_{date_str}.xlsx"
    
    # Validate token
    if not args.token:
        print("ERROR: JWT token required for API authentication.")
        print("       Pass via --token argument or set CC_JWT_TOKEN environment variable.")
        print("\n       To get a token, authenticate at the CC portal and extract from browser.")
        sys.exit(1)
    
    # Create API client
    api = APIClient(args.api_url, args.token)
    
    # Test API connectivity
    print(f"Testing API connectivity to {args.api_url}...")
    try:
        api.get("/api/health")
        print("  API health check: OK")
    except requests.exceptions.ConnectionError:
        print(f"ERROR: Cannot connect to API at {args.api_url}")
        print("       Make sure the CC API is running.")
        sys.exit(1)
    except requests.exceptions.HTTPError as e:
        if e.response.status_code == 401:
            print("ERROR: Authentication failed. Token may be expired or invalid.")
            sys.exit(1)
        # Health endpoint might not exist, continue anyway
        print(f"  API health check: {e.response.status_code} (continuing anyway)")
    
    # Generate workbook
    success = generate_workbook(api, args.output)
    
    if success:
        print(f"SUCCESS: Export complete. File saved to: {args.output}")
        sys.exit(0)
    else:
        print("COMPLETED with warnings (some sheets may be empty)")
        sys.exit(0)


if __name__ == "__main__":
    main()
