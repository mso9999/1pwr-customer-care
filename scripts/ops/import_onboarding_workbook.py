#!/usr/bin/env python3
"""Import Customer Onboarding workbook into 1PDB commissioning + fee status."""

from __future__ import annotations

import argparse
import csv
import logging
import os
import re
import sys
from collections import Counter
from datetime import date, datetime
from pathlib import Path
from typing import Any, Optional

ROOT = Path(__file__).resolve().parents[2] if len(Path(__file__).resolve().parents) > 2 else Path("/")
ACDB_API = Path(os.environ.get("ACDB_API", "/opt/cc-portal/backend"))
if not (ACDB_API / "customer_api.py").exists():
    ACDB_API = ROOT / "acdb-api"
if str(ACDB_API) not in sys.path:
    sys.path.insert(0, str(ACDB_API))

COMMISSIONING_STEPS = (
    "connection_fee_paid",
    "readyboard_fee_paid",
    "readyboard_tested",
    "readyboard_installed",
    "airdac_connected",
    "meter_installed",
    "customer_commissioned",
)

logging.basicConfig(level=logging.INFO, format="%(asctime)s [%(levelname)s] %(message)s")
log = logging.getLogger("import_onboarding_workbook")

ACCOUNT_RE = re.compile(r"^\d{3,4}[A-Z]{2,4}$")
IMPORT_TAG = "onboarding_import_2026-01"
DEFAULT_WORKBOOK = Path(
    "/Users/mattmso/Dropbox/1PWR/1PWR OM TEAM/20. Dashboards/"
    "Customer Onboarding Data_Jan-01, 2026.xlsx"
)


def _connect():
    import psycopg2

    url = os.environ.get("DATABASE_URL")
    if not url:
        raise SystemExit("DATABASE_URL is required")
    return psycopg2.connect(url)


def _norm_account(value: Any) -> str:
    return str(value or "").strip().upper().replace(" ", "")


def _as_date(value: Any) -> Optional[date]:
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    try:
        return datetime.fromisoformat(str(value)[:10]).date()
    except ValueError:
        return None


def _yn(value: Any) -> Optional[bool]:
    if value is None or value == "":
        return None
    text = str(value).strip().lower()
    if text in {"y", "yes", "true", "1", "pass"}:
        return True
    if text in {"n", "no", "false", "0", "fail"}:
        return False
    return None


def _header_map(header_row: tuple[Any, ...]) -> dict[str, int]:
    mapping: dict[str, int] = {}
    for idx, cell in enumerate(header_row):
        if cell is None:
            continue
        key = str(cell).strip()
        if key:
            mapping[key] = idx
    return mapping


def _cell(row: tuple[Any, ...], headers: dict[str, int], name: str) -> Any:
    idx = headers.get(name)
    if idx is None or idx >= len(row):
        return None
    return row[idx]


def _open_sheet_rows(path: Path, sheet_name: str) -> list[tuple[Any, ...]]:
    import openpyxl

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    if sheet_name not in wb.sheetnames:
        wb.close()
        return []
    rows = list(wb[sheet_name].iter_rows(values_only=True))
    wb.close()
    return rows


def load_customer_records(path: Path) -> list[dict[str, Any]]:
    rows = _open_sheet_rows(path, "Customer Records")
    if len(rows) < 3:
        return []
    headers = _header_map(rows[1])
    records: list[dict[str, Any]] = []
    for row in rows[2:]:
        account = _norm_account(_cell(row, headers, "Customer ID"))
        if not ACCOUNT_RE.match(account):
            continue
        records.append(
            {
                "account_number": account,
                "community": _cell(row, headers, "Concession"),
                "first_name": _cell(row, headers, "First name"),
                "last_name": _cell(row, headers, "Last name"),
                "national_id": _cell(row, headers, "ID number"),
                "phone_number": _cell(row, headers, "Phone Number"),
                "survey_id": _cell(row, headers, "PLOT NUMBERS"),
                "connection_fee_amount": _cell(row, headers, "Connection Fee Amount"),
                "connection_fee_paid": _as_date(_cell(row, headers, "Date Paid_CF")),
                "connection_fee_txn": _cell(row, headers, "Transaction ID_CF"),
                "readyboard_fee_amount": _cell(row, headers, "Readyboard Payment Amount"),
                "readyboard_fee_paid": _as_date(_cell(row, headers, "Date Paid_RB")),
                "readyboard_fee_txn": _cell(row, headers, "Transaction ID_RB"),
                "readyboard_installed": _yn(_cell(row, headers, "Readyboard (Y/N)")),
                "readyboard_installed_date": _as_date(_cell(row, headers, "Date Installed")),
                "readyboard_tested": _yn(_cell(row, headers, "Readyboard Test (Pass/Fail)")),
                "readyboard_tested_date": _as_date(_cell(row, headers, "Date of Test")),
                "house_wiring_test_passed": _yn(_cell(row, headers, "House Wiring Test (Pass/Fail)")),
                "house_wiring_test_date": _as_date(_cell(row, headers, "Date of Test_HW")),
                "airdac_connected": _yn(_cell(row, headers, "Airdac (Y/N)")),
                "airdac_connected_date": _as_date(_cell(row, headers, "Date Installed_AD")),
                "meter_installed": _yn(_cell(row, headers, "Smartmeter (Y/N)")),
                "meter_installed_date": _as_date(_cell(row, headers, "Date Installed_SM")),
                "meter_serial": _cell(row, headers, "Smartmeter number"),
                "customer_commissioned": _yn(_cell(row, headers, "Commisioned?(Y/N)")),
                "customer_commissioned_date": _as_date(_cell(row, headers, "Commissioning Date")),
                "notes": _cell(row, headers, "Notes"),
                "proof_urls": [],
            }
        )
    return records


def load_mak_plot_crosswalk(path: Path) -> dict[str, dict[str, Any]]:
    rows = _open_sheet_rows(path, "Mak plot numbers")
    if len(rows) < 2:
        return {}
    headers = _header_map(rows[0])
    crosswalk: dict[str, dict[str, Any]] = {}
    for row in rows[1:]:
        account = _norm_account(
            _cell(row, headers, "Customer Code")
            or _cell(row, headers, "Customer ID")
            or _cell(row, headers, "Account")
        )
        if not ACCOUNT_RE.match(account):
            continue
        crosswalk[account] = {
            "survey_id": _cell(row, headers, "Plot No") or _cell(row, headers, "Plot Number"),
            "meter_serial": _cell(row, headers, "Meter Code") or _cell(row, headers, "Meter"),
        }
    return crosswalk


def load_mak_records(path: Path) -> dict[str, dict[str, Any]]:
    rows = _open_sheet_rows(path, "MAK records")
    if len(rows) < 2:
        return {}
    headers = _header_map(rows[0])
    mak: dict[str, dict[str, Any]] = {}
    for row in rows[1:]:
        account = _norm_account(
            _cell(row, headers, "Customer Code")
            or _cell(row, headers, "Customer ID")
        )
        if not ACCOUNT_RE.match(account):
            continue
        urls = []
        for col in headers:
            if "dropbox" in col.lower() or "proof" in col.lower() or "contract" in col.lower():
                value = _cell(row, headers, col)
                if value and str(value).startswith("http"):
                    urls.append(str(value).strip())
        mak[account] = {
            "connection_fee_paid": _as_date(
                _cell(row, headers, "Date of payment of Connection fee")
                or _cell(row, headers, "Connection fee date")
            ),
            "readyboard_fee_paid": _as_date(
                _cell(row, headers, "Date of Readyboard Payment")
                or _cell(row, headers, "Readyboard payment date")
            ),
            "proof_urls": urls,
        }
    return mak


def merge_workbook_sheets(records: list[dict[str, Any]], path: Path) -> None:
    crosswalk = load_mak_plot_crosswalk(path)
    mak = load_mak_records(path)
    for record in records:
        account = record["account_number"]
        plot = crosswalk.get(account)
        if plot:
            record["survey_id"] = record.get("survey_id") or plot.get("survey_id")
            record["meter_serial"] = record.get("meter_serial") or plot.get("meter_serial")
        mak_row = mak.get(account)
        if not mak_row:
            continue
        record["connection_fee_paid"] = record.get("connection_fee_paid") or mak_row.get("connection_fee_paid")
        record["readyboard_fee_paid"] = record.get("readyboard_fee_paid") or mak_row.get("readyboard_fee_paid")
        record["proof_urls"] = list(dict.fromkeys(record.get("proof_urls", []) + mak_row.get("proof_urls", [])))


def _customer_snapshot(cur, account: str) -> Optional[dict[str, Any]]:
    cur.execute(
        """
        SELECT c.id, c.onboarding_import_tag,
               c.connection_fee_paid, c.readyboard_fee_paid, c.customer_commissioned
        FROM accounts a
        JOIN customers c ON c.id = a.customer_id
        WHERE a.account_number = %s
        LIMIT 1
        """,
        (account,),
    )
    row = cur.fetchone()
    if not row:
        return None
    return {
        "customer_id": int(row[0]),
        "onboarding_import_tag": row[1],
        "connection_fee_paid": bool(row[2]),
        "readyboard_fee_paid": bool(row[3]),
        "customer_commissioned": bool(row[4]),
    }


def _resolve_fee_transaction(
    cur,
    account: str,
    payment_type: str,
    txn_ref: Any,
) -> Optional[int]:
    ref = str(txn_ref or "").strip()
    if ref:
        cur.execute(
            """
            SELECT id FROM transactions
            WHERE account_number = %s
              AND payment_reference = %s
            ORDER BY id DESC
            LIMIT 1
            """,
            (account, ref),
        )
        row = cur.fetchone()
        if row:
            return int(row[0])
        cur.execute(
            """
            SELECT id FROM transactions
            WHERE account_number = %s
              AND payment_reference ILIKE %s
            ORDER BY id DESC
            LIMIT 1
            """,
            (account, f"%{ref}%"),
        )
        row = cur.fetchone()
        if row:
            return int(row[0])
    cur.execute(
        """
        SELECT t.id
        FROM transactions t
        WHERE t.account_number = %s
          AND COALESCE(t.payment_category, '') = %s
          AND COALESCE(t.payment_reference, '') LIKE 'mm:%%'
        ORDER BY t.transaction_date DESC NULLS LAST, t.id DESC
        LIMIT 1
        """,
        (account, payment_type),
    )
    row = cur.fetchone()
    return int(row[0]) if row else None


def _upsert_fee_verification(
    cur,
    account: str,
    payment_type: str,
    amount: Any,
    paid_date: Optional[date],
    txn_ref: Any,
    *,
    apply: bool,
) -> str:
    txn_id = _resolve_fee_transaction(cur, account, payment_type, txn_ref)
    if not txn_id:
        return "fee_txn_unresolved"
    cur.execute(
        """
        SELECT id, status FROM payment_verifications
        WHERE transaction_id = %s AND payment_type = %s
        ORDER BY id DESC
        LIMIT 1
        """,
        (txn_id, payment_type),
    )
    row = cur.fetchone()
    if row:
        if apply and row[1] != "verified":
            cur.execute(
                """
                UPDATE payment_verifications
                SET status = 'verified',
                    verified_by = 'onboarding_import',
                    verified_at = COALESCE(%s::timestamp, NOW()),
                    note = COALESCE(note, 'onboarding workbook import')
                WHERE id = %s
                """,
                (paid_date, row[0]),
            )
        return "fee_verification_linked"
    if not apply:
        return "fee_verification_would_create"
    cur.execute(
        """
        INSERT INTO payment_verifications
            (transaction_id, account_number, payment_type, amount, status, verified_by, verified_at, note)
        VALUES (%s, %s, %s, %s, 'verified', 'onboarding_import', COALESCE(%s::timestamp, NOW()), %s)
        """,
        (
            txn_id,
            account,
            payment_type,
            float(amount or 0),
            paid_date,
            "onboarding workbook import",
        ),
    )
    return "fee_verification_created"


def _store_proof_urls(cur, customer_id: int, urls: list[str], *, apply: bool) -> None:
    if not urls or not apply:
        return
    for url in urls:
        cur.execute(
            """
            SELECT id FROM payment_proofs
            WHERE customer_id = %s AND external_url = %s
            LIMIT 1
            """,
            (customer_id, url),
        )
        if cur.fetchone():
            continue
        cur.execute(
            """
            INSERT INTO payment_proofs
                (customer_id, file_path, file_name, content_type, size_bytes, sha256,
                 uploaded_by, note, external_url)
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)
            """,
            (
                customer_id,
                "",
                "workbook-link",
                "text/uri-list",
                0,
                "",
                "onboarding_import",
                "MAK workbook proof/contract link",
                url,
            ),
        )


def _table_has_column(cur, table: str, column: str) -> bool:
    cur.execute(
        """
        SELECT 1
        FROM information_schema.columns
        WHERE table_schema = 'public' AND table_name = %s AND column_name = %s
        LIMIT 1
        """,
        (table, column),
    )
    return cur.fetchone() is not None


def apply_record(
    conn,
    record: dict[str, Any],
    *,
    apply: bool,
    force: bool,
) -> dict[str, str]:
    account = record["account_number"]
    cur = conn.cursor()
    snapshot = _customer_snapshot(cur, account)
    if not snapshot:
        return {"account_number": account, "outcome": "unmatched_account"}
    customer_id = snapshot["customer_id"]
    if snapshot.get("onboarding_import_tag") == IMPORT_TAG and not force:
        return {
            "account_number": account,
            "outcome": "skipped_import_tag",
            "customer_id": str(customer_id),
        }
    if not apply:
        return {
            "account_number": account,
            "outcome": "would_update",
            "customer_id": str(customer_id),
            "before_connection_fee_paid": str(snapshot["connection_fee_paid"]),
            "before_readyboard_fee_paid": str(snapshot["readyboard_fee_paid"]),
            "before_commissioned": str(snapshot["customer_commissioned"]),
        }

    fee_notes: list[str] = []
    if record.get("connection_fee_paid"):
        fee_notes.append(
            _upsert_fee_verification(
                cur,
                account,
                "connection_fee",
                record.get("connection_fee_amount"),
                record.get("connection_fee_paid"),
                record.get("connection_fee_txn"),
                apply=apply,
            )
        )
    if record.get("readyboard_fee_paid"):
        fee_notes.append(
            _upsert_fee_verification(
                cur,
                account,
                "readyboard_fee",
                record.get("readyboard_fee_amount"),
                record.get("readyboard_fee_paid"),
                record.get("readyboard_fee_txn"),
                apply=apply,
            )
        )

    for step in COMMISSIONING_STEPS:
        date_key = f"{step}_date"
        if step in ("connection_fee_paid", "readyboard_fee_paid"):
            paid_date = record.get(step)
            if paid_date:
                cur.execute(
                    f"UPDATE customers SET {step} = TRUE, {date_key} = %s WHERE id = %s",
                    (paid_date, customer_id),
                )
            continue
        flag = record.get(step)
        if flag is True:
            cur.execute(
                f"UPDATE customers SET {step} = TRUE, {date_key} = COALESCE(%s, {date_key}, CURRENT_DATE) WHERE id = %s",
                (record.get(date_key), customer_id),
            )
        elif flag is False:
            cur.execute(
                f"UPDATE customers SET {step} = FALSE, {date_key} = NULL WHERE id = %s",
                (customer_id,),
            )

    sets = ["onboarding_import_tag = %s", "updated_at = NOW()"]
    params: list[Any] = [IMPORT_TAG, customer_id]
    if _table_has_column(cur, "customers", "house_wiring_test_passed"):
        sets.insert(0, "house_wiring_test_date = COALESCE(%s, house_wiring_test_date)")
        sets.insert(0, "house_wiring_test_passed = COALESCE(%s, house_wiring_test_passed)")
        params = [
            record.get("house_wiring_test_passed"),
            record.get("house_wiring_test_date"),
            IMPORT_TAG,
            customer_id,
        ]
    cur.execute(
        f"UPDATE customers SET {', '.join(sets)} WHERE id = %s",
        params,
    )
    if record.get("survey_id") and _table_has_column(cur, "accounts", "survey_id"):
        cur.execute(
            "UPDATE accounts SET survey_id = COALESCE(%s, survey_id) WHERE account_number = %s",
            (str(record["survey_id"])[:64], account),
        )
    if record.get("meter_serial"):
        meter_col = "meter_id" if _table_has_column(cur, "meters", "meter_id") else (
            "meter_serial" if _table_has_column(cur, "meters", "meter_serial") else None
        )
        if meter_col:
            cur.execute(
                f"""
                UPDATE meters SET {meter_col} = %s
                WHERE account_number = %s
                  AND ({meter_col} IS NULL OR {meter_col} = '')
                """,
                (str(record["meter_serial"])[:64], account),
            )
    _store_proof_urls(cur, customer_id, record.get("proof_urls", []), apply=apply)

    after = _customer_snapshot(cur, account) or snapshot
    return {
        "account_number": account,
        "outcome": "updated",
        "customer_id": str(customer_id),
        "before_connection_fee_paid": str(snapshot["connection_fee_paid"]),
        "after_connection_fee_paid": str(after["connection_fee_paid"]),
        "before_readyboard_fee_paid": str(snapshot["readyboard_fee_paid"]),
        "after_readyboard_fee_paid": str(after["readyboard_fee_paid"]),
        "before_commissioned": str(snapshot["customer_commissioned"]),
        "after_commissioned": str(after["customer_commissioned"]),
        "fee_actions": ";".join(fee_notes),
    }


def write_report(path: Path, rows: list[dict[str, str]]) -> None:
    if not rows:
        return
    fieldnames = sorted({k for row in rows for k in row.keys()})
    with path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--workbook", type=Path, default=DEFAULT_WORKBOOK)
    parser.add_argument("--report-csv", type=Path, default=Path("/tmp/onboarding_import.csv"))
    parser.add_argument("--apply", action="store_true")
    parser.add_argument("--force", action="store_true", help="Re-apply rows already tagged with import tag")
    args = parser.parse_args()

    if not args.workbook.exists():
        log.error("Workbook not found: %s", args.workbook)
        return 2

    records = load_customer_records(args.workbook)
    merge_workbook_sheets(records, args.workbook)
    log.info("Loaded %d customer rows from workbook", len(records))
    conn = _connect()
    results: list[dict[str, str]] = []
    try:
        for record in records:
            results.append(apply_record(conn, record, apply=args.apply, force=args.force))
        if args.apply:
            conn.commit()
    finally:
        conn.close()
    write_report(args.report_csv, results)
    log.info("Summary: %s", dict(Counter(r["outcome"] for r in results)))
    log.info("Report: %s", args.report_csv)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
