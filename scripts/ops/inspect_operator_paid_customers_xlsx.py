#!/usr/bin/env python3
"""
Inspect operator "paid customers" workbooks (e.g. docs/MAS, MAK & SEH Paid Customers *.xlsx).

Expected header row (case-insensitive substring match):
  Customer ID | Connection Fee Amount | Date Paid_CF | Transaction ID_CF |
  Readyboard Payment Amount | Date Paid_RB | Transaction ID_RB

Use project venv so openpyxl resolves:
  ./acdb-api/.venv/bin/python scripts/ops/inspect_operator_paid_customers_xlsx.py <path.xlsx>

This script does not write to 1PDB or SparkMeter. For a full backfill, extend
``backfill_merchant_payments_from_exports.py`` patterns (dedup by receipt /
amount / account) only after finance sign-off.
"""

from __future__ import annotations

import argparse
import re
import sys
from pathlib import Path


def _norm(s: object) -> str:
    return re.sub(r"\s+", " ", str(s or "").strip().lower())


def _find_header_row(ws, max_scan: int = 5):
    for r in range(1, max_scan + 1):
        vals = [_norm(ws.cell(r, c).value) for c in range(1, ws.max_column + 1)]
        joined = " ".join(v for v in vals if v)
        if "customer id" in joined and "connection fee" in joined:
            return r
    return None


def _col_map(header_row: list[object]) -> dict[str, int]:
    """1-based column index by canonical key."""
    m: dict[str, int] = {}
    aliases = {
        "account": ("customer id",),
        "cf_amt": ("connection fee amount",),
        "cf_date": ("date paid_cf", "date paid cf"),
        "cf_txn": ("transaction id_cf", "transaction id cf"),
        "rb_amt": ("readyboard payment amount",),
        "rb_date": ("date paid_rb", "date paid rb"),
        "rb_txn": ("transaction id_rb", "transaction id rb"),
    }
    for col, raw in enumerate(header_row, start=1):
        h = _norm(raw)
        if not h:
            continue
        for key, needles in aliases.items():
            if key in m:
                continue
            for n in needles:
                if h == n or (n in h and len(h) < 80):
                    m[key] = col
                    break
    return m


def main() -> int:
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument("xlsx", type=Path, help="Path to .xlsx")
    ap.add_argument("--max-rows", type=int, default=0, help="Limit data rows (0 = all)")
    args = ap.parse_args()

    try:
        import openpyxl
    except ImportError:
        print("Install openpyxl (see acdb-api/requirements.txt) or use acdb-api/.venv/bin/python", file=sys.stderr)
        return 2

    if not args.xlsx.is_file():
        print(f"Not found: {args.xlsx}", file=sys.stderr)
        return 1

    wb = openpyxl.load_workbook(args.xlsx, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    hr = _find_header_row(ws)
    if hr is None:
        print("Could not locate header row (need Customer ID + Connection Fee Amount).", file=sys.stderr)
        return 1

    header = [ws.cell(hr, c).value for c in range(1, ws.max_column + 1)]
    cmap = _col_map(header)
    print("Header row:", hr)
    print("Column map:", cmap)
    need = ("account", "cf_amt", "cf_txn")
    missing = [k for k in need if k not in cmap]
    if missing:
        print("Missing required columns:", missing, file=sys.stderr)
        return 1

    n = 0
    cf_rows = rb_rows = 0
    for r in range(hr + 1, ws.max_row + 1):
        if args.max_rows and n >= args.max_rows:
            break
        acct = ws.cell(r, cmap["account"]).value
        if acct is None or str(acct).strip() == "":
            continue
        n += 1
        cf_amt = ws.cell(r, cmap["cf_amt"]).value
        cf_txn = ws.cell(r, cmap.get("cf_txn", 0)).value if "cf_txn" in cmap else None
        rb_amt = ws.cell(r, cmap["rb_amt"]).value if "rb_amt" in cmap else None
        rb_txn = ws.cell(r, cmap["rb_txn"]).value if "rb_txn" in cmap else None
        if cf_amt not in (None, "", 0):
            cf_rows += 1
        if rb_amt not in (None, "", 0):
            rb_rows += 1
        if n <= 5:
            print(
                f"  sample r{r}: account={acct!r} CF amt={cf_amt!r} ref={cf_txn!r} "
                f"RB amt={rb_amt!r} ref={rb_txn!r}"
            )

    print(f"--- dry-run summary ---\nData rows scanned: {n}\nRows with CF amount: {cf_rows}\nRows with RB amount: {rb_rows}")
    wb.close()
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
