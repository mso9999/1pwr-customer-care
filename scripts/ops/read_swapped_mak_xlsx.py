#!/usr/bin/env python3
"""
Dump columns and rows from docs/ops/SWAPPED MAK CUSTOMERS.xlsx (when present in repo).

Usage:
  pip install openpyxl   # if needed
  python3 scripts/ops/read_swapped_mak_xlsx.py

Cross-check on CC server (after SSH): compare to ThunderCloud via fix_mak_drift.py / rca_mak_drift.py.
"""
from pathlib import Path

try:
    import openpyxl
except ImportError:
    raise SystemExit("Install: pip install openpyxl")

ROOT = Path(__file__).resolve().parents[2]
XLSX = ROOT / "docs" / "ops" / "SWAPPED MAK CUSTOMERS.xlsx"


def main():
    if not XLSX.is_file():
        print(f"Missing file: {XLSX}")
        print("Commit and push the workbook from your Mac, then pull here.")
        return 1
    wb = openpyxl.load_workbook(XLSX, read_only=True, data_only=True)
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        print(f"\n=== Sheet: {sheet} ===")
        rows = list(ws.iter_rows(values_only=True))
        for i, row in enumerate(rows[:50]):
            print(i + 1, row)
        if len(rows) > 50:
            print(f"... ({len(rows) - 50} more rows)")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
