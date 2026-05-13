#!/usr/bin/env python3
"""
Backfill operator "paid customers" workbooks (MAS / MAK / SEH style) into 1PDB.

Balance invariant (non-negotiable)
----------------------------------
``get_balance_kwh()`` sums ``CASE WHEN is_payment THEN kwh_value ELSE 0 END``.
This script **never** inserts a payment row with a non-zero ``kwh_value``.
Historical fee rows use :func:`balance_engine.record_fee_transaction`, which sets
``kwh_value = 0`` (same as merchant fee backfill and portal fee recording), so
the kWh ledger is unchanged. **SparkMeter credit is off by default** and this
script does not call ``credit_sparkmeter``.

Dedup keys (per fee slot)
-------------------------
1. Account must exist in ``accounts``.
2. ``payment_verifications``: existing **verified** row for the same
   ``(account_number, payment_type)`` → skip (finance already marked paid).
3. ``transactions.payment_reference`` (case-insensitive trim) globally; if the
   ref exists for another account or amount mismatch → skip with ``conflict``.
4. ``sms_inbound_log.receipt_key`` → skip (SMS path already ingested).
5. Fuzzy duplicate: same account, ``payment_category``, amount (±1c), and time
   window ±24h on an existing payment txn → skip.

Operator policy (``--require-payment-reference``)
---------------------------------------------------
When this flag is set, fee slots whose workbook **Transaction ID_CF** / **Transaction ID_RB**
(cell trimmed via the same normalization as dedup) is empty are **skipped** before any DB write.
Rows **with** payment references can be auto-verified without separate finance sign-off;
**kWh is never changed** (fee rows use ``kwh_value = 0``).

Apply order (each accepted slot)
--------------------------------
1. ``record_fee_transaction`` (fee category, ``kwh_value=0``, provenance in
   ``source_table``).
2. ``INSERT payment_verifications`` (pending).
3. Auto-verify that row (same pattern as merchant export backfill).
4. ``apply_fee_payment_category_to_debt`` to decrement ``customers.fee_debt_*``.
5. ``derive_payment_steps_for_accounts`` (best-effort onboarding sync).

Environment
-----------
``DATABASE_URL`` — required for dedup and for ``--apply``. If unset, the script
parses the workbook only (no DB checks / no writes).

Examples
--------
  PYTHONPATH=acdb-api python3 scripts/ops/backfill_operator_paid_customers_xlsx.py \\
      --xlsx "docs/MAS, MAK & SEH Paid Customers 13-05-2026.xlsx" --dry-run

  DATABASE_URL=postgresql://... PYTHONPATH=acdb-api python3 \\
      scripts/ops/backfill_operator_paid_customers_xlsx.py \\
      --xlsx path/to/paid_customers.xlsx --apply --require-payment-reference \
      --report-csv /tmp/op-fee.csv

Fee debt alignment only (no new ``transactions`` rows) lives in
``backfill_operator_fee_debt_align.py``.
"""

from __future__ import annotations

import argparse
import csv
import logging
import os
import re
import sys
from collections import Counter
from dataclasses import dataclass
from datetime import datetime, timedelta, timezone
from pathlib import Path
from typing import Any, Iterator, Optional

ROOT = Path(__file__).resolve().parents[2]
ACDB_API = ROOT / "acdb-api"
if str(ACDB_API) not in sys.path:
    sys.path.insert(0, str(ACDB_API))
OPS = ROOT / "scripts" / "ops"
if str(OPS) not in sys.path:
    sys.path.insert(0, str(OPS))

from inspect_operator_paid_customers_xlsx import (  # noqa: E402
    _col_map,
    _find_header_row,
)

LOG_FMT = "%(asctime)s [%(levelname)s] %(message)s"
logging.basicConfig(format=LOG_FMT, level=logging.INFO)
logger = logging.getLogger("cc-ops.operator-paid-xlsx")

VERIFIED_BY = "operator_paid_customers_xlsx"
SOURCE = "portal"
_AMOUNT_EPSILON = 0.005


@dataclass
class OperatorFeeSlot:
    account: str
    fee_type: str  # connection_fee | readyboard_fee
    amount: float
    paid_at: datetime | None
    reference: str | None
    workbook_row: int


def _norm_ref(ref: object) -> str:
    return re.sub(r"\s+", "", str(ref or "").strip())


def _parse_amount(val: object) -> float | None:
    if val is None or val == "":
        return None
    if isinstance(val, (int, float)):
        return float(val)
    s = str(val).strip().replace(",", "")
    if not s:
        return None
    try:
        return float(s)
    except ValueError:
        return None


def _parse_datetime(val: object) -> datetime | None:
    if val is None or val == "":
        return None
    if isinstance(val, datetime):
        dt = val
    else:
        try:
            from openpyxl.utils.datetime import from_excel

            if isinstance(val, (int, float)):
                dt = from_excel(val)
            else:
                s = str(val).strip()
                if not s:
                    return None
                dt = datetime.fromisoformat(s.replace("Z", "+00:00"))
        except Exception:
            return None
    if dt.tzinfo is None:
        dt = dt.replace(tzinfo=timezone.utc)
    else:
        dt = dt.astimezone(timezone.utc)
    return dt


def _normalize_account(raw: object) -> str:
    if raw is None:
        return ""
    if isinstance(raw, (int, float)) and not isinstance(raw, bool):
        s = str(int(raw)) if float(raw).is_integer() else str(raw)
    else:
        s = str(raw).strip()
    return s.upper()


def iter_fee_slots_from_workbook(
    path: Path,
    *,
    sheet: str | None = None,
    max_rows: int = 0,
) -> Iterator[OperatorFeeSlot]:
    import openpyxl

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb[sheet] if sheet else wb[wb.sheetnames[0]]
        hr = _find_header_row(ws)
        if hr is None:
            raise SystemExit(
                "Could not locate header row (need Customer ID + Connection Fee)."
            )
        header = [ws.cell(hr, c).value for c in range(1, ws.max_column + 1)]
        cmap = _col_map(header)
        need = ("account", "cf_amt")
        missing = [k for k in need if k not in cmap]
        if missing:
            raise SystemExit(f"Missing required columns: {missing}")

        n = 0
        for r in range(hr + 1, ws.max_row + 1):
            if max_rows and n >= max_rows:
                break
            acct_raw = ws.cell(r, cmap["account"]).value
            account = _normalize_account(acct_raw)
            if not account:
                continue
            n += 1

            cf_amt = _parse_amount(ws.cell(r, cmap["cf_amt"]).value)
            cf_date = _parse_datetime(ws.cell(r, cmap.get("cf_date", 0)).value) if "cf_date" in cmap else None
            cf_txn = _norm_ref(ws.cell(r, cmap.get("cf_txn", 0)).value) if "cf_txn" in cmap else None
            if cf_amt is not None and cf_amt > _AMOUNT_EPSILON:
                yield OperatorFeeSlot(
                    account=account,
                    fee_type="connection_fee",
                    amount=cf_amt,
                    paid_at=cf_date,
                    reference=cf_txn or None,
                    workbook_row=r,
                )

            if "rb_amt" in cmap:
                rb_amt = _parse_amount(ws.cell(r, cmap["rb_amt"]).value)
                rb_date = _parse_datetime(ws.cell(r, cmap.get("rb_date", 0)).value) if "rb_date" in cmap else None
                rb_txn = _norm_ref(ws.cell(r, cmap.get("rb_txn", 0)).value) if "rb_txn" in cmap else None
                if rb_amt is not None and rb_amt > _AMOUNT_EPSILON:
                    yield OperatorFeeSlot(
                        account=account,
                        fee_type="readyboard_fee",
                        amount=rb_amt,
                        paid_at=rb_date,
                        reference=rb_txn or None,
                        workbook_row=r,
                    )
    finally:
        wb.close()


def _account_exists(conn, account_number: str) -> bool:
    cur = conn.cursor()
    cur.execute(
        "SELECT 1 FROM accounts WHERE account_number = %s LIMIT 1",
        (account_number,),
    )
    return cur.fetchone() is not None


def _has_verified_fee(conn, account_number: str, payment_type: str) -> bool:
    cur = conn.cursor()
    try:
        cur.execute(
            """
            SELECT 1 FROM payment_verifications
            WHERE account_number = %s
              AND payment_type = %s
              AND status = 'verified'
            LIMIT 1
            """,
            (account_number, payment_type),
        )
        return cur.fetchone() is not None
    except Exception as exc:
        logger.warning("Fee verification lookup failed for %s: %s", account_number, exc)
        return False


def _payment_ref_taken(conn, ref: str) -> tuple[int, str] | None:
    receipt = (ref or "").strip()
    if not receipt:
        return None
    cur = conn.cursor()
    cur.execute(
        """
        SELECT id, account_number FROM transactions
        WHERE lower(trim(payment_reference)) = lower(trim(%s))
          AND payment_reference IS NOT NULL AND trim(payment_reference) <> ''
        LIMIT 1
        """,
        (receipt,),
    )
    row = cur.fetchone()
    return (int(row[0]), str(row[1])) if row else None


def _conflict_for_reference(
    conn,
    receipt: str,
    account_number: str,
    amount: float,
) -> str | None:
    receipt = (receipt or "").strip()
    if not receipt:
        return None
    dup = _payment_ref_taken(conn, receipt)
    if not dup:
        return None
    txn_id, existing_account = dup
    cur = conn.cursor()
    cur.execute(
        "SELECT transaction_amount FROM transactions WHERE id = %s",
        (txn_id,),
    )
    row = cur.fetchone()
    existing_amount = float(row[0]) if row and row[0] is not None else None
    if existing_account != account_number or (
        existing_amount is not None and abs(existing_amount - amount) > 0.01
    ):
        return (
            f"reference {receipt!r} already on txn {txn_id} "
            f"acct={existing_account} amount={existing_amount}"
        )
    return None


def _ref_in_inbound_log(conn, receipt: str) -> bool:
    receipt = (receipt or "").strip()
    if not receipt:
        return False
    cur = conn.cursor()
    try:
        cur.execute(
            "SELECT 1 FROM sms_inbound_log WHERE receipt_key = %s LIMIT 1",
            (receipt,),
        )
        return cur.fetchone() is not None
    except Exception:
        return False


def _fuzzy_fee_duplicate(
    conn,
    account_number: str,
    fee_type: str,
    amount: float,
    paid_at: datetime | None,
) -> bool:
    if paid_at is None:
        return False
    window_start = paid_at - timedelta(hours=24)
    window_end = paid_at + timedelta(hours=24)
    cur = conn.cursor()
    cur.execute(
        """
        SELECT 1 FROM transactions
        WHERE account_number = %s
          AND is_payment = true
          AND payment_category = %s
          AND transaction_amount BETWEEN %s AND %s
          AND transaction_date BETWEEN %s AND %s
        LIMIT 1
        """,
        (
            account_number,
            fee_type,
            amount - 0.01,
            amount + 0.01,
            window_start,
            window_end,
        ),
    )
    return cur.fetchone() is not None


def _resolve_meter(conn, account_number: str) -> str:
    cur = conn.cursor()
    cur.execute(
        """
        SELECT meter_id FROM meters
        WHERE account_number = %s AND status = 'active'
        LIMIT 1
        """,
        (account_number,),
    )
    row = cur.fetchone()
    return row[0] if row else ""


def _source_table_tag(slot: OperatorFeeSlot) -> str:
    ref = (slot.reference or "noref")[:20]
    return f"opxlsx:{slot.fee_type[:2]}:{ref}:r{slot.workbook_row}"[:50]


def _create_verification_entry(
    conn,
    transaction_id: int,
    account_number: str,
    payment_type: str,
    amount: float,
) -> int:
    cur = conn.cursor()
    cur.execute(
        """
        INSERT INTO payment_verifications
            (transaction_id, account_number, payment_type, amount)
        VALUES (%s, %s, %s, %s)
        RETURNING id
        """,
        (transaction_id, account_number, payment_type, amount),
    )
    return int(cur.fetchone()[0])


def _auto_verify_fee(conn, txn_id: int, note: str) -> None:
    cur = conn.cursor()
    cur.execute(
        """
        UPDATE payment_verifications
        SET status = 'verified',
            verified_by = %s,
            verified_at = NOW(),
            note = COALESCE(%s, note)
        WHERE transaction_id = %s AND status = 'pending'
        """,
        (VERIFIED_BY, note, txn_id),
    )


def process_fee_slot(
    conn,
    slot: OperatorFeeSlot,
    *,
    apply: bool,
    xlsx_name: str,
    require_payment_reference: bool = False,
) -> dict[str, Any]:
    from balance_engine import record_fee_transaction
    from fee_debt import apply_fee_payment_category_to_debt

    if require_payment_reference and not _norm_ref(slot.reference or ""):
        return {
            "workbook_row": slot.workbook_row,
            "account_number": slot.account,
            "fee_type": slot.fee_type,
            "amount": slot.amount,
            "reference": slot.reference,
            "paid_at": slot.paid_at.isoformat() if slot.paid_at else None,
            "outcome": "skipped",
            "reason": "missing_payment_reference",
        }

    base: dict[str, Any] = {
        "workbook_row": slot.workbook_row,
        "account_number": slot.account,
        "fee_type": slot.fee_type,
        "amount": slot.amount,
        "reference": slot.reference,
        "paid_at": slot.paid_at.isoformat() if slot.paid_at else None,
    }

    if not _account_exists(conn, slot.account):
        return {**base, "outcome": "skipped", "reason": "unknown_account"}

    if _has_verified_fee(conn, slot.account, slot.fee_type):
        return {**base, "outcome": "skipped", "reason": "already_verified_fee_type"}

    if slot.reference:
        conflict = _conflict_for_reference(conn, slot.reference, slot.account, slot.amount)
        if conflict:
            return {**base, "outcome": "conflict", "reason": conflict}
        if _payment_ref_taken(conn, slot.reference):
            return {**base, "outcome": "skipped", "reason": "duplicate_payment_reference"}
        if _ref_in_inbound_log(conn, slot.reference):
            return {**base, "outcome": "skipped", "reason": "sms_inbound_log_receipt"}

    if slot.paid_at and _fuzzy_fee_duplicate(conn, slot.account, slot.fee_type, slot.amount, slot.paid_at):
        return {**base, "outcome": "skipped", "reason": "fuzzy_duplicate_fee_slot"}

    source_table = _source_table_tag(slot)
    note = f"operator paid customers xlsx {xlsx_name} row {slot.workbook_row}"
    ts = slot.paid_at or datetime.now(timezone.utc)
    meter_id = _resolve_meter(conn, slot.account)

    summary = (
        f"would_insert fee_txn source_table={source_table!r} "
        f"category={slot.fee_type} amount={slot.amount:.2f} ref={slot.reference!r}"
    )
    if not apply:
        return {**base, "outcome": "would_insert", "would_insert_summary": summary}

    txn_id, _bal = record_fee_transaction(
        conn,
        slot.account,
        meter_id,
        amount_currency=slot.amount,
        payment_category=slot.fee_type,
        source=SOURCE,
        timestamp=ts,
        payment_reference=slot.reference or None,
        extra_columns={
            "source_table": source_table,
        },
    )
    _create_verification_entry(conn, txn_id, slot.account, slot.fee_type, slot.amount)
    _auto_verify_fee(conn, txn_id, note)
    apply_fee_payment_category_to_debt(conn, slot.account, slot.fee_type, slot.amount)
    try:
        from onboarding_derive import derive_payment_steps_for_accounts

        derive_payment_steps_for_accounts(conn, [slot.account])
    except Exception as exc:
        logger.warning("Onboarding derive failed for %s: %s", slot.account, exc)
    try:
        from onboarding_fee_trace import clear_listed_missing_if_fee_verified

        clear_listed_missing_if_fee_verified(conn, [slot.account])
    except Exception as exc:
        logger.warning("Fee trace clear failed for %s: %s", slot.account, exc)

    return {
        **base,
        "outcome": "inserted",
        "transaction_id": txn_id,
        "would_insert_summary": summary.replace("would_insert ", ""),
    }


def write_report(path: Path, rows: list[dict[str, Any]]) -> None:
    if not rows:
        path.write_text("outcome\n", encoding="utf-8")
        return
    fieldnames = sorted({key for row in rows for key in row.keys()})
    with path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)


def print_summary(rows: list[dict[str, Any]]) -> None:
    counts = Counter(row.get("outcome", "unknown") for row in rows)
    logger.info("Summary: %s", dict(counts))


def main() -> int:
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument("--xlsx", type=Path, required=True, help="Path to operator workbook")
    ap.add_argument("--sheet", type=str, default=None, help="Worksheet name (default: first sheet)")
    ap.add_argument("--max-rows", type=int, default=0, help="Limit data rows scanned (0=all)")
    ap.add_argument("--report-csv", type=Path, default=None)
    g = ap.add_mutually_exclusive_group()
    g.add_argument(
        "--dry-run",
        action="store_false",
        dest="apply",
        help="Parse + dedup only (default)",
    )
    g.add_argument("--apply", action="store_true", dest="apply", help="Write to 1PDB")
    ap.set_defaults(apply=False)
    ap.add_argument(
        "--require-payment-reference",
        action="store_true",
        help=(
            "Only process fee slots whose workbook Transaction ID_CF / ID_RB "
            "(trimmed) is non-empty for that row. Dry-run and --apply honor this. "
            "Logs a gate summary (skipped vs passed)."
        ),
    )
    args = ap.parse_args()
    apply = bool(args.apply)

    if not args.xlsx.is_file():
        logger.error("File not found: %s", args.xlsx)
        return 1

    slots = list(
        iter_fee_slots_from_workbook(args.xlsx, sheet=args.sheet, max_rows=args.max_rows)
    )
    logger.info("Parsed %d fee slot(s) from %s", len(slots), args.xlsx)

    database_url = os.environ.get("DATABASE_URL", "").strip()
    rows: list[dict[str, Any]] = []

    if not database_url:
        logger.warning("DATABASE_URL not set; skipping DB dedup and apply")
        for slot in slots:
            if args.require_payment_reference and not _norm_ref(slot.reference or ""):
                rows.append(
                    {
                        "workbook_row": slot.workbook_row,
                        "account_number": slot.account,
                        "fee_type": slot.fee_type,
                        "amount": slot.amount,
                        "reference": slot.reference,
                        "paid_at": slot.paid_at.isoformat() if slot.paid_at else None,
                        "outcome": "skipped",
                        "reason": "missing_payment_reference",
                    }
                )
            else:
                rows.append(
                    {
                        "workbook_row": slot.workbook_row,
                        "account_number": slot.account,
                        "fee_type": slot.fee_type,
                        "amount": slot.amount,
                        "reference": slot.reference,
                        "paid_at": slot.paid_at.isoformat() if slot.paid_at else None,
                        "outcome": "parsed_only",
                        "reason": "no_database_url",
                    }
                )
        if args.require_payment_reference:
            miss = sum(1 for r in rows if r.get("reason") == "missing_payment_reference")
            passed = len(slots) - miss
            logger.info(
                "Payment-reference gate (no DB): %d skipped (missing CF/RB txn ref), %d passed gate (of %d slots)",
                miss,
                passed,
                len(slots),
            )
        print_summary(rows)
        if args.report_csv:
            write_report(args.report_csv, rows)
        return 0

    import psycopg2

    conn = psycopg2.connect(database_url)
    conn.autocommit = False
    xlsx_name = args.xlsx.name
    try:
        for slot in slots:
            try:
                row = process_fee_slot(conn, slot, apply=apply, xlsx_name=xlsx_name, require_payment_reference=args.require_payment_reference)
            except Exception as exc:
                row = {
                    "workbook_row": slot.workbook_row,
                    "account_number": slot.account,
                    "fee_type": slot.fee_type,
                    "amount": slot.amount,
                    "reference": slot.reference,
                    "outcome": "error",
                    "reason": str(exc),
                }
                logger.exception("Row %s: %s", slot.workbook_row, exc)
            rows.append(row)
            ref_s = slot.reference or ""
            logger.info(
                "row=%s acct=%s type=%s amt=%.2f ref=%r -> %s%s",
                slot.workbook_row,
                slot.account,
                slot.fee_type,
                slot.amount,
                ref_s[:24] if ref_s else None,
                row.get("outcome"),
                f" ({row.get('reason')})" if row.get("reason") else "",
            )

        if apply:
            conn.commit()
            logger.info("Committed operator paid-customers backfill")
        else:
            conn.rollback()
        if args.require_payment_reference:
            miss = sum(1 for r in rows if r.get("reason") == "missing_payment_reference")
            passed = len(slots) - miss
            logger.info(
                "Payment-reference gate: %d skipped (missing CF/RB txn ref), %d passed gate (of %d slots)",
                miss,
                passed,
                len(slots),
            )
        print_summary(rows)
        if args.report_csv:
            write_report(args.report_csv, rows)
        return 0
    except Exception:
        conn.rollback()
        raise
    finally:
        conn.close()


if __name__ == "__main__":
    raise SystemExit(main())
